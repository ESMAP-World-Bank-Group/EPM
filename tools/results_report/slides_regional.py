# -*- coding: utf-8 -*-
"""Regional slide charts: maps, country mix, and the two corridor stories.

Same conventions as slide_dispatch.py / slide_seasonal.py -- the deck has to
look like one document, so palette, ink colour, legend column and 300 dpi
output all come from there.

The report cache only carries baseline and LC_Iso, so every other scenario is
read straight from its output_csv.  Two traps the readers below work around:

  * the Georgia-Romania submarine link is an *external* corridor, and external
    corridors carry neither TransmissionCapacity nor InterconUtilization nor
    CongestionShare in the outputs.  Its NTC is read back from the
    pExtTransferLimit file the scenario used, and utilisation is derived.
  * LC_FreeExpAll's pTransmissionMerged has no InterchangeExternal* rows at
    all, although its energy balance shows external trade.  Nothing can be
    drawn for its external corridors; only internal ones are.

    python slides_regional.py --chart all
"""

import argparse
import csv
import json
import math
from functools import lru_cache
import sys
from pathlib import Path

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
from matplotlib.colors import to_rgba
from matplotlib.lines import Line2D
from matplotlib.patches import Patch, Polygon, FancyArrowPatch
from matplotlib.ticker import MaxNLocator

HERE = Path(__file__).resolve().parent
sys.path.insert(0, str(HERE))
import extract as ex                                   # run layout + summary

RUN = ex.OUTVIEW / "simulations_run_20260825"
CACHE = HERE / "cache" / "simulations_run_20260825.json"
OUTDIR = HERE.parents[2] / "Data" / "results" / "slides"
YEARS = ex.YEARS

# --- house palette, copied from slide_dispatch.py so page and slide agree ----
COLORS = {
    "Nuclear": "#C8A8F0", "Coal": "#808890", "Gas": "#9A7040",
    "CCGT": "#B8921A", "OCGT": "#C4A820", "Diesel": "#6A7888",
    "HFO": "#7A7068", "Oil": "#7A7068", "Biomass": "#52C860",
    "Waste": "#8A9098", "Geothermal": "#D4A820", "Reservoir": "#1E9AF5",
    "ROR": "#5DADE2", "Hydro": "#1E9AF5", "PSH": "#0D7680",
    "Solar": "#FFD700", "PV": "#FFD700", "CSP": "#E8C547", "RPV": "#FFD700",
    "Onshore Wind": "#44DAEC", "Wind": "#44DAEC", "Offshore Wind": "#7CC8FA",
    "Battery": "#6A7BC8", "Storage": "#8290CE",
    "Imports": "#2E9EC8", "Exports": "#E8C547", "Storage Charge": "#0D7680",
    "Unmet demand": "#D9534F", "Demand": "#8B0000",
}
HATCHED = {"Imports", "Exports"}
INK = "#67788f"   # blue grey: titles and axis text, never black
SOFT = "#8a97a8"
GRID = "#e2e7ee"
LEGEND_IN = 1.25

COUNTRY = {"Armenia": "#1B6CA8", "Azerbaijan": "#36B5B5",
           "Georgia": "#E8C547", "Turkiye": "#6A7BC8"}
COUNTRIES = ["Armenia", "Azerbaijan", "Georgia", "Turkiye"]

# Map ink: the external ground stays behind the modelled zones.
EXT_FILL, EXT_EDGE = "#e8ecf3", "#d3dae5"
ZONE_FILL, ZONE_EDGE = "#f5f8fc", "#1a2333"
FLOW_HOT, FLOW_WARM, FLOW_COOL = "#c0392b", "#c0682a", "#1b6ca8"

SCEN_COLOR = {"LC_BSSC": "#1B6CA8", "LC_BSSC_VeryLow": "#7CC8FA",
              "LC_BSSC_Crisis": "#C0392B"}
SCEN_LABEL = {"LC_BSSC": "EU central", "LC_BSSC_VeryLow": "EU low",
              "LC_BSSC_Crisis": "EU crisis"}


def face(k, hatched):
    """Fill kwargs.  Matplotlib draws hatch lines in the edge colour, so a
    hatched band needs a pale face and a solid edge or the pattern vanishes."""
    c = COLORS.get(k, "#9aa5b4")
    if hatched:
        return dict(facecolor=to_rgba(c, .28), edgecolor=c, linewidth=.4,
                    hatch="///")
    return dict(facecolor=to_rgba(c, .88), edgecolor=to_rgba(c, .88),
                linewidth=.2)


def rc(fs):
    plt.rcParams.update({
        "font.family": "DejaVu Sans", "font.size": fs,
        "axes.edgecolor": "#ccd4de", "axes.linewidth": .6,
        "text.color": INK, "axes.labelcolor": SOFT,
        "xtick.color": SOFT, "ytick.color": SOFT,
        "xtick.major.size": 0, "ytick.major.size": 2,
        "ytick.major.width": .6, "ytick.major.pad": 2,
        "xtick.major.pad": 2, "xtick.labelsize": fs, "ytick.labelsize": fs,
    })


def save(fig, a, name):
    out = Path(a.out) if a.out else OUTDIR / name
    out.parent.mkdir(parents=True, exist_ok=True)
    fig.savefig(out, dpi=a.dpi, facecolor="white")
    print("%s  %.2f x %.2f in" % (out, a.width, a.height))
    return out


# ------------------------------------------------------------------- readers

def cache():
    return json.loads(CACHE.read_text(encoding="utf-8"))


def _tm(scen):
    import pandas as pd
    d = pd.read_csv(RUN / scen / "output_csv" / "pTransmissionMerged.csv")
    d["y"] = d["y"].astype(str)
    return d[d["y"].isin(YEARS)]


def corridors(scen):
    """Cache-shaped corridor block for any scenario, built from output_csv.

    extract.build_corridors leans on the run-level summary.csv, which has no
    column for the LC_FreeExpAll family; this reads the per-scenario file
    instead so every scenario in the run can be drawn.  `cong` is added on top
    of the cache shape: it exists for internal pairs only.
    """
    import pandas as pd
    tm = _tm(scen)
    cor = {}

    def slot(a, b, external=False):
        lo, hi = sorted([a, b])
        key = "%s|%s" % (lo, hi)
        if key not in cor:
            cor[key] = {"a": lo, "b": hi, "external": external,
                        "ntc": [0.0] * 16, "fwd": [0.0] * 16,
                        "rev": [0.0] * 16, "util": [0.0] * 16,
                        "cong": [0.0] * 16}
        return cor[key]

    # InterconUtilization and CongestionShare are *directional*: the model
    # reports z->z2 energy over NTC x 8760 h, and the share of hours the z->z2
    # flow sits on its limit.  Both denominators are the whole year, and a line
    # cannot be at its limit in both directions at once, so the average
    # utilisation of a line is the sum of its two rows, not the larger of them.
    # Taking the max threw away the minority direction: Armenia-Georgia 2035
    # read 95.9 % instead of 98.7 %, TRIPP 1 in 2035 read 35 % instead of 49 %.
    # Capacity is the same line seen twice, so it stays a max.
    for attr, field in (("InterconUtilization", "util"),
                        ("CongestionShare", "cong")):
        for _, r in tm[tm["attribute"] == attr].iterrows():
            c = slot(r["z"], r["uni"])
            c[field][YEARS.index(r["y"])] += float(r["value"] or 0)

    for _, r in tm[tm["attribute"] == "TransmissionCapacity"].iterrows():
        c = slot(r["z"], r["uni"])
        i = YEARS.index(r["y"])
        c["ntc"][i] = max(c["ntc"][i], float(r["value"] or 0))

    for _, r in tm[tm["attribute"] == "Interchange"].iterrows():
        c = slot(r["z"], r["uni"])                     # z is the source zone
        i = YEARS.index(r["y"])
        c["fwd" if r["z"] == c["a"] else "rev"][i] += float(r["value"] or 0) / 1e3

    for attr, into_z in (("InterchangeExternalImports", True),
                         ("InterchangeExternalExports", False)):
        for _, r in tm[tm["attribute"] == attr].iterrows():
            c = slot(r["z"], r["uni"], external=True)
            c["external"] = True
            i = YEARS.index(r["y"])
            fwd = (r["uni"] == c["a"]) if into_z else (r["z"] == c["a"])
            c["fwd" if fwd else "rev"][i] += float(r["value"] or 0) / 1e3

    f = ex.input_file_for(RUN, scen, "pExtTransferLimit")
    if f and f.exists():
        lim = pd.read_csv(f)
        lim.columns = [str(c).strip() for c in lim.columns]
        for (z, zext), g in lim.groupby(["z", "zext"]):
            c = slot(z, zext, external=True)
            c["external"] = True
            for i, y in enumerate(YEARS):
                if y in g.columns:
                    c["ntc"][i] = float(pd.to_numeric(g[y], errors="coerce").max() or 0)

    # isCongested fires whenever |flow - limit| < 1e-5, which is always true
    # on a line whose limit is still zero: those corridors come back as 100 %
    # congested in each direction.  A line that does not exist is not congested.
    for c in cor.values():
        for i in range(16):
            if c["ntc"][i] <= 0:
                c["cong"][i] = 0.0

    # External links report no utilisation, so it is the load factor of the
    # energy actually moved against what the NTC allows over 8760 h.
    for c in cor.values():
        if not c["external"]:
            continue
        for i in range(16):
            if c["ntc"][i] > 0:
                c["util"][i] = (c["fwd"][i] + c["rev"][i]) * 1e6 / (c["ntc"][i] * 8760.0)
    return cor


def hourly_price(scen, zone, year, hours):
    """pHours-weighted mean price, the only average that means anything here."""
    import pandas as pd
    hp = pd.read_csv(RUN / scen / "output_csv" / "pHourlyPrice.csv")
    hp["y"] = hp["y"].astype(str)
    hp = hp[(hp["z"] == zone) & (hp["y"] == str(year))]
    w = [hours.get("%s|%s|%s" % (q, d, t), 0.0)
         for q, d, t in zip(hp["q"], hp["d"], hp["t"])]
    tot = sum(w)
    return sum(v * wi for v, wi in zip(hp["value"], w)) / tot if tot else float("nan")


def trade_price(fname, zext, year, hours):
    """pHours-weighted mean of an exogenous external trade price file."""
    import pandas as pd
    df = pd.read_csv(ex.DATA / "trade" / fname)
    df["year"] = df["year"].astype(str)
    df = df[(df["zext"] == zext) & (df["year"] == str(year))]
    num = den = 0.0
    for _, r in df.iterrows():
        for h in ex.HOURS:
            w = hours.get("%s|%s|%s" % (r["q"], r["d"], h), 0.0)
            num += w * float(r[h])
            den += w
    return num / den if den else float("nan")


# ----------------------------------------------------------------- map layer

def flow_color(u):
    return FLOW_HOT if u >= .85 else (FLOW_WARM if u >= .6 else FLOW_COOL)


def draw_base(ax, geo, fs):
    for rings in geo["ext"].values():
        for ring in rings:
            ax.add_patch(Polygon(ring, closed=True, facecolor=EXT_FILL,
                                 edgecolor=EXT_EDGE, linewidth=.4, zorder=1))
    for rings in geo["zones"].values():
        for ring in rings:
            ax.add_patch(Polygon(ring, closed=True, facecolor=ZONE_FILL,
                                 edgecolor=ZONE_EDGE, linewidth=.45, zorder=2))
    x0, y0, x1, y1 = geo["box"]
    for name in geo["ext"]:
        lon, lat = geo["centroids"][name]
        if x0 < lon < x1 and y0 < lat < y1:
            ax.text(lon, lat, name, fontsize=fs - 1.2, color="#8d97a6",
                    ha="center", va="center", zorder=3, clip_on=True)
    ax.set_xlim(x0, x1)
    ax.set_ylim(y0, y1)
    # Degrees of longitude shrink with latitude; without this the region reads
    # as a squashed band rather than a map.
    ax.set_aspect(1.0 / math.cos(math.radians((y0 + y1) / 2)))
    ax.set_xticks([])
    ax.set_yticks([])
    for s in ax.spines.values():
        s.set_visible(False)


def _tbox(text, span, fs):
    """Half-width and half-height of a label, in degrees."""
    return (.0042 * span * len(text) * (fs / 6.) + .10,
            .19 * (fs / 6.))


def label_zones(ax, geo, fs, links=(), extra=(), pad=1.0):
    """Model zone names, nudged off the corridors they would otherwise sit on.

    Turkiye is nine zones and the maps are unreadable without the names, but a
    centroid is exactly where the arrows meet.  Each label is scored on a small
    ring of candidate slots: the score is how far the nearest obstacle sits
    outside the box the text occupies, measured in units of the box itself, so
    a value of 1 means just touching.  The offsets stay short (about a degree)
    because a zone name that drifts too far stops reading as that zone's."""
    x0, y0 = ax.get_xlim()[0], ax.get_ylim()[0]
    x1, y1 = ax.get_xlim()[1], ax.get_ylim()[1]
    span = x1 - x0

    # Sample the links: point-to-segment distance on a rotated box is fiddly
    # and a 0.12 degree step is finer than any label is small.
    obstacles = []
    for pa, pb in links:
        n = max(2, int(math.hypot(pb[0] - pa[0], pb[1] - pa[1]) / .12))
        for k in range(n + 1):
            f = k / float(n)
            obstacles.append((pa[0] + f * (pb[0] - pa[0]),
                              pa[1] + f * (pb[1] - pa[1]), .0, .0))
    for name in geo["ext"]:
        lon, lat = geo["centroids"][name]
        if x0 < lon < x1 and y0 < lat < y1:
            obstacles.append((lon, lat) + _tbox(name, span, fs))

    obstacles.extend(extra)

    def score(px, py, hw, hh):
        worst = 99.
        for ox, oy, ohw, ohh in obstacles:
            v = max(abs(px - ox) / (hw + ohw), abs(py - oy) / (hh + ohh))
            if v < worst:
                worst = v
        return worst

    order = sorted(geo["zones"], key=lambda n: -len(n))
    for name in order:
        lon, lat = geo["centroids"][name]
        if not (x0 < lon < x1 and y0 < lat < y1):
            continue
        hw, hh = _tbox(name, span, fs)
        best, best_s = (lon, lat), -99.
        for off in (0., .30, .55, .85, 1.15, 1.5):
            for ang in range(0, 360, 30):
                dx = off * math.cos(math.radians(ang))
                dy = off * math.sin(math.radians(ang)) * .62
                px, py = lon + dx, lat + dy
                if not (x0 + hw < px < x1 - hw and y0 + hh < py < y1 - hh):
                    continue
                s = score(px, py, hw, hh) - .22 * off      # stay home if you can
                if s > best_s:
                    best, best_s = (px, py), s
                if off == 0.:
                    break                                  # one centre, not 12
            if best_s > pad + .25:
                break
        obstacles.append((best[0], best[1], hw, hh))
        ax.text(best[0], best[1], name, fontsize=fs - 1.5, color="#5b6a7e",
                ha="center", va="center", zorder=7, clip_on=True,
                bbox=dict(boxstyle="round,pad=.10", fc="white", ec="none",
                          alpha=.70))


def fit_extent(ax, geo, w_in, h_in):
    """Crop the frame to the shape of the panel.

    The map keeps a true aspect ratio, so a wide short panel would otherwise
    show a correctly proportioned map hugging one side and half the picture
    empty.  Trim the frame instead, centred on the modelled zones rather than
    on the whole box so nothing that matters falls off the top."""
    x0, y0, x1, y1 = geo["box"]
    k = math.cos(math.radians((y0 + y1) / 2))
    lats = [pt[1] for rings in geo["zones"].values() for r in rings for pt in r]
    lons = [pt[0] for rings in geo["zones"].values() for r in rings for pt in r]
    cy = (min(lats) + max(lats)) / 2 if lats else (y0 + y1) / 2
    cx = (min(lons) + max(lons)) / 2 if lons else (x0 + x1) / 2
    want = (h_in / w_in) * (x1 - x0) * k
    if want < y1 - y0:
        lo = min(max(y0, cy - want / 2), y1 - want)
        ax.set_ylim(lo, lo + want)
    else:
        wl = (w_in / h_in) * (y1 - y0) / k
        lo = min(max(x0, cx - wl / 2), x1 - wl)
        ax.set_xlim(lo, lo + wl)


def draw_flows(ax, geo, cor, i, fs, skip_ext=False, scale=1.0):
    """One arrow per live corridor, pointing the way the net energy goes.

    Line widths and arrowheads are in points, so on the small-multiple panels
    they have to follow the panel width or the heads swallow the map."""
    live = []
    for key, c in cor.items():
        if skip_ext and c["external"]:
            continue
        gross = c["fwd"][i] + c["rev"][i]
        if gross <= 1e-3 and c["ntc"][i] <= 0:
            continue                                   # idle: not even a link
        if c["a"] not in geo["centroids"] or c["b"] not in geo["centroids"]:
            continue
        live.append((key, c, gross))
    if not live:
        return live
    top = max(g for _, _, g in live) or 1.0
    for key, c, gross in live:
        pa = geo["centroids"][c["a"]]
        pb = geo["centroids"][c["b"]]
        # fwd runs a -> b, so a net that runs the other way turns the head round.
        if c["fwd"][i] < c["rev"][i]:
            pa, pb = pb, pa
        lw = .45 + 2.5 * math.sqrt(max(gross, 0) / top)
        ax.add_patch(FancyArrowPatch(
            pa, pb, arrowstyle="-|>", mutation_scale=(3.2 + 2.0 * lw) * scale,
            linewidth=lw * scale, color=flow_color(c["util"][i]), alpha=.85,
            shrinkA=3 * scale, shrinkB=3 * scale, zorder=4))
    return live


def label_countries(ax, geo, zcmap, fs):
    pts = {}
    for z, c in zcmap.items():
        if c in COUNTRIES and z in geo["centroids"]:
            pts.setdefault(c, []).append(geo["centroids"][z])
    pos = {c: (sum(p[0] for p in ps) / len(ps),
               sum(p[1] for p in ps) / len(ps)) for c, ps in pts.items()}
    # Armenia and the Azerbaijani zones share a corner and their two names land
    # on top of each other.  Collision is judged on the box the text actually
    # occupies, not on the distance between anchors, then the pair is pushed
    # apart in latitude.
    x0, y0, x1, y1 = geo["box"]
    halfw = {c: .006 * (x1 - x0) * len(c) * (fs / 6.0) for c in pos}
    vgap = .030 * (y1 - y0) * (fs / 6.0)
    step = .020 * (y1 - y0)
    names = sorted(pos)
    for _ in range(14):
        moved = False
        for i, a in enumerate(names):
            for b in names[i + 1:]:
                if abs(pos[a][0] - pos[b][0]) > halfw[a] + halfw[b]:
                    continue
                if abs(pos[a][1] - pos[b][1]) > vgap:
                    continue
                up, dn = (a, b) if pos[a][1] >= pos[b][1] else (b, a)
                pos[up] = (pos[up][0], pos[up][1] + step)
                pos[dn] = (pos[dn][0], pos[dn][1] - step)
                moved = True
        if not moved:
            break
    for c, (lon, lat) in pos.items():
        ax.text(lon, lat, c, fontsize=fs - .8, color=INK, fontweight="bold",
                ha="center", va="center", zorder=6,
                bbox=dict(boxstyle="round,pad=.12", fc="white", ec="none",
                          alpha=.72))


def flow_legend(fig, right, fs, extra=None):
    h = [Line2D([], [], color=FLOW_COOL, lw=1.6, label="< 60% used"),
         Line2D([], [], color=FLOW_WARM, lw=1.6, label="60-85%"),
         Line2D([], [], color=FLOW_HOT, lw=1.6, label=">= 85%")]
    if extra:
        h += extra
    fig.legend(handles=h, loc="center left", ncol=1, fontsize=fs - .5,
               frameon=False, handlelength=1.1, handletextpad=.45,
               labelspacing=.5, borderpad=0, bbox_to_anchor=(right + .012, .5))


def map_right(a):
    return 1 - max(.86, min(LEGEND_IN, a.width * .18)) / a.width


# ------------------------------------------------------------------ chart 1

def chart_region_maps(a):
    d = cache()
    geo, cor = d["geo"], d["corridors"]["baseline"]
    years = [y.strip() for y in a.years.split(",")] if a.years else ["2025", "2030", "2035"]
    fs = 6.0
    rc(fs)
    fig, axes = plt.subplots(1, len(years), figsize=(a.width, a.height), dpi=a.dpi)
    axes = list(axes) if len(years) > 1 else [axes]
    panel = a.width * .96 / len(years)
    sc = min(1.0, panel / 4.4)
    fsp = max(4.2, fs * (.55 + .45 * sc))    # type follows the panel too
    for ax, y in zip(axes, years):
        i = YEARS.index(y)
        draw_base(ax, geo, fsp)
        fit_extent(ax, geo, panel, a.height * .80)
        draw_flows(ax, geo, cor, i, fs, scale=sc)
        ax.set_title(y, fontsize=fs + 1.5, fontweight="bold", loc="left",
                     pad=1.5, color=INK)
        ax.set_anchor("N")            # maps hug their year label, slack goes down
    if a.title:
        fig.suptitle(a.title, fontsize=fs + 1, fontweight="bold", color=INK,
                     x=.012, y=.995, ha="left", va="top")
    # Three panels side by side are width-starved, so the legend lies flat in
    # the strip the maps leave underneath rather than eating a fourth column.
    top = .93 if a.title else .995   # no headline: give the band back
    fig.tight_layout(pad=.2, w_pad=.3, rect=(0, .085, .995, top))
    h = [Line2D([], [], color=FLOW_COOL, lw=1.6, label="< 60% used"),
         Line2D([], [], color=FLOW_WARM, lw=1.6, label="60-85%"),
         Line2D([], [], color=FLOW_HOT, lw=1.6, label=">= 85%"),
         Line2D([], [], color=SOFT, lw=2.6, label="thicker = more energy")]
    fig.legend(handles=h, loc="lower center", ncol=4, fontsize=fs - .5,
               frameon=False, handlelength=1.3, handletextpad=.4,
               columnspacing=1.6, borderpad=0, bbox_to_anchor=(.5, -.005))
    return save(fig, a, "region_maps.png")


# ------------------------------------------------------------------ chart 2

def chart_region_generation(a):
    d = cache()
    years = [y.strip() for y in a.years.split(",")] if a.years else ["2025", "2030", "2035"]
    fs = 6.0
    rc(fs)
    fig, axes = plt.subplots(1, len(COUNTRIES), figsize=(a.width, a.height), dpi=a.dpi)

    used = []
    for ax, ctry in zip(axes, COUNTRIES):
        blk = d["annual"][ctry]["baseline"]
        xs = list(range(len(years)))
        for xi, y in zip(xs, years):
            i = YEARS.index(y)
            base = 0.0
            for f in d["fuel_order"]:
                v = blk["gen"].get(f)
                if not v or abs(v[i]) < 1e-3:
                    continue
                ax.bar(xi, v[i], bottom=base, width=.62, zorder=2, **face(f, False))
                base += v[i]
                if f not in used:
                    used.append(f)
            # The scope's trade block already drops flows internal to the
            # country's own zones, so this is the cross-border position only.
            imp = sum(t["imp"][i] for t in blk["trade"].values())
            exp = sum(t["exp"][i] for t in blk["trade"].values())
            if imp > 1e-3:
                ax.bar(xi, imp, bottom=base, width=.62, zorder=2,
                       **face("Imports", True))
                if "Imports" not in used:
                    used.append("Imports")
            if exp > 1e-3:
                ax.bar(xi, -exp, width=.62, zorder=2, **face("Exports", True))
                if "Exports" not in used:
                    used.append("Exports")
        ax.axhline(0, color="#8b96a5", linewidth=.6, zorder=3)
        ax.set_xticks(xs)
        ax.set_xticklabels(years, fontsize=fs - .5, color=SOFT)
        ax.set_xlim(xs[0] - .52, xs[-1] + .52)
        ax.set_title(ctry, fontsize=fs + .5, fontweight="bold", loc="left",
                     pad=2, color=INK)
        ax.yaxis.set_major_locator(MaxNLocator(4))
        ax.yaxis.grid(True, color=GRID, linewidth=.5, zorder=0)
        ax.set_axisbelow(True)
        for s in ("top", "right"):
            ax.spines[s].set_visible(False)
    axes[0].set_ylabel("TWh", fontsize=fs, labelpad=2)

    fig.suptitle(a.title or "Generation mix and cross-border trade, baseline",
                 fontsize=fs + 1, fontweight="bold", color=INK,
                 x=.012, y=.995, ha="left", va="top")
    # Turkiye is fifty times Armenia, so one shared axis would flatten the three
    # small systems into a line: each panel keeps its own scale, said plainly.
    fig.text(.012, .012, "y scales differ by country", fontsize=fs - 1.2,
             color=SOFT, ha="left", va="bottom")
    right = 1 - LEGEND_IN / a.width
    fig.tight_layout(pad=.3, w_pad=.6, rect=(.012, .045, right, .93))
    handles = [Patch(label=k, **face(k, k in HATCHED)) for k in used]
    lfs, lsp = fs, .42
    while len(handles) * lfs * (1 + lsp) > a.height * 72 * .92 and lfs > 4.2:
        lfs -= .2
        lsp = max(.16, lsp - .03)
    fig.legend(handles=handles, loc="center left", ncol=1, fontsize=lfs,
               frameon=False, handlelength=1.1, handleheight=.9,
               handletextpad=.45, labelspacing=lsp, borderpad=0,
               bbox_to_anchor=(right + .015, .48))
    return save(fig, a, "region_generation.png")


# ------------------------------------------------------------------ chart 3

def bssc_facts(scen, year):
    """Everything the callout prints, straight from the run."""
    d = cache()
    c = corridors(scen)["Georgia|Romania"]
    i = YEARS.index(str(year))
    out = {                                            # fwd is Georgia -> Romania
        "exp": c["fwd"][i], "imp": c["rev"][i],
        "ntc": c["ntc"][i], "util": c["util"][i], "cong": c["cong"][i],
        "first": next((YEARS[k] for k in range(16)
                       if c["fwd"][k] + c["rev"][k] > 1e-3), None),
        "cod": next((YEARS[k] for k in range(16) if c["ntc"][k] > 0), None),
        "ge_price": hourly_price(scen, "Georgia", year, d["hours"]),
    }
    imp_f = ex.input_file_for(RUN, scen, "pTradePrice")
    exp_f = ex.input_file_for(RUN, scen, "pTradePriceExport")
    out["ro_buy"] = trade_price(imp_f.name, "Romania", year, d["hours"]) if imp_f else None
    out["ro_sell"] = trade_price(exp_f.name, "Romania", year, d["hours"]) if exp_f else None
    return out


def chart_bssc_map(a):
    d = cache()
    geo = d["geo"]
    scen, year = a.scenario or "LC_BSSC", a.year or "2035"
    cor = corridors(scen)
    i = YEARS.index(str(year))
    f = bssc_facts(scen, year)
    fs = 6.0
    rc(fs)
    fig, ax = plt.subplots(figsize=(a.width, a.height), dpi=a.dpi)
    draw_base(ax, geo, fs)
    fit_extent(ax, geo, a.width * map_right(a) - .1, a.height - .1)
    live = draw_flows(ax, geo, cor, i, fs)
    label_zones(ax, geo, fs, [(geo["centroids"][c["a"]],
                              geo["centroids"][c["b"]]) for _, c, _ in live])
    ax.set_anchor("W")   # one panel: give the slack to the legend gutter

    # The callout sits over the Black Sea, beside the line it describes.  The
    # hours-at-the-limit line is honest about a gap: CongestionShare is only
    # written for internal pairs, never for an external link.
    ge = geo["centroids"]["Georgia"]
    # Bare dollar signs would open mathtext and swallow the text between them.
    txt = ("Georgia - Romania, %s\n"
           "%.1f TWh GE to RO   %.1f TWh RO to GE\n"
           "%.0f MW link, %.0f%% load factor\n"
           "hours at the limit: not reported for external links\n"
           "Georgia %.0f \\$/MWh   Romania %.0f \\$/MWh\n"
           "(EU central, exogenous export price, see table)"
           % (year, f["exp"], f["imp"], f["ntc"], 100 * f["util"],
              f["ge_price"], f["ro_sell"]))
    # Anchored in axes fractions, not degrees: the frame is cropped to the
    # panel shape, so a point picked on the map can end up outside it.
    ax.annotate(txt, xy=ge, xytext=(.995, .995), textcoords="axes fraction",
                fontsize=fs - 1.1, color=INK, ha="right", va="top", ma="left",
                zorder=8,
                bbox=dict(boxstyle="round,pad=.3", fc="white", ec="#c3ccd9",
                          linewidth=.5),
                arrowprops=dict(arrowstyle="-", color="#c3ccd9", linewidth=.5,
                                shrinkB=2))
    if a.title:
        fig.suptitle(a.title, fontsize=fs + 1, fontweight="bold", color=INK,
                     x=.012, y=.995, ha="left", va="top")
    right = map_right(a)
    top = .93 if a.title else .995   # no headline: give the band back
    fig.tight_layout(pad=.2, rect=(0, 0, right, top))
    flow_legend(fig, right, fs)
    return save(fig, a, "bssc_map.png")


# ------------------------------------------------------------------ chart 4

def chart_bssc_volume(a):
    scens = ["LC_BSSC", "LC_BSSC_VeryLow", "LC_BSSC_Crisis"]
    dat = {s: corridors(s)["Georgia|Romania"] for s in scens}
    # Every other year of the cable's life, anchored on the horizon: the drift
    # under the low EU price path is gradual, so a two-year step still shows it
    # while leaving room for the utilisation figure on every bar.
    live = [y for k, y in enumerate(YEARS)
            if any(dat[s]["ntc"][k] > 0 for s in scens)]
    years = ([y.strip() for y in a.years.split(",")] if a.years
             else live[(len(live) - 1) % 2::2])
    fs = 6.0
    rc(fs)
    fig, ax = plt.subplots(figsize=(a.width, a.height), dpi=a.dpi)

    bw = .78 / len(scens)
    for j, s in enumerate(scens):
        c, col = dat[s], SCEN_COLOR[s]
        xs = [k - .39 + bw * (j + .5) for k in range(len(years))]
        for x, y in zip(xs, years):
            i = YEARS.index(y)
            fwd, rev = c["fwd"][i], c["rev"][i]
            ax.bar(x, fwd, width=bw * .84, zorder=2,
                   facecolor=to_rgba(col, .88), edgecolor=to_rgba(col, .88),
                   linewidth=.2)
            ax.bar(x, rev, bottom=fwd, width=bw * .84, zorder=2,
                   facecolor=to_rgba(col, .22), edgecolor=col, linewidth=.45,
                   hatch="///")
            # A dot on the gross top reads better here than a tick across the
            # bar: at this width the tick disappeared into the bar edge.
            ax.plot([x], [fwd + rev], marker="o", markersize=2.6,
                    color="#2f3f57", zorder=5, linestyle="none")
            if len(years) < 6:
                ax.text(x, fwd + rev + .30, "%.0f%%" % (100 * c["util"][i]),
                        fontsize=fs - 1.0, color=INK, ha="center",
                        va="bottom", zorder=6)

    ax.set_xticks(range(len(years)))
    ax.set_xticklabels(years, fontsize=fs, color=INK, fontweight="bold")
    ax.set_ylabel("TWh on the link\nsolid = Georgia to Romania", fontsize=fs,
                  labelpad=2)
    ax.set_xlim(-.6, len(years) - .4)
    ax.set_ylim(0, ax.get_ylim()[1] * 1.10)
    ax.yaxis.set_major_locator(MaxNLocator(5))
    ax.yaxis.grid(True, color=GRID, linewidth=.5, zorder=0)
    ax.set_axisbelow(True)
    for s in ("top", "right"):
        ax.spines[s].set_visible(False)

    fig.suptitle(a.title or "Submarine cable volume against EU price levels",
                 fontsize=fs + 1, fontweight="bold", color=INK,
                 x=.012, y=.995, ha="left", va="top")
    right = 1 - LEGEND_IN / a.width
    fig.tight_layout(pad=.3, rect=(.012, 0, right, .92))
    h = [Patch(label=SCEN_LABEL[s], facecolor=to_rgba(SCEN_COLOR[s], .88),
               edgecolor=to_rgba(SCEN_COLOR[s], .88), linewidth=.2)
         for s in scens]
    h.append(Patch(label="Romania to\nGeorgia", facecolor="#ffffff",
                   edgecolor=SOFT, linewidth=.45, hatch="///"))
    h.append(Line2D([], [], color="#2f3f57", marker="o", markersize=3.0,
                    linestyle="none", label="load factor"))
    fig.legend(handles=h, loc="center left", ncol=1, fontsize=fs, frameon=False,
               handlelength=1.1, handleheight=.9, handletextpad=.45,
               labelspacing=.5, borderpad=0, bbox_to_anchor=(right + .015, .5))
    return save(fig, a, "bssc_volume_sensitivity.png")


# ------------------------------------------------------------------ chart 4b

@lru_cache(maxsize=None)
def annual_for(scen, ctry):
    """One country's annual block for any scenario in summary.csv.

    The cache only carries baseline and LC_Iso, so anything else is rebuilt
    from summary.csv with the same code the cache itself used."""
    d = cache()
    have = d["annual"][ctry]
    if scen in have:
        return have[scen]
    return ex.build_annual(ex.load_summary(RUN), scen, d["scopes"][ctry])


def chart_bssc_mix_delta(a):
    d = cache()
    scen = a.scenario or "LC_BSSC"
    years = ([y.strip() for y in a.years.split(",")] if a.years
             else ["2030", "2035", "2040"])
    fs = 7.0
    rc(fs)
    # One shared y axis across the four panels.  Per-panel scales make Armenia's
    # 0.3 TWh look like Georgia's 8, which is the opposite of what the chart is
    # for: the point is that the cable moves Georgia and Turkiye and barely
    # touches Armenia.
    fig, axes = plt.subplots(1, len(COUNTRIES), figsize=(a.width, a.height),
                             dpi=a.dpi, sharey=True)

    used = []
    for ax, ctry in zip(axes, COUNTRIES):
        base, alt = annual_for("baseline", ctry), annual_for(scen, ctry)
        # Tighter than unit spacing: the years of one country read as a group,
        # and the gaps between panels do the separating.
        xs = [k * .80 for k in range(len(years))]
        for xi, y in zip(xs, years):
            i = YEARS.index(y)
            up = dn = 0.0

            def put(key, v, hatched=False):
                """Gains stack above the zero line, losses below it."""
                nonlocal up, dn
                if abs(v) < 1e-3:
                    return
                ax.bar(xi, v, bottom=(up if v > 0 else dn), width=.62,
                       zorder=2, **face(key, hatched))
                if v > 0:
                    up += v
                else:
                    dn += v
                if key not in used:
                    used.append(key)

            for f in d["fuel_order"]:
                b, m = base["gen"].get(f), alt["gen"].get(f)
                if not b and not m:
                    continue
                put(f, (m[i] if m else 0.0) - (b[i] if b else 0.0))
            for key, side in (("Imports", "imp"), ("Exports", "exp")):
                b = sum(t[side][i] for t in base["trade"].values())
                m = sum(t[side][i] for t in alt["trade"].values())
                # Exports keep the sign convention of the mix chart: more sold
                # abroad reads downward, the same way it does in the baseline.
                put(key, (m - b) * (1 if key == "Imports" else -1), True)

            net = up + dn
            ax.plot([xi - .31, xi + .31], [net, net], color="#2f3f57",
                    linewidth=1.0, zorder=6, solid_capstyle="butt")

        ax.axhline(0, color="#8b96a5", linewidth=.6, zorder=3)
        ax.set_xticks(xs)
        # Four panels on a 5.9 in box: the year labels only clear each
        # other a good deal smaller than the panel titles.
        ax.set_xticklabels(years, fontsize=fs - 2.0, color=SOFT)
        ax.set_xlim(xs[0] - .52, xs[-1] + .52)
        ax.set_title(ctry, fontsize=fs + .5, fontweight="bold", loc="left",
                     pad=2, color=INK)
        ax.yaxis.set_major_locator(MaxNLocator(4))
        ax.yaxis.grid(True, color=GRID, linewidth=.5, zorder=0)
        ax.set_axisbelow(True)
        for sp in ("top", "right"):
            ax.spines[sp].set_visible(False)
    axes[0].set_ylabel("TWh vs baseline", fontsize=fs, labelpad=2)
    for ax in axes[1:]:
        ax.spines["left"].set_visible(False)
        ax.tick_params(axis="y", length=0)

    if a.title:
        fig.suptitle(a.title, fontsize=fs + 1, fontweight="bold", color=INK,
                     x=.012, y=.995, ha="left", va="top")
    right = 1 - LEGEND_IN / a.width
    top = .93 if a.title else .995
    fig.tight_layout(pad=.3, w_pad=.4, rect=(.012, .012, right, top))
    order = [f for f in d["fuel_order"] if f in used] +             [k for k in ("Imports", "Exports") if k in used]
    handles = [Patch(label=k, **face(k, k in HATCHED)) for k in order]
    handles.append(Line2D([], [], color="#2f3f57", lw=1.0, label="net"))
    lfs, lsp = fs, .42
    while len(handles) * lfs * (1 + lsp) > a.height * 72 * .92 and lfs > 4.2:
        lfs -= .2
        lsp = max(.16, lsp - .03)
    fig.legend(handles=handles, loc="center left", ncol=1, fontsize=lfs,
               frameon=False, handlelength=1.1, handleheight=.9,
               handletextpad=.45, labelspacing=lsp, borderpad=0,
               bbox_to_anchor=(right + .015, .48))
    return save(fig, a, "bssc_mix_delta.png")


# ------------------------------------------------------------------ chart 5

def country_generation(scen):
    """Annual generation in TWh per country, for a scenario in summary.csv."""
    summ = ex.load_summary(RUN)
    zc = ex.read_zcmap()
    g = summ[summ["attribute"] == "Energy: GWh"].copy()
    g["country"] = g["zone"].map(zc)
    t = g.groupby(["country", "year"])[scen].sum() / 1000.0
    return {c: {y: float(t.get((c, y), 0.0)) for y in YEARS} for c in COUNTRIES}


def chart_bssc_impact(a):
    years = [y.strip() for y in a.years.split(",")] if a.years else ["2030", "2035", "2040"]
    base = country_generation("baseline")
    bssc = country_generation(a.scenario or "LC_BSSC")
    fs = 6.0
    rc(fs)
    fig, ax = plt.subplots(figsize=(a.width, a.height), dpi=a.dpi)

    for xi, y in enumerate(years):
        up = dn = 0.0
        for c in COUNTRIES:
            v = bssc[c][y] - base[c][y]
            if abs(v) < 5e-3:
                continue
            ax.bar(xi, v, bottom=(up if v > 0 else dn), width=.5, zorder=2,
                   facecolor=to_rgba(COUNTRY[c], .88),
                   edgecolor=to_rgba(COUNTRY[c], .88), linewidth=.2)
            if v > 0:
                up += v
            else:
                dn += v
        net = sum(bssc[c][y] - base[c][y] for c in COUNTRIES)
        ax.plot([xi - .32, xi + .32], [net] * 2, color="#2f3f57", linewidth=1.2,
                solid_capstyle="butt", zorder=5)
        ax.text(xi + .35, net, "%+.1f" % net, fontsize=fs - .8, color=INK,
                ha="left", va="center", zorder=6)

    ax.axhline(0, color="#8b96a5", linewidth=.6, zorder=3)
    ax.set_xticks(range(len(years)))
    ax.set_xticklabels(years, fontsize=fs + .5, color=INK, fontweight="bold")
    ax.set_ylabel("TWh generated, vs baseline", fontsize=fs, labelpad=2)
    ax.set_xlim(-.55, len(years) - .25)
    lo, hi = ax.get_ylim()
    ax.set_ylim(lo * 1.18, hi * 1.18)
    ax.yaxis.set_major_locator(MaxNLocator(5))
    ax.yaxis.grid(True, color=GRID, linewidth=.5, zorder=0)
    ax.set_axisbelow(True)
    for s in ("top", "right"):
        ax.spines[s].set_visible(False)

    fig.suptitle(a.title or "Generation shift under LC_BSSC, against baseline",
                 fontsize=fs + 1, fontweight="bold", color=INK,
                 x=.012, y=.995, ha="left", va="top")
    right = 1 - LEGEND_IN / a.width
    fig.tight_layout(pad=.3, rect=(.012, 0, right, .92))
    h = [Patch(label=c, facecolor=to_rgba(COUNTRY[c], .88),
               edgecolor=to_rgba(COUNTRY[c], .88), linewidth=.2)
         for c in COUNTRIES]
    h.append(Line2D([], [], color="#2f3f57", lw=1.2, label="Region, net"))
    fig.legend(handles=h, loc="center left", ncol=1, fontsize=fs, frameon=False,
               handlelength=1.1, handleheight=.9, handletextpad=.45,
               labelspacing=.5, borderpad=0, bbox_to_anchor=(right + .015, .5))
    return save(fig, a, "bssc_impact.png")


# ------------------------------------------------------------------ chart 6

def capacity_growth(scen, y0="2025", y1="2040", floor=50.0):
    """Corridors whose NTC grows over the horizon, MW, largest first."""
    cor = corridors(scen)
    i0, i1 = YEARS.index(y0), YEARS.index(y1)
    out = [(k, c, c["ntc"][i1] - c["ntc"][i0])
           for k, c in cor.items() if not c["external"]]
    return sorted([r for r in out if r[2] >= floor], key=lambda r: -r[2])


def chart_freeexp_map(a):
    d = cache()
    geo = d["geo"]
    scen = a.scenario or "LC_FreeExpAll"
    year = a.year or "2040"
    cor = corridors(scen)
    i = YEARS.index(str(year))
    fs = 6.0
    rc(fs)
    fig, ax = plt.subplots(figsize=(a.width, a.height), dpi=a.dpi)
    draw_base(ax, geo, fs)
    fit_extent(ax, geo, a.width * map_right(a) - .1, a.height - .1)
    live = draw_flows(ax, geo, cor, i, fs)
    segs = [(geo["centroids"][c["a"]], geo["centroids"][c["b"]])
            for _, c, _ in live]
    boxes = []
    ax.set_anchor("W")   # one panel: give the slack to the legend gutter

    # The labels are the chart, and several of them land on the same crowded
    # corner (Nakhchivan, western Turkiye).  Each one is tried on both sides of
    # its own corridor at growing distance and keeps the first slot that clears
    # the labels already placed and the country names, with a leader line back
    # to the midpoint so the link it belongs to stays readable.
    x0, y0, x1, y1 = geo["box"]
    taken = []
    for z, c in d["zcmap"].items():
        if c in COUNTRIES and z in geo["centroids"]:
            taken.append(geo["centroids"][z])
    for name in geo["ext"]:                          # Iran, Russia, Iraq...
        pt = geo["centroids"].get(name)
        if pt and x0 < pt[0] < x1 and y0 < pt[1] < y1:
            taken.append(pt)
    grown = list(capacity_growth(scen, "2025", str(year)))
    for key, c, dmw in grown:
        pa, pb = geo["centroids"][c["a"]], geo["centroids"][c["b"]]
        mid = ((pa[0] + pb[0]) / 2, (pa[1] + pb[1]) / 2)
        dx, dy = pb[0] - pa[0], pb[1] - pa[1]
        n = math.hypot(dx, dy) or 1.0
        nx, ny = -dy / n, dx / n                      # unit normal to the link
        best, best_gap = None, -1.0
        for step in range(7):
            off = .7 + .7 * step
            for sign in (1, -1):
                pos = (mid[0] + nx * off * sign, mid[1] + ny * off * sign)
                if not (x0 + .4 < pos[0] < x1 - .4
                        and y0 + .3 < pos[1] < y1 - .3):
                    continue
                gap = min([math.hypot(pos[0] - q[0], (pos[1] - q[1]) * 1.6)
                           for q in taken] or [99])
                if gap > best_gap:
                    best, best_gap = pos, gap
                if gap > 2.2:
                    break
            if best_gap > 2.2:
                break
        pos = best or mid
        taken.append(pos)
        ax.annotate("+%.0f MW" % dmw, xy=mid, xytext=pos, textcoords="data",
                    fontsize=fs - 1.1, color="#12356e", fontweight="bold",
                    ha="center", va="center", zorder=8,
                    bbox=dict(boxstyle="round,pad=.16", fc="white", ec="#c3ccd9",
                              linewidth=.4, alpha=.96),
                    arrowprops=dict(arrowstyle="-", color="#7f8b9c",
                                    linewidth=.5, shrinkA=1, shrinkB=1))
        boxes.append(pos + _tbox("+%.0f MW" % dmw,
                                 ax.get_xlim()[1] - ax.get_xlim()[0], fs))

    label_zones(ax, geo, fs, segs, boxes)
    if a.title:
        fig.suptitle(a.title, fontsize=fs + 1, fontweight="bold", color=INK,
                     x=.012, y=.995, ha="left", va="top")
    right = map_right(a)
    top = .93 if a.title else .995   # no headline: give the band back
    fig.tight_layout(pad=.2, rect=(0, 0, right, top))
    flow_legend(fig, right, fs)
    return save(fig, a, "freeexp_map.png")


# ------------------------------------------------------------------ chart 7

def top_corridors(scen, year, n=10):
    cor = corridors(scen)
    i = YEARS.index(str(year))
    rows = [(k, c["fwd"][i] + c["rev"][i], c["util"][i]) for k, c in cor.items()]
    return sorted([r for r in rows if r[1] > 1e-3], key=lambda r: -r[1])[:n]


def chart_freeexp_corridors(a):
    scen = a.scenario or "LC_FreeExpAll"
    year = a.year or "2040"
    rows = top_corridors(scen, year, a.top)
    fs = 6.0
    rc(fs)
    fig, ax = plt.subplots(figsize=(a.width, a.height), dpi=a.dpi)
    for x, (k, twh, util) in enumerate(rows):
        col = flow_color(util)
        ax.bar(x, twh, width=.66, zorder=2, facecolor=to_rgba(col, .88),
               edgecolor=to_rgba(col, .88), linewidth=.2)
        ax.text(x, twh + .4, "%.0f%%" % (100 * util), fontsize=fs - 1.2,
                color=INK, ha="center", va="bottom", zorder=5)
    ax.set_xticks(range(len(rows)))
    ax.set_xticklabels([k.replace("|", "\n") for k, _, _ in rows],
                       fontsize=fs - 1.2, color=SOFT)
    ax.set_ylabel("TWh, both directions", fontsize=fs, labelpad=2)
    ax.set_xlim(-.7, len(rows) - .3)
    ax.set_ylim(0, ax.get_ylim()[1] * 1.08)
    ax.yaxis.set_major_locator(MaxNLocator(5))
    ax.yaxis.grid(True, color=GRID, linewidth=.5, zorder=0)
    ax.set_axisbelow(True)
    for s in ("top", "right"):
        ax.spines[s].set_visible(False)

    fig.suptitle(a.title or "Busiest corridors, %s %s (label = utilisation)"
                 % (scen, year),
                 fontsize=fs + 1, fontweight="bold", color=INK,
                 x=.012, y=.995, ha="left", va="top")
    right = map_right(a)
    fig.tight_layout(pad=.3, rect=(.012, 0, right, .92))
    flow_legend(fig, right, fs)
    return save(fig, a, "freeexp_corridors.png")



def build_blocks(step=3):
    """Equal blocks of `step` years over the build horizon.

    2025 is the starting stock, not a build year, so the blocks start in 2026
    and 15 years divide cleanly into five."""
    yrs = YEARS[1:]
    return [yrs[k:k + step] for k in range(0, len(yrs) - step + 1, step)]


def chart_freeexp_build(a):
    """What the model builds, corridor by corridor, in MW.

    One panel per corridor whose NTC grows, one bar per three-year block: the
    pale plinth is the 2025 stock, the coloured part is what free expansion
    adds, and the figure on top is the average utilisation over the block.
    Averaging over three years takes out the year-to-year hydrology noise that
    makes a single year's load factor unreadable."""
    import pandas as pd
    scen = a.scenario or "LC_FreeExpAll"
    rows = capacity_growth(scen, "2025", "2040")
    blocks = build_blocks()
    zc = cache()["zcmap"]

    def cross(key):
        a_, b_ = key.split("|")
        return zc.get(a_) != zc.get(b_)

    # The first question anyone asks of this chart is why so little crosses a
    # border, so name the cross-border candidates the model was offered and
    # turned down, with the load factor of the existing link it would have
    # doubled.
    cor = corridors(scen)
    i40 = YEARS.index("2040")
    built = {k for k, _, _ in rows}
    turned = []
    for _, r in pd.read_csv(ex.input_file_for(RUN, scen,
                                              "pNewTransmission")).iterrows():
        k = "|".join(sorted([str(r["From"]), str(r["To"])]))
        if k in built or k in [t[0] for t in turned] or not cross(k):
            continue
        turned.append((k, cor[k]["util"][i40] if k in cor else 0.0))
    fs = 6.6
    rc(fs)
    fig, axes = plt.subplots(1, len(rows), figsize=(a.width, a.height),
                             dpi=a.dpi, sharey=True)
    top = 0.0
    for ax, (key, c, dmw) in zip(axes, rows):
        base = c["ntc"][0]
        xs = [k * .80 for k in range(len(blocks))]
        for xi, blk in zip(xs, blocks):
            idx = [YEARS.index(y) for y in blk]
            ntc = c["ntc"][idx[-1]]
            if ntc <= 0:
                continue
            live = [c["util"][i] for i in idx if c["ntc"][i] > 0]
            u = sum(live) / len(live) if live else 0.0
            plinth = min(base, ntc)
            if plinth > 0:
                ax.bar(xi, plinth, width=.62, zorder=2, facecolor="#e8ecf2",
                       edgecolor="#cfd7e2", linewidth=.3)
            add = max(ntc - base, 0)
            if add > 0:
                col = to_rgba(flow_color(u), .88)
                ax.bar(xi, add, bottom=plinth, width=.62, zorder=2,
                       facecolor=col, edgecolor=col, linewidth=.2)
            ax.text(xi, ntc, "%.0f%%" % (100 * u), fontsize=fs - 1.6, color=INK,
                    ha="center", va="bottom", zorder=5)
            top = max(top, ntc)
        ax.set_xticks(xs)
        ax.set_xticklabels(["%s-%s" % (b[0][-2:], b[-1][-2:]) for b in blocks],
                           fontsize=fs - 2.0, color=SOFT)
        ax.set_xlim(xs[0] - .52, xs[-1] + .52)
        # Two lines: "AzerbaijanMain - Nakhchivan" on one line runs straight
        # into the next panel's title.
        # Two lines: "AzerbaijanMain - Nakhchivan" on one line runs straight
        # into the next panel's title.  The arrow flags a border crossing.
        ax.set_title((u"↔ " if cross(key) else "") + key.replace("|", "\n"),
                     fontsize=fs, fontweight="bold", loc="left", pad=2,
                     color=INK, linespacing=1.05)
        ax.yaxis.set_major_locator(MaxNLocator(4))
        ax.yaxis.grid(True, color=GRID, linewidth=.5, zorder=0)
        ax.set_axisbelow(True)
        for sp in ("top", "right"):
            ax.spines[sp].set_visible(False)
    axes[0].set_ylabel("NTC, MW", fontsize=fs, labelpad=2)
    for ax in axes[1:]:
        ax.spines["left"].set_visible(False)
        ax.tick_params(axis="y", length=0)
    axes[0].set_ylim(0, top * 1.14)                # headroom for the figures

    if a.title:
        fig.suptitle(a.title, fontsize=fs + 1, fontweight="bold", color=INK,
                     x=.012, y=.995, ha="left", va="top")
    note = u"↔ crosses a border.  Cross-border candidates offered and not built: "            + ", ".join("%s (%.0f%% used in 2040)" % (k.replace("|", "-"), 100 * u)
                       for k, u in turned)
    fig.text(.012, .012, note, fontsize=fs - 1.4, color=SOFT, ha="left",
             va="bottom")
    right = map_right(a)
    fig.tight_layout(pad=.3, w_pad=.4,
                     rect=(.012, .058, right, .93 if a.title else .995))
    flow_legend(fig, right, fs, extra=[
        Patch(facecolor="#e8ecf2", edgecolor="#cfd7e2", linewidth=.3,
              label="2025 stock")])
    return save(fig, a, "freeexp_build.png")

# ------------------------------------------------------------------ benefits

# Every project is judged against the counterfactual that shares its EU price
# assumption.  Comparing a Crisis-price project with the central-price baseline
# would credit the project with the price path, which is not a project benefit.
PRICE_PATHS = [("EU central", "baseline", ""),
               ("EU very low", "LC_Base_VeryLow", "_VeryLow"),
               ("EU crisis", "LC_Base_Crisis", "_Crisis"),
               ("EU central + CBAM", "LC_Base_CBAM", "_CBAM")]
PROJECTS = [("BSSC", "LC_BSSC"), ("All\nprojects", "LC_AllProjects")]

# Stacking and legend order.  Signs are cost signs: a positive NPV is money the
# system spends, so a benefit is counterfactual minus scenario.
COMPS = [
    ("exp_ext", "Export revenue, external", "#c9a227"),
    ("fuel",    "Fuel cost",                "#2f3f57"),
    ("capex",   "Generation capex",         "#8fb3d4"),
    ("fom",     "Fixed O&M",                "#6f97bd"),
    ("vom",     "Variable O&M",             "#33628f"),
    ("res",     "Reserve and unserved",     "#8a97a8"),
    ("trans",   "Transmission capex, internal", "#7fb0a3"),
    ("imp_ext", "Import cost, external",    "#e0cd8f"),
    ("newcap",  "Transmission capex, external", "#c0392b"),
]
COMPS_COUNTRY = COMPS + [
    ("exp_int", "Export revenue, internal", "#eccb72"),
    ("imp_int", "Import cost, internal",    "#d7dde5"),
    ("shared",  "Trade shared benefits",    "#9c8ab5"),
]

# pCosts headers -> component, sign.  Anything the model reports that is not one
# of the named blocks lands in "res"; those lines are pennies, but they have to
# go somewhere or the decomposition stops adding up to the NPV.
_CMAP = {
    "Fixed O&M: $m": ("fom", 1), "Variable O&M: $m": ("vom", 1),
    "Fuel costs: $m": ("fuel", 1), "Transmission costs: $m": ("trans", 1),
    "Import costs with external zones: $m": ("imp_ext", 1),
    # generate_report writes this one as a positive magnitude, but base.gms:679
    # subtracts it from the objective.  Flip it or the NPV will not reconcile.
    "Export revenues with external zones: $m": ("exp_ext", -1),
    "Import costs with internal zones: $m": ("imp_int", 1),
    "Export revenues with internal zones: $m": ("exp_int", 1),
    "Trade shared benefits: $m": ("shared", 1),
}
_SKIP = {"NPV of system cost: $m"}


@lru_cache(maxsize=1)
def _rr():
    """Discount factor by year, read back out of the model's own output.

    pCostsMerged carries both the undiscounted yearly cost and the discounted
    cumulated one, so the ratio of the two is pRR * pWeightYear exactly.
    Rebuilding it from DR would miss the half-year convention EPM uses.
    """
    import pandas as pd
    d = pd.read_csv(RUN / "baseline" / "output_csv" / "pCostsMerged.csv")
    key = "Fuel costs: $m"
    un = d[(d.attribute == "Costs") & (d.uni == key)].groupby("y").value.sum()
    cu = d[(d.attribute == "DiscountedWeightedCostsCumulated")
           & (d.uni == key)].groupby("y").value.sum()
    step = cu.diff()
    step.iloc[0] = cu.iloc[0]
    return (step / un).to_dict()


@lru_cache(maxsize=1)
def _ext_capex():
    """NPV of the external interconnector annuities, by scenario and zone, $m.

    pExtTransferLimit carries no investment variable, so the model gets those
    corridors for free: their cost lives in trade/pExtTransmissionCost.csv,
    which GAMS never reads.  tools/ext_transmission_cost.py is the one place
    that reads it and works out which phases each scenario buys and when, so
    call it rather than re-deriving the ramp here.
    """
    sys.path.insert(0, str(HERE.parents[0]))
    import ext_transmission_cost as etc
    b = etc.build(ex.DATA, "LC_Baseline")
    weight, rr, years = b["weight"], b["rr"], b["years"]
    out = {}
    for d in b["rows_detail"]:
        if d["capex_musd"] is None:
            continue
        f = sum(weight[y] * rr[y] for y in years if y >= d["commissioning"])
        per = out.setdefault(d["scenario"], {})
        per[d["z"]] = per.get(d["z"], 0.0) + d["annuity_musd_per_yr"] * f
    out["baseline"] = out.get("LC_Baseline", {})      # run folder name
    return out


@lru_cache(maxsize=1)
def benefit_npv():
    """Discounted system cost by scenario, country and component, $m.

    Generation capex is not in pCosts at all, so it is discounted here from the
    yearly annuities in summary.csv.  External interconnector capex is not in
    the model either, and comes from _ext_capex.
    """
    import pandas as pd
    scens = ["baseline"] + [c for _, c, _ in PRICE_PATHS[1:]] + \
            [p + s for _, p in PROJECTS for _, _, s in PRICE_PATHS]
    sm = pd.read_csv(RUN / "summary.csv")
    zc = dict(sm[["zone", "country"]].drop_duplicates().values)
    inv = sm[sm.attribute == "Investment costs: $m"]
    rr = _rr()

    rows = []
    for s in scens:
        d = pd.read_csv(RUN / s / "output_csv" / "pCostsMerged.csv")
        d = d[d.attribute == "DiscountedWeightedCostsCumulated"]
        d = d[d.y == d.y.max()]
        for _, r in d.iterrows():
            if r["uni"] in _SKIP:
                continue
            comp, sign = _CMAP.get(r["uni"], ("res", 1))
            rows.append((s, r["z"], comp, sign * float(r["value"])))
        for _, r in inv.iterrows():
            v = float(r[s]) * rr.get(float(r["year"]), 0.0)
            if v:
                rows.append((s, r["zone"], "capex", v))

    ext = _ext_capex()
    for s in scens:
        for z, v in ext.get(s, {}).items():
            rows.append((s, z, "newcap", v))

    df = pd.DataFrame(rows, columns=["scen", "zone", "comp", "npv"])
    df["country"] = df.zone.map(zc).fillna(df.zone)
    return df


def _deltas(df, scen, cf, comps, country=None):
    """Counterfactual minus scenario, by component: positive is a benefit."""
    d = df[df.scen.isin([scen, cf])]
    if country:
        d = d[d.country == country]
    g = d.groupby(["scen", "comp"]).npv.sum()
    return {k: float(g.get((cf, k), 0.0)) - float(g.get((scen, k), 0.0))
            for k, _, _ in comps}


def _stack(ax, xs, vals, comps, width, fs):
    """One stacked bar per x, positives up and negatives down, net ruled on top."""
    pos = [0.0] * len(xs)
    neg = [0.0] * len(xs)
    for key, _, col in comps:
        for i, x in enumerate(xs):
            v = vals[i].get(key, 0.0)
            if abs(v) < 1e-9:
                continue
            base = pos[i] if v > 0 else neg[i]
            ax.bar(x, v, bottom=base, width=width, zorder=3,
                   facecolor=col, edgecolor=col, linewidth=.2)
            if v > 0:
                pos[i] += v
            else:
                neg[i] += v
    span = max(max(pos), 1e-6) - min(min(neg), 0.0)
    for i, x in enumerate(xs):
        net = pos[i] + neg[i]
        ax.plot([x - width / 2, x + width / 2], [net, net], color="#1d2735",
                linewidth=1.2, zorder=6, solid_capstyle="butt")
        ax.text(x, pos[i] + .015 * span, "%+.1f" % net, fontsize=fs - 1.0,
                color="#1d2735", fontweight="bold", ha="center", va="bottom",
                zorder=7)
    return pos, neg


def _comp_legend(fig, right, fs, comps):
    fig.legend(handles=[Patch(facecolor=c, edgecolor=c, label=lab)
                        for _, lab, c in comps]
               + [Line2D([0], [0], color="#1d2735", linewidth=1.2,
                         label="Net benefit")],
               loc="center left", bbox_to_anchor=(right + .012, .5),
               frameon=False, fontsize=fs - 1.3, labelspacing=.45,
               handlelength=1.0, handleheight=.8, borderaxespad=0)


def _benefit_axes(ax, centres, glabels, xs, blabels, fs, note_y=-.135):
    """Two rows of x labels: the project on the bar, the group underneath."""
    ax.set_xticks(xs)
    ax.set_xticklabels(blabels, fontsize=fs - 1.3, color=SOFT,
                       linespacing=.95)
    for cx, lab in zip(centres, glabels):
        ax.text(cx, note_y, lab, transform=ax.get_xaxis_transform(),
                fontsize=fs, fontweight="bold", color=INK, ha="center",
                va="top")
    ax.axhline(0, color="#b9c3d0", linewidth=.7, zorder=4)
    ax.set_ylabel("NPV difference vs counterfactual, $bn", fontsize=fs,
                  labelpad=2)
    ax.yaxis.set_major_locator(MaxNLocator(6))
    ax.yaxis.grid(True, color=GRID, linewidth=.5, zorder=0)
    ax.set_axisbelow(True)
    for sp in ("top", "right"):
        ax.spines[sp].set_visible(False)


def _grouped(n_groups, n_bars, gap=.42):
    """Bar centres for n_bars inside each of n_groups, and the group centres."""
    xs, centres = [], []
    for gi in range(n_groups):
        c = gi * 1.0
        centres.append(c)
        for bi in range(n_bars):
            xs.append(c + (bi - (n_bars - 1) / 2) * gap)
    return xs, centres


def chart_benefit_regional(a):
    """Economic benefit of each project, against its own price counterfactual.

    Stacked NPV differences in $bn: bars above zero are money the region saves
    or earns, bars below zero are money it spends.  The rule on top of each bar
    is the net, which is what the project is worth once the new interconnector
    is paid for.
    """
    df = benefit_npv()
    fs = 6.8
    rc(fs)
    fig, ax = plt.subplots(figsize=(a.width, a.height), dpi=a.dpi)
    xs, centres = _grouped(len(PRICE_PATHS), len(PROJECTS))
    vals = [{k: v / 1e3 for k, v in
             _deltas(df, base + suf, cf, COMPS).items()}
            for _, cf, suf in PRICE_PATHS for _, base in PROJECTS]
    _stack(ax, xs, vals, COMPS, .36, fs)
    ax.set_xlim(-.62, len(PRICE_PATHS) - 1 + .62)
    _benefit_axes(ax, centres, [p for p, _, _ in PRICE_PATHS], xs,
                  [b for _ in PRICE_PATHS for b, _ in PROJECTS], fs)
    if a.title:
        fig.suptitle(a.title, fontsize=fs + 1, fontweight="bold", color=INK,
                     x=.012, y=.995, ha="left", va="top")
    right = 1 - 1.72 / a.width
    fig.tight_layout(pad=.3, rect=(.012, .09, right,
                                   .93 if a.title else .995))
    _comp_legend(fig, right, fs, COMPS)
    return save(fig, a, "benefit_regional.png")


def chart_benefit_country(a):
    """The same decomposition per country, for one price path.

    Internal import and export cancel across the region but not inside it, so
    the country panel carries three extra blocks and shows who is paid and who
    pays for the same regional net.
    """
    df = benefit_npv()
    label = a.scenario or "EU central"
    look = {p: (cf, suf) for p, cf, suf in PRICE_PATHS}
    if label not in look:                                # accept a raw suffix
        label = next((p for p, _, s in PRICE_PATHS if s == label), "EU central")
    cf, sfx = look[label]
    fs = 6.8
    rc(fs)
    fig, ax = plt.subplots(figsize=(a.width, a.height), dpi=a.dpi)
    xs, centres = _grouped(len(COUNTRIES), len(PROJECTS))
    vals = [{k: v / 1e3 for k, v in
             _deltas(df, base + sfx, cf, COMPS_COUNTRY, country=ctry).items()}
            for ctry in COUNTRIES for _, base in PROJECTS]
    _stack(ax, xs, vals, COMPS_COUNTRY, .36, fs)
    ax.set_xlim(-.62, len(COUNTRIES) - 1 + .62)
    _benefit_axes(ax, centres, COUNTRIES, xs,
                  [b for _ in COUNTRIES for b, _ in PROJECTS], fs)
    if a.title:
        fig.suptitle(a.title, fontsize=fs + 1, fontweight="bold", color=INK,
                     x=.012, y=.995, ha="left", va="top")
    fig.text(.012, .012, "%s price path.  Internal trade nets to zero over the "
             "region, not within a country." % label,
             fontsize=fs - 1.4, color=SOFT, ha="left", va="bottom")
    right = 1 - 1.72 / a.width
    fig.tight_layout(pad=.3, rect=(.012, .125, right,
                                   .93 if a.title else .995))
    _comp_legend(fig, right, fs, COMPS_COUNTRY)
    return save(fig, a, "benefit_country.png")


def table_benefits(a):
    """The numbers behind the two benefit charts, $m of NPV."""
    df = benefit_npv()
    out = OUTDIR / "table_benefits.csv"
    out.parent.mkdir(parents=True, exist_ok=True)
    with out.open("w", newline="", encoding="utf-8") as fh:
        w = csv.writer(fh)
        w.writerow(["Price path", "Project", "Scope"]
                   + [lab for _, lab, _ in COMPS_COUNTRY] + ["Net benefit"])
        for plabel, cf, suf in PRICE_PATHS:
            for blabel, base in PROJECTS:
                for scope in ["Region"] + COUNTRIES:
                    d = _deltas(df, base + suf, cf, COMPS_COUNTRY,
                                country=None if scope == "Region" else scope)
                    w.writerow([plabel, blabel.replace(chr(10), " "), scope]
                               + ["%.0f" % d[k] for k, _, _ in COMPS_COUNTRY]
                               + ["%.0f" % sum(d.values())])
    print(out)
    return out


def table_benefits_xlsx(a):
    """The same numbers as a workbook, for anyone who wants to slice them.

    Four sheets: the headline blocks the deck shows, the full regional
    decomposition, the same by country, and the notes that keep the figures
    from being misread once they leave this repo.
    """
    import pandas as pd
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    df = benefit_npv()
    labs = [lab for _, lab, _ in COMPS_COUNTRY]

    def block(scope):
        rows = []
        for plabel, cf, suf in PRICE_PATHS:
            for blabel, base in PROJECTS:
                d = _deltas(df, base + suf, cf, COMPS_COUNTRY,
                            country=None if scope == "Region" else scope)
                rows.append([plabel, blabel.replace(chr(10), " ")]
                            + [round(d[k]) for k, _, _ in COMPS_COUNTRY]
                            + [round(sum(d.values()))])
        return pd.DataFrame(rows, columns=["EU price path", "Project"]
                            + labs + ["Net benefit"])

    reg = block("Region")
    cty = pd.concat([block(c).assign(Country=c) for c in COUNTRIES])
    cty = cty[["EU price path", "Project", "Country"] + labs + ["Net benefit"]]
    head = reg[["EU price path", "Project", "Export revenue, external",
                "Fuel cost", "Generation capex", "Fixed O&M",
                "Transmission capex, external", "Net benefit"]]
    notes = pd.DataFrame({"Note": [
        "NPV over %s-%s, $m, discounted on the model's own year factors "
        "(DR = 6%%, half-year convention)." % (YEARS[0], YEARS[-1]),
        "Each project is compared with the counterfactual sharing its EU price "
        "path, not with the central baseline: %s."
        % ", ".join("%s -> %s" % (p, c) for p, c, _ in PRICE_PATHS),
        "Positive = benefit, i.e. counterfactual cost minus scenario cost.",
        "'Transmission capex, external' is the capex of the external "
        "interconnectors, "
        "which the model does not price: pExtTransferLimit carries no "
        "investment variable. Source trade/pExtTransmissionCost.csv via "
        "tools/ext_transmission_cost.py.",
        "'Generation capex' is discounted here from the yearly annuities in "
        "summary.csv; pCosts does not carry it.",
        "Export revenue with external zones is written as a positive magnitude "
        "by generate_report.gms but subtracted from the objective "
        "(base.gms:679); the sign is flipped here.",
        "Components sum back to the model's 'NPV of system cost' within "
        "0.6 $m on 211,000 $m, for all twelve scenarios.",
        "Internal import and export cancel across the region but not within a "
        "country, which is why the country sheet carries three extra blocks.",
        "Source run: %s." % RUN.name,
    ]})

    out = OUTDIR / "table_benefits.xlsx"
    out.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(out, engine="openpyxl") as w:
        head.to_excel(w, sheet_name="Regional summary", index=False)
        reg.to_excel(w, sheet_name="Regional full", index=False)
        cty.to_excel(w, sheet_name="By country", index=False)
        notes.to_excel(w, sheet_name="Notes", index=False)
        hf = Font(bold=True, color="FF3F4E63")
        fill = PatternFill("solid", fgColor="FFEDF1F6")
        for name, ws in w.sheets.items():
            ws.freeze_panes = "A2"
            for c in ws[1]:
                c.font, c.fill = hf, fill
                c.alignment = Alignment(wrap_text=True, vertical="bottom",
                                        horizontal="center")
            ws.row_dimensions[1].height = 15 if name == "Notes" else 30
            for j, col in enumerate(ws.iter_cols(min_row=1), start=1):
                wide = max(len(str(c.value or "")) for c in col)
                ws.column_dimensions[get_column_letter(j)].width = (
                    110 if name == "Notes" else min(max(wide + 2, 9), 16))
                for c in col[1:]:
                    if isinstance(c.value, (int, float)):
                        c.number_format = "+#,##0;-#,##0;0"
            if name != "Notes":                       # net column stands out
                for c in ws[get_column_letter(ws.max_column)][1:]:
                    c.font = Font(bold=True)
    print(out)
    return out


def table_levels_xlsx(a):
    """System cost in levels, not differences, for one price path.

    Same decomposition as the benefit charts, but each scenario stands on its
    own: counterfactual, BSSC and All projects side by side in $m of NPV.  The
    total is the model's objective plus the external interconnector capex the
    model never sees, so it is the full bill, not the reported one.
    """
    import pandas as pd
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    label = a.scenario or "EU central"
    look = {p: (cf, suf) for p, cf, suf in PRICE_PATHS}
    if label not in look:                                # accept a raw suffix
        label = next((p for p, _, s in PRICE_PATHS if s == label), "EU central")
    cf, sfx = look[label]
    cols = [("Counterfactual", cf)] + [(b.replace(chr(10), " "), s + sfx)
                                       for b, s in PROJECTS]

    df = benefit_npv()

    def block(scope):
        d = df if scope == "Region" else df[df.country == scope]
        g = d.groupby(["scen", "comp"]).npv.sum()
        rows = []
        for key, lab, _ in COMPS_COUNTRY:
            rows.append([lab] + [round(float(g.get((s, key), 0.0)))
                                 for _, s in cols])
        rows.append(["Total system cost, NPV"]
                    + [round(sum(float(g.get((s, k), 0.0))
                                 for k, _, _ in COMPS_COUNTRY))
                       for _, s in cols])
        return pd.DataFrame(rows, columns=["Component"] + [c for c, _ in cols])

    reg = block("Region")
    cty = pd.concat([block(c).assign(Country=c) for c in COUNTRIES])
    cty = cty[["Country", "Component"] + [c for c, _ in cols]]
    notes = pd.DataFrame({"Note": [
        "%s price path. Scenarios: %s."
        % (label, ", ".join("%s = %s" % (c, s) for c, s in cols)),
        "NPV over %s-%s, $m, discounted on the model's own year factors "
        "(DR = 6%%, half-year convention). Levels, not differences."
        % (YEARS[0], YEARS[-1]),
        "Positive = money the system spends. Export revenue is negative "
        "because it offsets the bill.",
        "Everything above 'Transmission capex, external' sums to the model's "
        "'NPV of system cost'; the total row adds the external interconnector "
        "capex, which the model does not price (pExtTransferLimit carries no "
        "investment variable). Source trade/pExtTransmissionCost.csv via "
        "tools/ext_transmission_cost.py.",
        "'Generation capex' is discounted here from the yearly annuities in "
        "summary.csv; pCosts does not carry it.",
        "Export revenue with external zones is written as a positive magnitude "
        "by generate_report.gms but subtracted from the objective "
        "(base.gms:679); the sign is flipped here.",
        "Internal import, export and shared benefits cancel across the region "
        "but not within a country.",
        "Source run: %s." % RUN.name,
    ]})

    out = OUTDIR / "table_costs_absolute.xlsx"
    out.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(out, engine="openpyxl") as w:
        reg.to_excel(w, sheet_name="Region", index=False)
        cty.to_excel(w, sheet_name="By country", index=False)
        notes.to_excel(w, sheet_name="Notes", index=False)
        hf = Font(bold=True, color="FF3F4E63")
        fill = PatternFill("solid", fgColor="FFEDF1F6")
        for name, ws in w.sheets.items():
            ws.freeze_panes = "B2" if name == "Region" else "A2"
            for c in ws[1]:
                c.font, c.fill = hf, fill
                c.alignment = Alignment(wrap_text=True, vertical="bottom",
                                        horizontal="center")
            for j, col in enumerate(ws.iter_cols(min_row=1), start=1):
                wide = max(len(str(c.value or "")) for c in col)
                ws.column_dimensions[get_column_letter(j)].width = (
                    110 if name == "Notes" else min(max(wide + 2, 10), 26))
                for c in col[1:]:
                    if isinstance(c.value, (int, float)):
                        c.number_format = "#,##0;-#,##0;0"
            for row in ws.iter_rows(min_row=2):          # total rows in bold
                if any(str(c.value or "").startswith("Total") for c in row[:2]):
                    for c in row:
                        c.font = Font(bold=True)
    print(out)
    return out


# -------------------------------------------------------------------- tables

def table_bssc(a):
    import pandas as pd
    cost = pd.read_csv(ex.DATA / "trade" / "pExtTransmissionCost.csv")
    row = cost[(cost["z"] == "Georgia") & (cost["phase"] == "BSSC")].iloc[0]
    f = bssc_facts("LC_BSSC", 2035)
    mw, musd = float(row["CapacityMW"]), float(row["CapexMUSD"])
    out = OUTDIR / "table_bssc.csv"
    out.parent.mkdir(parents=True, exist_ok=True)
    with out.open("w", newline="", encoding="utf-8") as fh:
        w = csv.writer(fh)
        w.writerow(["Item", "Value"])
        w.writerow(["Route", "Georgia - Romania, Black Sea submarine (%s)"
                    % row["family"]])
        w.writerow(["Capacity (MW)", "%.0f" % mw])
        w.writerow(["COD in the model / first year of flow",
                    "%s / %s" % (f["cod"], f["first"])])
        w.writerow(["Capex", "%.0f $m  (%.0f $/kW)" % (musd, musd * 1e3 / mw)])
        w.writerow(["Economic life (years)", "%.0f" % float(row["Life"])])
        w.writerow(["2035 flow (TWh)", "%.2f to Romania / %.2f to Georgia"
                    % (f["exp"], f["imp"])])
        w.writerow(["2035 load factor (%)", "%.0f" % (100 * f["util"])])
        w.writerow(["2035 price, Georgia / Romania ($/MWh)",
                    "%.0f / %.0f" % (f["ge_price"], f["ro_sell"])])
    print(out)
    return out


def table_freeexp(a):
    import pandas as pd
    scen = a.scenario or "LC_FreeExpAll"
    proj = pd.read_csv(ex.input_file_for(RUN, scen, "pNewTransmission"))
    # CostPerLine is $m per line (base.gms multiplies it by 1e6), so this is
    # the per-MW price the model actually paid for what it built.
    unit = {}
    for _, r in proj.iterrows():
        k = "|".join(sorted([str(r["From"]), str(r["To"])]))
        unit[k] = float(r["CostPerLine"]) / float(r["CapacityPerLine"])

    add = _tm(scen)
    add = add[(add["attribute"] == "NewTransmissionCapacity") & (add["value"] > 5)]
    first = {}
    for _, r in add.iterrows():
        k = "|".join(sorted([r["z"], r["uni"]]))
        first[k] = min(first.get(k, "9999"), r["y"])

    rows = [[k.replace("|", " - "), dmw, first.get(k, ""),
             dmw * unit.get(k, float("nan"))]
            for k, c, dmw in capacity_growth(scen, "2025", "2040")]
    out = OUTDIR / "table_freeexp_projects.csv"
    out.parent.mkdir(parents=True, exist_ok=True)
    with out.open("w", newline="", encoding="utf-8") as fh:
        w = csv.writer(fh)
        w.writerow(["Route", "Added MW by 2040", "First year built", "Capex ($m)"])
        keep, rest = rows[:7], rows[7:]
        for r in keep:
            w.writerow([r[0], "%.0f" % r[1], r[2],
                        "" if r[3] != r[3] else "%.0f" % r[3]])
        if rest:
            w.writerow(["Others (%d corridors)" % len(rest),
                        "%.0f" % sum(r[1] for r in rest), "",
                        "%.0f" % sum(r[3] for r in rest if r[3] == r[3])])
    print(out)
    return out


# Project names for the corridors LC_AllProjects switches on.  The model only
# knows zone pairs, so the names come from pre-analysis/data/reference_lines.csv
# and from the phase names in the external capex catalogue.
PROJECT_NAMES = {
    ("Georgia", "Romania", "BSSC"): "Black Sea Submarine Cable (BSSC)",
    ("Georgia", "Romania", "GECO"): "Georgia-EU corridor, GECO phase",
    ("Trakia", "Bulgaria", "EWTC"): "East-West Transmission Corridor (EWTC)",
    ("Trakia", "Greece", "EWTC"): "East-West Transmission Corridor (EWTC)",
    ("AzerbaijanMain", "Kazakhstan", "TransCaspian"):
        "Trans-Caspian interconnector",
    ("AzerbaijanMain", "Nakhchivan"): "Zangezur corridor",
    ("Georgia", "Armenia"): "Caucasus Transmission Network (CTN)",
    ("EastAna", "Georgia"): "BSTN Extension, 3rd B2B at Akhaltsikhe",
    ("Nakhchivan", "EastAna"):
        "Azerbaijan-Turkiye-Europe Green Corridor, Zangezur phase",
}


def _yr(v):
    """Commissioning years arrive as 2031.0 from the capex catalogue."""
    try:
        return "%d" % float(v)
    except (TypeError, ValueError):
        return str(v)


def table_allprojects(a):
    """The transmission projects LC_AllProjects contains, one row per line.

    Two blocks, because they are not priced the same way: the external
    corridors widen pExtTransferLimit, which carries no investment variable, so
    the model gets them for free and their capex only enters as a post-hoc
    charge; the internal and cross-border candidates go through
    pNewTransmission and the model pays for them in the objective.
    """
    import pandas as pd
    scen = a.scenario or "LC_AllProjects"

    sys.path.insert(0, str(HERE.parents[0]))
    import ext_transmission_cost as etc
    detail = [d for d in etc.build(ex.DATA, "LC_Baseline")["rows_detail"]
              if d["scenario"] == scen]

    rows = []
    for d in sorted(detail, key=lambda d: d["commissioning"]):
        rows.append(["External", PROJECT_NAMES.get(
            (d["z"], d["zext"], d["phase"]), d["phase"]),
            "%s - %s" % (d["z"], d["zext"]), d["mw"], _yr(d["commissioning"]),
            d["capex_musd"]])

    # What the model was offered, and what it actually built.  All four
    # candidates are taken at their earliest entry today, but do not assume it:
    # read the build year back out of the results.
    cand = pd.read_csv(ex.input_file_for(RUN, scen, "pNewTransmission"))
    tm = _tm(scen)
    built = tm[(tm["attribute"] == "NewTransmissionCapacity") & (tm["value"] > 1)]
    year = {}
    for _, r in built.iterrows():
        k = "|".join(sorted([str(r["z"]), str(r["uni"])]))
        year[k] = min(year.get(k, "9999"), str(r["y"]))
    for _, r in cand.iterrows():
        if int(r["MaximumNumOfLines"]) < 1:
            continue
        fr, to = str(r["From"]), str(r["To"])
        k = "|".join(sorted([fr, to]))
        rows.append(["Internal",
                     PROJECT_NAMES.get((fr, to),
                                       PROJECT_NAMES.get((to, fr), "")),
                     "%s - %s" % (fr, to), float(r["CapacityPerLine"]),
                     _yr(year.get(k, "not built")), float(r["CostPerLine"])])

    df = pd.DataFrame(rows, columns=["Block", "Project", "Line (A - B)",
                                     "Capacity (MW)", "COD",
                                     "Capex ($m)"])
    df["COD"] = df["COD"].astype(str)
    df = df.sort_values(["Block", "COD"])   # External block first, then COD
    out = OUTDIR / "table_allprojects.csv"
    out.parent.mkdir(parents=True, exist_ok=True)
    df.to_csv(out, index=False, encoding="utf-8")
    print(out)

    xl = OUTDIR / "table_allprojects.xlsx"
    _write_project_xlsx(xl, df, scen)
    print(xl)
    return xl


def _write_project_xlsx(path, df, scen):
    import pandas as pd
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    heads = {"External": "External corridors - capacity is exogenous "
                         "(pExtTransferLimit), capex is not in the model "
                         "objective",
             "Internal": "Internal and cross-border candidates - the model "
                         "chooses them and pays for them (pNewTransmission)"}
    lines = []
    for block in ["External", "Internal"]:
        sub = df[df.Block == block]
        lines.append([heads[block], "", "", "", ""])
        for _, r in sub.iterrows():
            lines.append([r["Project"], r["Line (A - B)"], r["Capacity (MW)"],
                          r["COD"], r["Capex ($m)"]])
        lines.append(["Subtotal", "", sub["Capacity (MW)"].sum(), "",
                      sub["Capex ($m)"].sum()])
    lines.append(["Total", "", df["Capacity (MW)"].sum(), "",
                  df["Capex ($m)"].sum()])
    tab = pd.DataFrame(lines, columns=["Project", "Line (A - B)",
                                       "Capacity (MW)", "COD", "Capex ($m)"])
    notes = pd.DataFrame({"Note": [
        "Scenario %s, run %s." % (scen, RUN.name),
        "COD for the external corridors is the year pExtTransferLimit steps "
        "up; for the candidates it is the year the model actually builds, read "
        "back from pTransmissionMerged.",
        "Capex for the external corridors comes from "
        "trade/pExtTransmissionCost.csv, which GAMS never reads: those "
        "corridors are free to the model and their cost is charged afterwards "
        "(3,865 $m of NPV at DR = 6%).",
        "Capex for the candidates is CostPerLine in pNewTransmission, which "
        "the model does carry in the objective.",
        "Economic life is 30 years for every line here.",
        "Project names: pre-analysis/data/reference_lines.csv.",
    ]})

    with pd.ExcelWriter(path, engine="openpyxl") as w:
        tab.to_excel(w, sheet_name="Projects", index=False)
        notes.to_excel(w, sheet_name="Notes", index=False)
        hf = Font(bold=True, color="FF3F4E63")
        fill = PatternFill("solid", fgColor="FFEDF1F6")
        band = PatternFill("solid", fgColor="FFF6F8FB")
        for name, ws in w.sheets.items():
            ws.freeze_panes = "A2"
            for c in ws[1]:
                c.font, c.fill = hf, fill
                c.alignment = Alignment(wrap_text=True, vertical="bottom")
            for j, col in enumerate(ws.iter_cols(min_row=1), start=1):
                wide = max(len(str(c.value or "")) for c in col)
                ws.column_dimensions[get_column_letter(j)].width = (
                    110 if name == "Notes" else min(max(wide + 2, 10), 56))
                for c in col[1:]:
                    if isinstance(c.value, (int, float)):
                        c.number_format = "#,##0"
            if name == "Projects":
                for row in ws.iter_rows(min_row=2):
                    head = str(row[0].value or "")
                    if head.startswith(("External corridors",
                                        "Internal and cross-border")):
                        for c in row:
                            c.font, c.fill = Font(bold=True, italic=True), band
                    elif head.startswith(("Subtotal", "Total")):
                        for c in row:
                            c.font = Font(bold=True)


# ---------------------------------------------------------------------- main

# The nine All Projects candidates, in the order of the slide-64 table.  The
# corridor keys are alphabetical because corridors() sorts each pair, so the
# solid half of every bar is the first-named zone feeding the second.
AP_LINES = [
    ("External", "BSSC + GECO",    "Georgia|Romania"),
    ("External", "EWTC 1",         "Bulgaria|Trakia"),
    ("External", "EWTC 2",         "Greece|Trakia"),
    ("External", "Trans-Caspian",  "AzerbaijanMain|Kazakhstan"),
    ("Internal", "TRIPP 1",        "AzerbaijanMain|Nakhchivan"),
    ("Internal", "CTN",            "Armenia|Georgia"),
    ("Internal", "BSTN Extension", "EastAna|Georgia"),
    ("Internal", "TRIPP 2",        "EastAna|Nakhchivan"),
]

AP_SHORT = {"AzerbaijanMain": "AZ Main", "EastAna": "East Ana.",
            "Nakhchivan": "Nakhch.", "Georgia": "Georgia", "Armenia": "Armenia",
            "Trakia": "Trakia", "Romania": "Romania", "Bulgaria": "Bulgaria",
            "Greece": "Greece", "Kazakhstan": "Kazakhstan"}


def chart_allprojects_lines(a):
    """Utilisation and volume of the nine project corridors, on one axis.

    Plotting TWh directly would put a 5.2 GW cable (45 TWh of annual capacity)
    next to a 166 MW tie (1.4 TWh) and flatten every small link, so the bar
    length is the load factor - annual energy against NTC x 8760 h - split into
    the two directions, and the volume itself is printed at the bar end.  That
    is the same load-factor definition for internal and external links, which
    the model's own InterconUtilization is not: it is reported per direction
    for internal pairs and not at all for external ones.
    """
    scen = a.scenario or "LC_AllProjects"
    cor = corridors(scen)
    years = ([y.strip() for y in a.years.split(",")] if a.years
             else ["2035", "2040"])
    col = {years[0]: FLOW_COOL, years[-1]: FLOW_WARM}

    fs = 6.0
    rc(fs)
    fig, ax = plt.subplots(figsize=(a.width, a.height), dpi=a.dpi)

    bh = .74 / len(years)
    ticks, labels = [], []
    for row, (block, name, key) in enumerate(AP_LINES):
        c = cor.get(key)
        ticks.append(row)
        if not c:
            labels.append(name)
            continue
        # Orient every row on its dominant direction in the last year plotted,
        # so the solid half always carries the main flow and a hatched majority
        # means the link ran backwards that year.
        last = YEARS.index(years[-1])
        flip = c["rev"][last] > c["fwd"][last]
        za, zb = (c["b"], c["a"]) if flip else (c["a"], c["b"])
        labels.append(name + chr(10) + "%s to %s" % (AP_SHORT.get(za, za),
                                                     AP_SHORT.get(zb, zb)))
        for j, y in enumerate(years):
            i = YEARS.index(y)
            cap = c["ntc"][i] * 8760.0 / 1e6                      # TWh a year
            yy = row - .37 + bh * (j + .5)
            ax.barh(yy, 100, height=bh * .86, zorder=1,
                    facecolor="#eef2f7", edgecolor="#e0e6ee", linewidth=.4)
            if cap <= 0:
                ax.text(1.5, yy, "not built", fontsize=fs - 1.2, color=SOFT,
                        ha="left", va="center", zorder=4)
                continue
            f0, r0 = (c["rev"][i], c["fwd"][i]) if flip else (c["fwd"][i],
                                                               c["rev"][i])
            fwd, rev = 100 * f0 / cap, 100 * r0 / cap
            k = col[y]
            ax.barh(yy, fwd, height=bh * .86, zorder=2,
                    facecolor=to_rgba(k, .88), edgecolor=to_rgba(k, .88),
                    linewidth=.2)
            ax.barh(yy, rev, left=fwd, height=bh * .86, zorder=2,
                    facecolor=to_rgba(k, .26), edgecolor=k, linewidth=.45,
                    hatch="///")
            # Both figures live to the right of the envelope: inside the bar
            # the percentage kept landing on a hatched band and went unreadable.
            ax.text(109, yy, "%.0f%%" % (fwd + rev), fontsize=fs - .4,
                    color=INK, ha="right", va="center", zorder=4,
                    fontweight="bold")
            ax.text(113, yy, "%.1f TWh" % (f0 + r0), fontsize=fs - .4,
                    color=SOFT, ha="left", va="center", zorder=4)

    # A hairline between the external block and the internal one: the two are
    # priced and financed differently and the slide table separates them too.
    cut = sum(1 for b, _, _ in AP_LINES if b == "External") - .5
    ax.axhline(cut, color="#c9d3e0", linewidth=.7, zorder=3)

    ax.set_yticks(ticks)
    ax.set_yticklabels(labels, fontsize=fs - .4, color=INK)
    ax.invert_yaxis()
    ax.set_xlim(0, 100)
    ax.set_ylim(len(AP_LINES) - .45, -.55)
    ax.set_xticks([0, 25, 50, 75, 100])
    ax.set_xticklabels(["0", "25", "50", "75", "100%"], fontsize=fs)
    ax.set_xlabel("Average line utilization   |   "
                  "solid = first-named zone to second", fontsize=fs, labelpad=2)
    ax.xaxis.grid(True, color=GRID, linewidth=.5, zorder=0)
    ax.set_axisbelow(True)
    for s in ("top", "right", "left"):
        ax.spines[s].set_visible(False)

    fig.suptitle(a.title or "Lines utilization, All Projects scenario",
                 fontsize=fs + 1, fontweight="bold", color=INK,
                 x=.012, y=.995, ha="left", va="top")
    right = 1 - .60 / a.width
    fig.tight_layout(pad=.3, rect=(.012, 0, right, .90))
    h = [Patch(label=y, facecolor=to_rgba(col[y], .88),
               edgecolor=to_rgba(col[y], .88), linewidth=.2) for y in years]
    fig.legend(handles=h, loc="upper right", bbox_to_anchor=(.997, .93),
               frameon=False, fontsize=fs - .4, handlelength=1.1,
               handleheight=.85, labelspacing=.35, borderaxespad=0)
    return save(fig, a, "allprojects_lines.png")


# ------------------------------------------------- external trade valuation

EXTZ_COLOR = {"Romania": "#1B6CA8", "Bulgaria": "#36B5B5", "Greece": "#7CC8FA",
              "Kazakhstan": "#C0682A", "Russia": "#9aa5b4",
              "Iran": "#B8921A", "Iraq": "#C4A820", "Syria": "#7A7068"}
EXTZ_OWNER = {"Romania": "Georgia", "Russia": "Georgia", "Bulgaria": "Turkiye",
              "Greece": "Turkiye", "Kazakhstan": "Azerbaijan",
              "Iran": "Turkiye", "Iraq": "Turkiye", "Syria": "Turkiye"}


@lru_cache(maxsize=None)
def _hours():
    return ex.read_hours()


@lru_cache(maxsize=None)
def _price_file(scen, param):
    import pandas as pd
    df = pd.read_csv(ex.input_file_for(RUN, scen, param))
    df["year"] = df["year"].astype(str)
    return df


def ext_price(scen, param, zext, year):
    """One external price file flattened to (q, d, t, p, h).

    The trade files are wide: 24 hour columns per season and day type.  The
    charts need them long and carrying the hour weight, because every average
    that means anything here is weighted by pHours.
    """
    import pandas as pd
    df = _price_file(scen, param)
    df = df[(df["zext"] == zext) & (df["year"] == str(year))]
    h = _hours()
    rows = [(r["q"], r["d"], t, float(r[t]),
             h.get("%s|%s|%s" % (r["q"], r["d"], t), 0.0))
            for _, r in df.iterrows() for t in ex.HOURS]
    return pd.DataFrame(rows, columns=["q", "d", "t", "p", "h"])


def ext_trade(scen):
    """Every external corridor of a scenario: NTC and annual energy each way.

    Keyed on (internal zone, external zone) rather than the sorted pair
    corridors() uses, because revenue has a direction.
    """
    import pandas as pd
    tm = _tm(scen)
    out = {}

    def slot(z, zext):
        return out.setdefault((z, zext), {"ntc": [0.0] * 16, "exp": [0.0] * 16,
                                          "imp": [0.0] * 16})

    for attr, field in (("InterchangeExternalExports", "exp"),
                        ("InterchangeExternalImports", "imp")):
        for _, r in tm[tm["attribute"] == attr].iterrows():
            slot(r["z"], r["uni"])[field][YEARS.index(r["y"])] += float(r["value"] or 0)

    f = ex.input_file_for(RUN, scen, "pExtTransferLimit")
    if f and f.exists():
        lim = pd.read_csv(f)
        lim.columns = [str(c).strip() for c in lim.columns]
        for (z, zext), g in lim.groupby(["z", "zext"]):
            c = slot(z, zext)
            for i, y in enumerate(YEARS):
                if y in g.columns:
                    c["ntc"][i] = float(pd.to_numeric(g[y], errors="coerce").max() or 0)
    return out


def merchant_value(prices, ntc, gwh, sell):
    """Value the model's annual energy at the hours it would actually flow.

    The outputs carry no hourly corridor flow, so the price to apply is not
    obvious.  A plain pHours mean is wrong: a link at 85 % load factor skips
    the cheapest hours, and on the Turkish ties that gap is 15 $/MWh.  So the
    model's annual GWh is poured into the most profitable hours first, capped
    at NTC.  Checked against the Georgia-Romania cable, where a pure price rule
    (export whenever the Georgian price is under the Romanian one) reproduces
    the model to within 0.3 %, which is what says the links are merchant.

    Returns (value in $m, flow-weighted price, hours at full NTC).
    """
    if ntc <= 0 or gwh <= 0 or not len(prices):
        return 0.0, float("nan"), 0.0
    d = prices.sort_values("p", ascending=not sell)
    left, val, used = gwh * 1e3, 0.0, 0.0
    for p, h in zip(d["p"], d["h"]):
        take = min(ntc * h, left)
        if take <= 0:
            break
        val += take * p
        used += take / ntc
        left -= take
    return val / 1e6, val / (gwh * 1e3), used


def ext_revenue(scen):
    """Export revenue and import cost by external zone, $m a year."""
    tr = ext_trade(scen)
    rev, cost = {}, {}
    for (z, zext), c in tr.items():
        rv, ct = [0.0] * 16, [0.0] * 16
        for i, y in enumerate(YEARS):
            if c["ntc"][i] <= 0:
                continue
            if c["exp"][i] > 0:
                rv[i] = merchant_value(ext_price(scen, "pTradePriceExport", zext, y),
                                       c["ntc"][i], c["exp"][i], True)[0]
            if c["imp"][i] > 0:
                ct[i] = merchant_value(ext_price(scen, "pTradePrice", zext, y),
                                       c["ntc"][i], c["imp"][i], False)[0]
        rev[(z, zext)], cost[(z, zext)] = rv, ct
    return rev, cost


def chart_export_revenue(a):
    scen = a.scenario or "LC_AllProjects"
    rev, cost = ext_revenue(scen)
    keys = [k for k in sorted(rev, key=lambda k: -sum(rev[k]))
            if sum(rev[k]) + sum(cost[k]) > 0]

    fs = 6.0
    rc(fs)
    fig, ax = plt.subplots(figsize=(a.width, a.height), dpi=a.dpi)
    x = list(range(len(YEARS)))
    base = [0.0] * 16
    for z, zext in keys:
        v = rev[(z, zext)]
        col = EXTZ_COLOR.get(zext, "#9aa5b4")
        ax.bar(x, v, bottom=base, width=.74, zorder=2,
               facecolor=to_rgba(col, .88), edgecolor=to_rgba(col, .88),
               linewidth=.2)
        base = [b + s for b, s in zip(base, v)]
    net = [b - sum(cost[k][i] for k in keys) for i, b in enumerate(base)]
    ax.plot(x, net, marker="o", markersize=2.4, linewidth=.9, color="#2f3f57",
            zorder=5)

    ax.set_xticks(x)
    ax.set_xticklabels([y if y[-1] in "05" else "" for y in YEARS], fontsize=fs,
                       color=INK)
    ax.set_ylabel("$ million a year", fontsize=fs, labelpad=2)
    ax.set_xlim(-.7, 15.7)
    ax.yaxis.grid(True, color=GRID, linewidth=.5, zorder=0)
    ax.set_axisbelow(True)
    for s in ("top", "right"):
        ax.spines[s].set_visible(False)

    fig.suptitle(a.title or "Export revenue by external market",
                 fontsize=fs + 1, fontweight="bold", color=INK,
                 x=.012, y=.995, ha="left", va="top")
    right = 1 - LEGEND_IN / a.width
    fig.tight_layout(pad=.3, rect=(.012, 0, right, .92))
    h = [Patch(label="%s (%s)" % (zext, EXTZ_OWNER.get(zext, z)),
               facecolor=to_rgba(EXTZ_COLOR.get(zext, "#9aa5b4"), .88),
               edgecolor=to_rgba(EXTZ_COLOR.get(zext, "#9aa5b4"), .88),
               linewidth=.2) for z, zext in keys]
    h.append(Line2D([], [], marker="o", markersize=2.4, linewidth=.9,
                    color="#2f3f57", label="Net of import cost"))
    fig.legend(handles=h, loc="upper left", bbox_to_anchor=(right + .012, .93),
               frameon=False, fontsize=fs - .4, handlelength=1.1,
               handleheight=.85, labelspacing=.35, borderaxespad=0)
    return save(fig, a, "export_revenue.png")


def cable_hourly(scen, zint, zext, year):
    """Hour-by-hour direction of an external link, from the price spread.

    Validated on Georgia-Romania: exporting whenever the domestic marginal cost
    sits below the external export price, importing whenever it sits above the
    external import price, at full NTC in both cases, reproduces the model's
    annual energy to within 0.3 % in every year of every BSSC scenario.  The
    outputs carry no hourly corridor flow, so this is the only way to read the
    link's shape.
    """
    i = YEARS.index(str(year))
    ntc = ext_trade(scen)[(zint, zext)]["ntc"][i]
    import pandas as pd
    hp = pd.read_csv(RUN / scen / "output_csv" / "pHourlyPrice.csv")
    hp["y"] = hp["y"].astype(str)
    g = hp[(hp["z"] == zint) & (hp["y"] == str(year))][["q", "d", "t", "value"]]
    pe = ext_price(scen, "pTradePriceExport", zext, year).rename(columns={"p": "pe"})
    pi = ext_price(scen, "pTradePrice", zext, year).rename(columns={"p": "pi"})
    m = (g.rename(columns={"value": "pg"})
         .merge(pe, on=["q", "d", "t"])
         .merge(pi[["q", "d", "t", "pi"]], on=["q", "d", "t"]))
    m["mw"] = ((m["pg"] < m["pe"]).astype(float)
               - (m["pg"] > m["pi"]).astype(float)) * ntc
    m["exp"] = (m["mw"] > 0) * m["h"] * ntc / 1e3                      # GWh
    m["imp"] = (m["mw"] < 0) * m["h"] * ntc / 1e3
    return m


def chart_bssc_seasonality(a):
    scen = a.scenario or "LC_BSSC"
    year = a.year or "2035"
    m = cable_hourly(scen, "Georgia", "Romania", year)
    qs = sorted(m["q"].unique())
    QC = {"Q1": "#1B6CA8", "Q2": "#36B5B5", "Q3": "#C0682A", "Q4": "#6A7BC8"}

    fs = 6.0
    rc(fs)
    fig, axes = plt.subplots(1, 2, figsize=(a.width, a.height), dpi=a.dpi,
                             gridspec_kw={"width_ratios": [1, 1.25]})
    ax = axes[0]
    for j, q in enumerate(qs):
        g = m[m["q"] == q]
        e, i = g["exp"].sum(), -g["imp"].sum()
        c = QC.get(q, "#9aa5b4")
        ax.bar(j, e, width=.62, zorder=2, facecolor=to_rgba(c, .88),
               edgecolor=to_rgba(c, .88), linewidth=.2)
        ax.bar(j, i, width=.62, zorder=2, facecolor=to_rgba(c, .26),
               edgecolor=c, linewidth=.45, hatch="///")
        ax.text(j, e + 60, "%.2f" % (e / 1e3), fontsize=fs - .6, color=INK,
                ha="center", va="bottom", zorder=4)
        ax.text(j, i - 60, "%.2f" % (-i / 1e3), fontsize=fs - .6, color=SOFT,
                ha="center", va="top", zorder=4)
    ax.axhline(0, color="#c9d3e0", linewidth=.7, zorder=3)
    ax.set_xticks(range(len(qs)))
    ax.set_xticklabels(qs, fontsize=fs, color=INK, fontweight="bold")
    ax.set_ylabel("TWh to Romania\nhatched = to Georgia", fontsize=fs,
                  labelpad=2)
    ax.set_yticks([-500, 0, 1000, 2000, 3000])
    ax.set_yticklabels(["-0.5", "0", "1", "2", "3"], fontsize=fs)
    ax.set_xlim(-.6, len(qs) - .4)
    ax.set_ylim(-750, 3300)
    ax.yaxis.grid(True, color=GRID, linewidth=.5, zorder=0)

    ax = axes[1]
    for q in qs:
        g = m[m["q"] == q]
        prof = [(gg["mw"] * gg["h"]).sum() / gg["h"].sum()
                for _, gg in g.groupby("t")]
        ax.plot(range(1, 25), prof, linewidth=1.1, color=QC.get(q, "#9aa5b4"),
                zorder=3, label=q)
    ax.axhline(0, color="#c9d3e0", linewidth=.7, zorder=2)
    ax.set_xticks([1, 6, 12, 18, 24])
    ax.set_xticklabels(["1h", "6h", "12h", "18h", "24h"], fontsize=fs)
    ax.set_ylabel("Mean flow, MW", fontsize=fs, labelpad=2)
    ax.set_xlim(1, 24)
    ax.set_ylim(-420, 1420)
    ax.yaxis.grid(True, color=GRID, linewidth=.5, zorder=0)
    ax.legend(frameon=False, fontsize=fs - .6, loc="lower left", ncol=4,
              handlelength=1.1, labelspacing=.25, columnspacing=.9,
              borderaxespad=.2)

    for x in axes:
        x.set_axisbelow(True)
        for s in ("top", "right"):
            x.spines[s].set_visible(False)
    fig.suptitle(a.title or "Submarine cable seasonality and daily shape, %s"
                 % year, fontsize=fs + 1, fontweight="bold", color=INK,
                 x=.012, y=.995, ha="left", va="top")
    fig.tight_layout(pad=.3, w_pad=1.4, rect=(.012, 0, .995, .90))
    return save(fig, a, "bssc_seasonality.png")


# ------------------------------------------------------------- expansion map

EXP_GREY = "#b9c2ce"
EXP_HOT = "#12356e"


def free_expansions(scen="LC_FreeExpAll", ref="baseline", year="2040",
                    floor=50.0):
    """Internal corridors the free-expansion run reinforced, MW, largest first.

    Measured against the reference run rather than against the corridor's own
    2025 capacity: TRIPP is committed in both, so its +800 MW is not a choice
    the model made here.  The floor drops the sub-MW numerical dust that a
    continuous expansion variable leaves on every link."""
    a, b = corridors(scen), corridors(ref)
    i = YEARS.index(str(year))
    out = []
    for k, c in a.items():
        if c["external"]:
            continue
        d = c["ntc"][i] - (b[k]["ntc"][i] if k in b else 0.0)
        if d >= floor:
            out.append((k, c, d))
    return sorted(out, key=lambda r: -r[2])


def chart_freeexp_expansion(a):
    """The regional grid in flat grey, with the reinforced links called out."""
    d = cache()
    geo = d["geo"]
    scen = a.scenario or "LC_FreeExpAll"
    ref = "baseline"
    year = a.year or "2040"
    i = YEARS.index(str(year))
    cor = corridors(scen)
    grown = free_expansions(scen, ref, year)
    hot = {k for k, _, _ in grown}
    fs = 6.0
    rc(fs)
    fig, ax = plt.subplots(figsize=(a.width, a.height), dpi=a.dpi)
    draw_base(ax, geo, fs)
    fit_extent(ax, geo, a.width - .1, a.height - .1)

    # Every link that exists at all, one flat grey stroke: this map is about
    # where the reinforcement lands, not about how hard anything is used.
    segs = []
    for key, c in cor.items():
        if c["a"] not in geo["centroids"] or c["b"] not in geo["centroids"]:
            continue
        if c["ntc"][i] <= 0 and c["fwd"][i] + c["rev"][i] <= 1e-3:
            continue
        pa, pb = geo["centroids"][c["a"]], geo["centroids"][c["b"]]
        if key in hot:
            continue
        segs.append((pa, pb))
        ax.plot([pa[0], pb[0]], [pa[1], pb[1]], color=EXP_GREY, linewidth=.9,
                solid_capstyle="round", zorder=4)
    for key, c, dmw in grown:
        pa, pb = geo["centroids"][c["a"]], geo["centroids"][c["b"]]
        ax.plot([pa[0], pb[0]], [pa[1], pb[1]], color="white", linewidth=3.4,
                solid_capstyle="round", zorder=5)
        ax.plot([pa[0], pb[0]], [pa[1], pb[1]], color=EXP_HOT, linewidth=1.9,
                solid_capstyle="round", zorder=6)
        segs.append((pa, pb))

    # Same placement search as the flow map: try both sides of the link at
    # growing distance and keep the slot that clears everything already on the
    # map, with a leader back to the link it belongs to.
    x0, y0, x1, y1 = geo["box"]
    boxes = []
    taken = [geo["centroids"][z] for z in geo["zones"]]
    for name in geo["ext"]:
        pt = geo["centroids"].get(name)
        if pt and x0 < pt[0] < x1 and y0 < pt[1] < y1:
            taken.append(pt)
    for key, c, dmw in grown:
        pa, pb = geo["centroids"][c["a"]], geo["centroids"][c["b"]]
        mid = ((pa[0] + pb[0]) / 2, (pa[1] + pb[1]) / 2)
        dx, dy = pb[0] - pa[0], pb[1] - pa[1]
        n = math.hypot(dx, dy) or 1.0
        nx, ny = -dy / n, dx / n
        best, best_gap = None, -1.0
        for step in range(7):
            off = .7 + .7 * step
            for sign in (1, -1):
                pos = (mid[0] + nx * off * sign, mid[1] + ny * off * sign)
                if not (x0 + .4 < pos[0] < x1 - .4
                        and y0 + .3 < pos[1] < y1 - .3):
                    continue
                gap = min([math.hypot(pos[0] - q[0], (pos[1] - q[1]) * 1.6)
                           for q in taken] or [99])
                if gap > best_gap:
                    best, best_gap = pos, gap
                if gap > 2.2:
                    break
            if best_gap > 2.2:
                break
        pos = best or mid
        taken.append(pos)
        ax.annotate("+%.0f MW" % dmw, xy=mid, xytext=pos, textcoords="data",
                    fontsize=fs - 1.1, color=EXP_HOT, fontweight="bold",
                    ha="center", va="center", zorder=9,
                    bbox=dict(boxstyle="round,pad=.16", fc="white",
                              ec="#c3ccd9", linewidth=.4, alpha=.96),
                    arrowprops=dict(arrowstyle="-", color="#7f8b9c",
                                    linewidth=.5, shrinkA=1, shrinkB=1))
        boxes.append(pos + _tbox("+%.0f MW" % dmw,
                                 ax.get_xlim()[1] - ax.get_xlim()[0],
                                 fs))

    # Last, so the zone names dodge the +MW callouts rather than the reverse:
    # the callouts are the point of this map.
    label_zones(ax, geo, fs, segs, boxes)

    h = [Line2D([], [], color=EXP_GREY, lw=1.1, label="Existing network"),
         Line2D([], [], color=EXP_HOT, lw=1.8,
                label="Reinforced under free expansion")]
    ax.legend(handles=h, loc="lower left", fontsize=fs - 1.0, frameon=False,
              handlelength=1.3, handletextpad=.5, borderpad=.2,
              borderaxespad=.3)
    if a.title:
        fig.suptitle(a.title, fontsize=fs + 1, fontweight="bold", color=INK,
                     x=.012, y=.995, ha="left", va="top")
    fig.tight_layout(pad=.2, rect=(0, 0, 1, .93 if a.title else .995))
    return save(fig, a, "freeexp_expansion_map.png")


CHARTS = {
    "region_maps": (chart_region_maps, 9.4, 2.5),
    "region_generation": (chart_region_generation, 5.6, 2.5),
    "bssc_map": (chart_bssc_map, 9.4, 3.2),
    "bssc_volume": (chart_bssc_volume, 5.86, 2.45),
    "bssc_mix_delta": (chart_bssc_mix_delta, 5.86, 2.45),
    "bssc_impact": (chart_bssc_impact, 5.6, 2.3),
    "freeexp_map": (chart_freeexp_map, 9.4, 3.2),
    "freeexp_expansion": (chart_freeexp_expansion, 6.4, 3.0),
    "freeexp_corridors": (chart_freeexp_corridors, 5.6, 2.3),
    "freeexp_build": (chart_freeexp_build, 9.4, 2.6),
    "benefit_regional": (chart_benefit_regional, 6.0, 3.1),
    "benefit_country": (chart_benefit_country, 6.0, 3.1),
    "allprojects_lines": (chart_allprojects_lines, 5.6, 2.6),
    "export_revenue": (chart_export_revenue, 5.86, 2.45),
    "bssc_seasonality": (chart_bssc_seasonality, 5.86, 2.45),
    "tables": (None, 5.6, 2.3),
}


def main():
    p = argparse.ArgumentParser()
    p.add_argument("--chart", default="all",
                   help="one of %s, or all" % ", ".join(CHARTS))
    p.add_argument("--scenario", default=None)
    p.add_argument("--year", default=None)
    p.add_argument("--years", default=None)
    p.add_argument("--top", type=int, default=10)
    p.add_argument("--width", type=float, default=None)
    p.add_argument("--height", type=float, default=None)
    p.add_argument("--dpi", type=int, default=300)
    p.add_argument("--title", default=None)
    p.add_argument("--out", default=None)
    a = p.parse_args()

    every = a.chart == "all"
    for n in (list(CHARTS) if every else [a.chart]):
        fn, w, h = CHARTS[n]
        if every:                                      # per-chart slide sizes
            a.width, a.height, a.out = w, h, None
        else:
            a.width = a.width or w
            a.height = a.height or h
        if n == "tables":
            table_bssc(a)
            table_freeexp(a)
            table_benefits(a)
            table_benefits_xlsx(a)
            table_levels_xlsx(a)
            table_allprojects(a)
            continue
        fn(a)
        plt.close("all")


if __name__ == "__main__":
    main()
