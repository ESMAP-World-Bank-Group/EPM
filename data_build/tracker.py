# -*- coding: utf-8 -*-
"""Data tracking workbook, DERIVED from the build.

    python data_build/tracker.py --config data_build/build_casa.yaml

The workbook is not a source of information: it is a view. Everything it shows
comes from build_report.json, hence from what the build actually does, plus the
"sources" and "leads" blocks of the YAML. It therefore cannot drift away from the
data, which was the flaw of the hand-kept tracker on Black Sea.

Runs the build in --check mode if the report is missing or older than the YAML.
"""

import argparse
import json
import os
import re
import subprocess
import sys

import yaml
import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

HERE = os.path.dirname(os.path.abspath(__file__))

# Colour code taken from Black Sea: green done, amber partial, red to do, grey out of scope.
FILLS = {"G": PatternFill("solid", fgColor="C6EFCE"),
         "Y": PatternFill("solid", fgColor="FFEB9C"),
         "R": PatternFill("solid", fgColor="FFC7CE"),
         "B": PatternFill("solid", fgColor="D9D9D9")}
HDR = PatternFill("solid", fgColor="1F4E79")
SUB = PatternFill("solid", fgColor="D6E4F7")
PRIO = {"P1": PatternFill("solid", fgColor="FCE4D6"),
        "P2": PatternFill("solid", fgColor="FFF2CC"),
        "P3": PatternFill("solid", fgColor="F2F2F2")}
WHITE = Font(color="FFFFFF", bold=True, size=10)
BOLD = Font(bold=True, size=9)
SMALL = Font(size=9)
TOP = Alignment(vertical="top", wrap_text=True)
THIN = Border(*[Side(style="thin", color="BFBFBF")] * 4)


def remaining(entry):
    """What is left to do, or an empty string when the todo is a status report."""
    txt = str(entry.get("todo", "")).strip()
    return "" if txt.upper().startswith("DONE") else txt


def covers(source):
    """Countries a source can speak for, as text. The YAML holds a list of country
    codes, so nothing has to be re-typed when the perimeter changes."""
    c = source.get("covers")
    if isinstance(c, (list, tuple)):
        return ", ".join(str(x) for x in c)
    return str(c or "")


def state_of(entry):
    """Progress of a resource, derived from the build alone.

    Two facts are enough. Did the build transform the file (action), and is there
    any declared work left (todo)? Nothing is typed in by hand here, so nothing can
    stay green by oversight.
    """
    if entry.get("work") == "SKIP":
        return "B", "out of scope"
    touched = entry.get("action") not in ("keep", "shared")
    # A todo starting with DONE is not a leftover: it is the record of what was
    # done, kept in the YAML next to the decision it explains.
    todo = bool(remaining(entry))
    if touched and not todo:
        return "G", "done"
    if touched:
        # DONE = the data is in place; any remaining todo is an addition (scenario
        # variant, better source expected), not a gap.
        if entry.get("work") == "DONE":
            return "Y", "filled, more to come"
        return "Y", "structure adapted, data to do"
    if not todo:
        return "G", "inherited, fit for purpose"
    return "R", "inherited 2020, to process"


def phase_of(entry):
    """Phase number, read at the start of the todo. Convention: 'PHASE 3. ...'."""
    m = re.match(r"\s*PHASE\s+(\d+)", str(entry.get("todo", "")), re.I)
    return "P" + m.group(1) if m else ""


def sheet(wb, title, widths, headers, first=False):
    ws = wb.active if first else wb.create_sheet()
    ws.title = title
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=i, value=h)
        c.fill, c.font, c.alignment, c.border = HDR, WHITE, TOP, THIN
    ws.freeze_panes = "A2"
    return ws


def put(ws, row, values, fill=None, font=SMALL):
    for i, v in enumerate(values, 1):
        # folded YAML scalars keep a trailing newline
        c = ws.cell(row=row, column=i, value=v.strip() if isinstance(v, str) else v)
        c.alignment, c.border, c.font = TOP, THIN, font
        if fill:
            c.fill = fill
    return row + 1


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--config", required=True)
    ap.add_argument("--out", default=os.path.join(HERE, "DATA_SOURCES_casa.xlsx"))
    args = ap.parse_args()

    cfg = yaml.safe_load(open(args.config, "r", encoding="utf-8"))
    report_path = os.path.join(HERE, "build_report.json")

    # The report must mirror the current YAML, otherwise re-run the build dry.
    if (not os.path.isfile(report_path)
            or os.path.getmtime(report_path) < os.path.getmtime(args.config)):
        print("Report missing or stale, re-running the build with --check.")
        subprocess.check_call([sys.executable, os.path.join(HERE, "build.py"),
                               "--config", args.config, "--check"])

    report = json.load(open(report_path, "r", encoding="utf-8"))
    sources = report.get("sources") or {}
    rows = report.get("resources") or []

    wb = openpyxl.Workbook()

    # ── Sheet 1: tracking ──────────────────────────────────────────────────
    ws = sheet(wb, "Tracking",
               [13, 22, 30, 34, 9, 8, 26, 13, 6, 20, 20, 11, 60, 70],
               ["Block", "Resource", "File", "Inherited content", "Rows", "Phase",
                "State", "Work", "Prio", "Sources", "Source sheet", "Vintage",
                "What was done", "What is left"], first=True)
    r, block, tally = 2, None, {"G": 0, "Y": 0, "R": 0, "B": 0}
    for e in rows:
        if e.get("group") != block:
            block = e.get("group")
            c = ws.cell(row=r, column=1, value=block or "(unclassified)")
            c.fill, c.font, c.alignment, c.border = SUB, Font(bold=True, size=9), TOP, THIN
            for i in range(2, 15):
                ws.cell(row=r, column=i).fill = SUB
            r += 1
        code, label = state_of(e)
        tally[code] += 1
        n = e.get("rows_out")
        r = put(ws, r, [block, e["resource"], e.get("path", ""), e.get("what", ""),
                        n if n != "-" else "-", phase_of(e), label,
                        e.get("work", ""), e.get("priority", ""),
                        ", ".join(e.get("source") or []),
                        e.get("sheet", ""), str(e.get("vintage", "") or ""),
                        e.get("note", ""), e.get("todo", "")])
        for i in (7,):
            ws.cell(row=r - 1, column=i).fill = FILLS[code]
        p = PRIO.get(e.get("priority"))
        if p:
            ws.cell(row=r - 1, column=9).fill = p

    r += 1
    put(ws, r, ["SUMMARY", "{} done | {} partial | {} to do | {} out of scope"
                .format(tally["G"], tally["Y"], tally["R"], tally["B"])], font=BOLD)

    # ── Sheet 2: sources ───────────────────────────────────────────────────
    ws = sheet(wb, "Sources", [18, 46, 12, 13, 20, 30, 34, 80],
               ["Key", "Source", "Date", "Grade", "Access", "Covers",
                "Where to find it", "What it contains"])
    r = 2
    for key, s in sources.items():
        # Grade travels with the date: it is what stops a placeholder dated of today
        # from reading as fresh data. Same field the documentation page grades on.
        r = put(ws, r, [key, s.get("name", ""), str(s.get("date", "")),
                        s.get("grade", ""), s.get("access", ""), covers(s),
                        s.get("where", ""), s.get("note", "")])

    # ── Sheet 3: leads ─────────────────────────────────────────────────────
    ws = sheet(wb, "Leads", [14, 34, 30, 20, 62, 70],
               ["Country", "Organisation", "Address", "Status", "What was found",
                "What we do with it"])
    r = 2
    # The status reuses the tracking colour code: tested and available is green,
    # tested and empty is red. An untested lead stays grey.
    STAT = {"AVAILABLE": "G", "PARTIAL": "Y", "NOTHING": "R"}
    for lead in (cfg.get("leads") or []):
        r = put(ws, r, list(lead))
        status = str(lead[3]) if len(lead) > 3 else ""
        code = next((c for k, c in STAT.items() if k in status), "B")
        ws.cell(row=r - 1, column=4).fill = FILLS[code]

    wb.save(args.out)
    print("Written: {}".format(args.out))
    print("{} done | {} partial | {} to do | {} out of scope"
          .format(tally["G"], tally["Y"], tally["R"], tally["B"]))


if __name__ == "__main__":
    main()
