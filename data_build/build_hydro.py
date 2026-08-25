# -*- coding: utf-8 -*-
"""Rebuild the seasonal availability of the fleet, hydro plant by hydro plant.

    python data_build/build_hydro.py

Reads:
    mappings/seasons_months.csv   the season set and the months behind it
    mappings/hydro_plants.csv     one line per hydro unit of the model, hand-edited
    mappings/availability.csv     the outage rate of a new thermal (tech, fuel) pair
    extracted/pGenDataInput.csv   the fleet, to see whom the table is missing
    extracted/thermal_derating.csv  the DeCA derating of an existing thermal unit
    <reference>/supply/pAvailabilityCustom.csv   the 2020 values, used as the base
    the DeCA AssumptionBooks                     sheets HydroPP and Hydro Profile

RUN ORDER IS build_fleet.py, THEN build_costs.py, THEN build_hydro.py. The fleet is
read here to catch the units the availability table does not cover, and the derating
this script applies is the one build_costs.py extracts on its way through the DeCA
thermal sheets.

Writes:
    extracted/pAvailabilityCustom.csv        taken as is by the build
    extracted/pAvailabilityDefault.csv       idem
    extracted/hydro_availability_report.csv  what each unit was given, and why

THE THERMAL SIDE IS A FOURTH METHOD, deca_outage, and it asks a different question of
an existing unit than of one still to be built.

A CANDIDATE is written at the pair rate of availability.csv, 0.90 across the seasons.
That is DeCA's own arithmetic for a plant in working order: it gives the same two
outage rates to every thermal unit of the five countries it covers, 5% forced and 5%
scheduled. A plant that does not exist yet has no history of dereliction, so the pair
rate is all there is to say about it, and it is what the two Kyrgyz coal candidates
get in place of the 1 the 2020 table gave them -- a fifth of the year of outage simply
missing -- and what NEPS_HERAT_ST_Coal_cand gets in place of 0.0329 in Q1 alone, which
is a typo and not an assumption.

AN EXISTING UNIT IS WRITTEN AT ITS OWN DERATING, because DeCA states an Available
capacity beside the Installed capacity of every plant it carries and the gap between
the two is the plant's condition, unit by unit. thermal_derating.csv holds the ratio,
already multiplied by the 0.90: 0.4655 for Bishkek, whose 812 MW book capacity is
available at 420, against 0.90 for the second Dushanbe set. Ignoring it would have
told the model that a Soviet CHP nobody has maintained since 1991 and a 2018 combined
cycle are the same machine. The 2020 model wrote 0.85 everywhere and named no source;
BISHKEK_CHP_2023 alone sat at 0.3085, which was the right instinct -- a real derating
hidden in the availability column -- and lands at 0.4655 once DeCA states it, so 310
MW effective out of the 666 the model carries rather than 205.

OUTSIDE THE FIVE COUNTRIES NOTHING MOVES. Pakistan and Afghanistan keep their 2020
availability, because the rule of this build is that DeCA wins where DeCA and 2020
disagree and on a Pakistani plant DeCA says nothing at all. Carrying the Central Asian
median south would be inventing a number for a fleet no source in hand describes, and
inventing it in the one place it would move the answer: Pakistan is the demand this
study asks whether Central Asian power is worth exporting to. The median is still
computed and written into thermal_derating.csv for the units inside the perimeter that
the mapping could not match to a book plant, where the country is described and only
the plant is not.

WHAT THE MAPPING DELIBERATELY DOES NOT NAME is hydro, solar, wind, imports and
storage. For those the column is not an outage rate but the resource itself -- the
seasonal water of a reservoir, a profile that already carries the weather -- and a
plant-level number must not be overwritten by a technology-level one.

Three methods, declared per unit in the mapping:

    deca_profile  the season is the mean of the DeCA monthly utilisation factors over
                  its months, weighted by the length of each month. Used where DeCA
                  carries a profile of its own for the plant and not the national
                  average.
    deca_level    the 2020 seasonal shape is kept and rescaled so that the annual mean
                  equals the DeCA capacity factor, yearly production over installed
                  capacity. Reservoir operation is a modelling choice that DeCA does
                  not describe; the energy of a year is a measurement, and that is what
                  is taken.
    keep_2020     untouched.

A rescaling that would push a season above 1 is clipped and reported: availability is
a share of the capacity, it cannot exceed it.

THE TABLE IS THEN COMPLETED AGAINST THE FLEET, and that is not a formality. The 2020
table covered its 403 units exactly; the 16 solar and wind candidates this build adds
from 2031 had no line at all, and a generator absent from pAvailability is not given a
default of one, it is given ZERO: base.gms:812 then writes vPwrOut = 0 for the whole
horizon, so the unit can be built and produces nothing. Solar happened to be rescued by
the generic fallback of the model, which carries a PV/Solar line; wind was not, its
technology being written WT here and OnshoreWind there, so EIGHT WIND CANDIDATES WERE
SILENTLY DEAD. A missing unit is completed with the value its own technology and fuel
already carry elsewhere in the table, and the step refuses to guess when that value is
not unanimous.
"""

import argparse
import collections
import csv
import io
import os

HERE = os.path.dirname(os.path.abspath(__file__))
REPO = os.path.dirname(HERE)

MONTH_DAYS = [31, 28, 31, 30, 31, 30, 31, 31, 30, 31, 30, 31]
MONTH_NAMES = ["January", "February", "March", "April", "May", "June", "July",
               "August", "September", "October", "November", "December"]
BOOK = "20250306_DeCA - AssumptionBook - {}_V5.1_Clean.xlsx"


def read_csv(path):
    with io.open(path, encoding="utf-8-sig", newline="") as fh:
        rd = csv.reader(fh)
        return next(rd), [r for r in rd if any(c.strip() for c in r)]


def write_csv(path, header, rows):
    with io.open(path, "w", encoding="utf-8", newline="") as fh:
        w = csv.writer(fh, lineterminator="\n")
        w.writerow(header)
        w.writerows(rows)


def num(x):
    """A number, or None. Empty cells and text both come back as None."""
    try:
        return float(str(x).strip())
    except (TypeError, ValueError):
        return None


def seasons(path):
    """[(season, months, hours)] read from the mapping, in file order."""
    header, rows = read_csv(path)
    i = {name.strip(): j for j, name in enumerate(header)}
    out = []
    for r in rows:
        months = [int(m) for m in r[i["months"]].split(",") if m.strip()]
        out.append((r[i["season"]].strip(), months,
                    24.0 * sum(MONTH_DAYS[m - 1] for m in months)))
    total = sum(h for _, _, h in out)
    if abs(total - 8760.0) > 24.0:
        raise ValueError("the seasons cover {:.0f} h, not a year".format(total))
    return out


def legacy_of(path):
    """season -> the season of the 2020 model it inherits its value from.

    The two sets overlap in name without overlapping in meaning: Q1, Q2 and Q4 stand
    for the same months in both, but the 2020 Q3 was one five-month summer where ours
    is two, Q3a and Q3b, and the 2020 Q5 was a 130 h peak block that has no successor
    here at all. The mapping says which feeds which so that neither set has to match
    the other.
    """
    header, rows = read_csv(path)
    i = {name.strip(): j for j, name in enumerate(header)}
    return dict((r[i["season"]].strip(), r[i["legacy"]].strip()) for r in rows)


def header_row(rows, label):
    """Index of the row that holds the column names, found by one of them."""
    for i, r in enumerate(rows):
        if any(isinstance(c, str) and c.strip() == label for c in r):
            return i
    raise KeyError("no header row carrying " + label)


def columns(row):
    """Column name -> index, first occurrence wins.

    The month names are repeated once per hydrology scenario, so the last occurrence
    would silently read a scenario nobody asked for.
    """
    out = {}
    for j, c in enumerate(row):
        if isinstance(c, str) and c.strip() and c.strip() not in out:
            out[c.strip()] = j
    return out


def scenario_column(rows, h, label):
    """Column where the wanted scenario block starts, above the header row."""
    for i in (h - 1, h - 2, h - 3):
        if i < 0:
            continue
        for j, c in enumerate(rows[i]):
            if isinstance(c, str) and c.strip() == label:
                return j
    return None


def read_book(path, scenario="Reference"):
    """Plant name -> capacity, yearly production, monthly profile."""
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    plants = {}

    rows = list(wb["HydroPP"].iter_rows(values_only=True))
    h = header_row(rows, "Country")
    col = columns(rows[h])
    for r in rows[h + 1:]:
        name = r[col["Name"]] if "Name" in col else None
        if not isinstance(name, str) or not name.strip():
            continue
        cap = num(r[col["Capacity (MW)"]]) if "Capacity (MW)" in col else None
        gwh = num(r[col["Yearly production [GWh]"]]) if "Yearly production [GWh]" in col else None
        plants[name.strip()] = {"cap": cap, "gwh": gwh, "profile": None}

    rows = list(wb["Hydro Profile"].iter_rows(values_only=True))
    h = header_row(rows, "Country")
    col = columns(rows[h])
    jan = scenario_column(rows, h, scenario)
    if jan is None:
        jan = col[MONTH_NAMES[0]]
    if not all(isinstance(rows[h][jan + k], str)
               and rows[h][jan + k].strip() == MONTH_NAMES[k] for k in range(12)):
        raise ValueError("the twelve months do not follow the scenario column in "
                         + os.path.basename(path))
    for r in rows[h + 1:]:
        name = r[col["Name"]]
        if not isinstance(name, str) or not name.strip():
            continue
        vals = [num(r[jan + k]) for k in range(12)]
        if any(v is None for v in vals):
            continue
        plants.setdefault(name.strip(), {"cap": None, "gwh": None, "profile": None})
        plants[name.strip()]["profile"] = vals
    wb.close()
    return plants


def to_seasons(monthly, season_set):
    """Monthly values -> one value per season, weighted by the length of the months."""
    out = []
    for _, months, _ in season_set:
        days = sum(MONTH_DAYS[m - 1] for m in months)
        out.append(sum(monthly[m - 1] * MONTH_DAYS[m - 1] for m in months) / days)
    return out


def annual(values, season_set):
    """Hour-weighted mean of a seasonal series."""
    hours = [h for _, _, h in season_set]
    return sum(v * h for v, h in zip(values, hours)) / sum(hours)


def for_pair(table, tech, fuel):
    """The row that speaks for this pair, the exact one first, then the wildcard."""
    return table.get((tech, fuel)) or table.get((tech, "*"))


def fleet_pairs(fleet_path):
    """(tech, fuel, zone, status) of every unit of the fleet, by generator name."""
    header, gens = read_csv(fleet_path)
    col = {c.strip(): j for j, c in enumerate(header)}
    return dict((r[0].strip(), (r[col["tech"]].strip(), r[col["fuel"]].strip(),
                                r[col["z"]].strip(), r[col["Status"]].strip()))
                for r in gens)


def complete(out, names, fleet_path, report):
    """The lines the fleet needs and the availability table does not have.

    A unit missing here is a unit at zero, not a unit at one, so the gap has to be
    closed rather than reported. The value is the one its own (tech, fuel) already
    carries in the table, by majority: the 24 solar units of the model are 22 at one
    in every season and 2 derated for a partial first year, and it is the 22 that
    describe the technology. A tie is refused rather than settled by file order, and
    so is a pair no unit of which is in the table, which is how a hydro candidate
    would be caught: reservoir operation is plant data, nothing about it could
    honestly be copied from a neighbour.
    """
    header, gens = read_csv(fleet_path)
    col = {c.strip(): j for j, c in enumerate(header)}
    have = set(r[0] for r in out)
    known = {}
    for r in out:
        g = r[0]
        for x in gens:
            if x[0].strip() == g:
                known.setdefault((x[col["tech"]].strip(), x[col["fuel"]].strip()),
                                 collections.Counter())[tuple(r[1:])] += 1
                break

    added = []
    for x in gens:
        g = x[0].strip()
        if g in have:
            continue
        key = (x[col["tech"]].strip(), x[col["fuel"]].strip())
        seen = known.get(key)
        if not seen:
            raise KeyError("no availability anywhere for " + " ".join(key)
                           + ", needed by " + g)
        top = seen.most_common(2)
        if len(top) > 1 and top[0][1] == top[1][1]:
            raise ValueError("the units of " + " ".join(key) + " are split evenly "
                             "between two availabilities, so " + g + " cannot be "
                             "completed")
        values = list(top[0][0])
        added.append([g] + values)
        report.append([g, x[col["z"]].strip(), "", "", "completed", "fleet_default",
                       "", "", ",".join(values),
                       "absent from the 2020 table, given what {0} of the {1} "
                       "{2} units carry".format(top[0][1], sum(seen.values()),
                                                " ".join(key))])
    return added


def main():
    ap = argparse.ArgumentParser(description="Rebuild the seasonal availability.")
    ap.add_argument("--reference", default=os.path.join("epm", "input", "data_casa_2020"))
    ap.add_argument("--books", default=os.path.join("..", "data_collection", "Mercados"))
    ap.add_argument("--scenario", default="Reference",
                    help="hydrology block of the Hydro Profile sheet")
    args = ap.parse_args()

    maps = os.path.join(HERE, "mappings")
    season_set = seasons(os.path.join(maps, "seasons_months.csv"))
    names = [q for q, _, _ in season_set]
    inherits = legacy_of(os.path.join(maps, "seasons_months.csv"))

    ref_path = os.path.join(REPO, args.reference, "supply", "pAvailabilityCustom.csv")
    ref_header, ref_rows = read_csv(ref_path)
    ref_col = {c.strip(): j for j, c in enumerate(ref_header)}
    missing = [inherits[q] for q in names if inherits[q] not in ref_col]
    if missing:
        raise KeyError("the reference has no column for " + ", ".join(missing))
    base, order = {}, []
    for r in ref_rows:
        g = r[ref_col["g"]].strip()
        order.append(g)
        base[g] = [num(r[ref_col[inherits[q]]]) or 0.0 for q in names]

    fleet_path = os.path.join(HERE, "extracted", "pGenDataInput.csv")
    fleet = fleet_pairs(fleet_path)
    aheader, arows = read_csv(os.path.join(maps, "availability.csv"))
    acol = {c.strip(): j for j, c in enumerate(aheader)}
    outage = dict(((r[acol["tech"]].strip(), r[acol["fuel"]].strip()),
                   r[acol["availability"]].strip()) for r in arows)
    dheader, drows = read_csv(os.path.join(HERE, "extracted",
                                           "thermal_derating.csv"))
    dcol = {c.strip(): j for j, c in enumerate(dheader)}
    derating = dict((r[0].strip(), (r[dcol["availability"]].strip(),
                                    r[dcol["source"]].strip())) for r in drows)

    header, rows = read_csv(os.path.join(maps, "hydro_plants.csv"))
    col = {c.strip(): j for j, c in enumerate(header)}
    plan = {}
    for r in rows:
        plan[r[col["g"]].strip()] = dict((k, r[col[k]].strip())
                                         for k in ("z", "book", "deca_plant", "method"))

    books = {}
    for cc in sorted(set(p["book"] for p in plan.values() if p["book"])):
        path = os.path.join(REPO, args.books, BOOK.format(cc))
        if not os.path.isfile(path):
            raise IOError("AssumptionBook missing: " + path)
        books[cc] = read_book(path, args.scenario)

    out, report = [], []
    for g in order:
        values = list(base.get(g, [0.0] * len(names)))
        p = plan.get(g)
        method = p["method"] if p else "not_hydro"
        deca = None
        if p and p["book"]:
            deca = books.get(p["book"], {}).get(p["deca_plant"])
        cf_ref = annual(values, season_set)
        cf_deca, applied, why = "", method, ""

        if method == "deca_profile":
            if deca and deca["profile"]:
                values = to_seasons(deca["profile"], season_set)
            else:
                applied, why = "keep_2020", "no monthly profile in DeCA for this plant"
        elif method == "deca_level":
            if deca and deca["cap"] and deca["gwh"]:
                cf = deca["gwh"] * 1000.0 / (deca["cap"] * 8760.0)
                cf_deca = round(cf, 4)
                if cf > 1.0:
                    applied, why = "keep_2020", "the DeCA capacity factor exceeds 1"
                elif cf_ref <= 0:
                    applied, why = "keep_2020", "the 2020 availability is zero"
                else:
                    values = [v * cf / cf_ref for v in values]
            else:
                applied, why = "keep_2020", "DeCA gives no capacity or no yearly production"

        rate, said = None, ""
        if p is None:
            tech, fuel, zone, status = fleet.get(g, ("", "", "", ""))
            pair = for_pair(outage, tech, fuel)
            if pair is None:
                # Hydro, solar, wind, imports, storage: the column is not an outage
                # rate there but the resource itself, and must not be overwritten.
                pass
            elif status != "1":
                rate = pair
                said = "the {0} rate of a new {1} on {2}".format(pair, tech, fuel)
            elif g in derating:
                rate, source = derating[g]
                said = "the derating of {0}, {1} available".format(source, rate)
            # An existing unit outside the DeCA perimeter falls through and keeps its
            # 2020 value, which is the whole of the Pakistani and Afghan fleet.
        if rate is not None:
            was = list(values)
            values = [float(rate)] * len(names)
            applied = "deca_outage"
            if ["{0:.6g}".format(v) for v in was] != ["{0:.6g}".format(v)
                                                      for v in values]:
                why = "was {0}, now {1}".format(
                    ", ".join("{0:.6g}".format(v) for v in was), said)
                report.append([g, zone, "", "", "thermal", applied,
                               "{0:.4f}".format(cf_ref), rate,
                               "{0:.4f}".format(float(rate)), why])

        clipped = [min(v, 1.0) for v in values]
        if clipped != values:
            why = (why + "; " if why else "") + "a season was clipped at 1"
        values = clipped

        out.append([g] + ["{0:.6g}".format(v) for v in values])
        if p:
            report.append([g, p["z"], p["book"], p["deca_plant"], method, applied,
                           "{0:.4f}".format(cf_ref), cf_deca,
                           "{0:.4f}".format(annual(values, season_set)), why])

    out += complete(out, names, fleet_path, report)

    # The default table is read only for a (zone, tech, fuel) the custom table does
    # not cover unit by unit, which today is none of them: input_treatment fills
    # pAvailability from the custom rows first and reaches for the default only for
    # what is left. It is written all the same, for the pairs whose availability is a
    # technology property and not plant data, so that a generator added later without
    # a line of its own lands on a stated number instead of on zero. It also carried
    # the two-season header of data_test, which no longer matches this model.
    zones = sorted(set((z, tech, fuel) for tech, fuel, z, _s in fleet.values()
                       if for_pair(outage, tech, fuel) is not None))
    write_csv(os.path.join(HERE, "extracted", "pAvailabilityDefault.csv"),
              ["z", "tech", "fuel"] + names,
              [[z, tech, fuel] + [for_pair(outage, tech, fuel)] * len(names)
               for z, tech, fuel in zones])

    write_csv(os.path.join(HERE, "extracted", "pAvailabilityCustom.csv"),
              ["g"] + names, out)
    write_csv(os.path.join(HERE, "extracted", "hydro_availability_report.csv"),
              ["g", "z", "book", "deca_plant", "method", "applied",
               "annual_2020", "annual_deca", "annual_new", "note"], report)

    thermal = [r for r in report if r[5] == "deca_outage"]
    hydro = [r for r in report if r[5] not in ("fleet_default", "deca_outage")]
    done = sum(1 for r in hydro if r[5] != "keep_2020")
    print("units      {0}, of which completed against the fleet {1}".format(
        len(out), sum(1 for r in report if r[5] == "fleet_default")))
    print("hydro      {0}, of which rebuilt from DeCA {1}".format(len(hydro), done))
    own = sum(1 for r in thermal if "derating" in r[9])
    print("thermal    {0} unit(s) moved, {1} on their own DeCA derating and {2} on the "
          "{3} pair rate(s)".format(len(thermal), own, len(thermal) - own,
                                    len(outage)))
    print("           {0} default row(s), {1} existing unit(s) left on 2020".format(
        len(zones), sum(1 for g, (t, f, z, st) in fleet.items()
                        if st == "1" and for_pair(outage, t, f) is not None
                        and g not in derating)))
    print("seasons    " + ", ".join("{0} ({1:.0f} h)".format(q, h)
                                    for q, _, h in season_set))


if __name__ == "__main__":
    main()
