# -*- coding: utf-8 -*-
"""The solar and wind profiles, on the time structure of build_demand.py.

    python data_build/build_vre.py

Reads:
    mappings/vre_profiles.csv              which DeCA rows each (zone, technology) reads
    mappings/seasons_months.csv            which months make each season
    extracted/pHours.csv                   the time structure, for its days
    extracted/demand_report.csv            which calendar day each block stands for
    <hourly>/PV_casa.csv, WT_casa.csv      the rescaled hourly year, one day at a time
    <reference>/pHours.csv                 the 2020 structure, for its block widths
    <reference>/supply/pVREProfile.csv     the fallback intra-day shape
    the five DeCA AssumptionBooks, sheet RenewableProfile

Writes:
    extracted/pVREProfile.csv
    extracted/vre_report.csv

RUN AFTER build_demand.py, which writes the pHours and the representative dates this
reads, and AFTER build_vre_hourly.py, which writes the rescaled hourly year.

WHY IT IS BUILT FROM TWO SOURCES. base.gms:812 writes the output of every solar and
wind unit as pVREgenProfile times the availability times the capacity, so this table
is the whole production of the renewable fleet, in level as well as in shape. Neither
source gives both.
  - DeCA gives the LEVEL and not the shape: RenewableProfile states a monthly
    utilization factor per plant, no hour of the day. Its own header says so, "Hourly
    profiles are also needed for all WPP and SPP power plants", a line the authors
    left for themselves.
  - The 2020 model gives the SHAPE and a level that is now five years old: its blocks
    are chronological hour-of-day groups, verifiable on the file itself, solar at zero
    in the first and last blocks and highest in the middle one.
A third source now settles the shape where it can. build_vre_hourly.py holds a
Renewables.ninja year per zone already brought onto the DeCA level, and
extracted/demand_report.csv records which calendar day each representative block stands
for. So for a zone that has one, the block is READ OFF THAT DAY: 12 December for the
winter peak day, 8 September for the second late-summer day, the weather that actually
happened on the day the load was measured on. Where no hourly series survived the
rescale, the day shape falls back to the 2020 profile normalised to a mean of one. Either
way the level is the DeCA seasonal utilization factor, the month-length weighted mean of
the monthly factors of the season over the plants the mapping selects.

WHAT THE MAPPING SAYS. One line per (zone, technology): which workbook to read and
which rows of it, by a list of substrings matched on the DeCA name. The five books name
no zone for their existing plants, so the factor is a COUNTRY MEAN over every plant of
that technology, candidates included, high yield and low yield sites together. That is
an average site, which is what a generic candidate is. A zone whose line names no
workbook keeps its 2020 level unchanged: Afghanistan and Pakistan are outside the DeCA
perimeter.
One trap the mapping steps around: the Tajik book labels its own regions with a KZ
prefix, Khatlon and Sougd and Gorno-Badakhshan all reading "KZ ...". Matching on the
workbook rather than on the name is what keeps Tajik sites out of Kazakhstan.

WHAT IT STILL DOES NOT DO, and it is now a shorter list than it was. Where the hourly
year is used, the three representative days of a season differ from each other and a
cloudy one can appear; where it is not, every representative day of the season carries
the same renewable day and there is no cloudy day and no calm day at all, only an average
one. The second case is what ten wind series fall back to, all of them in mountain zones
where MERRA-2 cannot see the terrain, and it understates the backup a large wind fleet
would need there. vre_report.csv names them in its shape column.

The days themselves are also chosen on the LOAD alone: build_demand.py picked the peak
day and the median days of each group without looking at the weather, so a still evening
of high demand is in the set only by luck. Choosing days on load and renewables jointly
is what the Poncelet clustering in pre-analysis/representative_days does, and it waits on
the Pakistani and Afghan hourly load that does not exist yet.

And the years do not line up everywhere. The load years behind the representative dates
are DeCA vintages, Kazakhstan 2022 and Tajikistan 2021 with the others unstated, against
a 2022 weather year. For Kazakhstan the pairing is a true one, same day of the same year;
elsewhere it is the same date of a different year, which is seasonally right and
synoptically a coincidence.
"""

import argparse
import collections
import csv
import glob
import io
import os

import openpyxl

HERE = os.path.dirname(os.path.abspath(__file__))
REPO = os.path.dirname(HERE)

BOOKS = r"C:/Users/wb590892/Documents/EPM_Models/ca_2026/data_collection/Mercados"
MONTH_LENGTH = [31, 28, 31, 30, 31, 30, 31, 31, 30, 31, 30, 31]
MONTHS = ["January", "February", "March", "April", "May", "June", "July", "August",
          "September", "October", "November", "December"]
HOURS_PER_DAY = 24
HOURLY = os.path.join(REPO, "pre-analysis", "representative_days", "input")

# The furthest the three representative days of a season may be moved, either way, to put
# the season back on its stated level. They are three real days standing for sixty, chosen
# on the load and not on the weather, so their own mean is not the season's and a
# correction is expected. A LARGE one means they are not a sample of that season at all:
# northern Kazakhstan needs 2.3 to lift its winter days to the stated solar factor, which
# says the days it drew were overcast ones, and an overcast day multiplied by 2.3 is a
# clear day with the wrong curvature rather than a better winter. The bound is symmetric
# because three unusually windy days are no more a season than three unusually calm ones.
SEASON_STRETCH = 1.5


def read_csv(path):
    with io.open(path, encoding="utf-8-sig", newline="") as fh:
        rd = csv.reader(fh)
        return next(rd), [r for r in rd if any(c.strip() for c in r)]


def write_csv(path, header, rows):
    with io.open(path, "w", encoding="utf-8", newline="") as fh:
        w = csv.writer(fh, lineterminator="\n")
        w.writerow(header)
        w.writerows(rows)


def dicts(path):
    header, rows = read_csv(path)
    keys = [h.strip() for h in header]
    return [dict(zip(keys, r)) for r in rows]


def num(x):
    try:
        return float(str(x).strip())
    except (TypeError, ValueError):
        return None


def seasons(path):
    out = collections.OrderedDict()
    for row in dicts(path):
        out[row["season"].strip()] = [int(m) for m in row["months"].split(",")]
    return out


def legacy_of(path):
    """season -> the season of the 2020 model it inherits its shape from.

    The two sets overlap in name without overlapping in meaning: Q1, Q2 and Q4 stand
    for the same months in both, but the 2020 Q3 was one five-month summer where ours
    is two, Q3a and Q3b, and the 2020 Q5 was a 130 h peak block that has no successor
    here at all. The mapping says which feeds which so that neither set has to match
    the other.
    """
    out = {}
    for row in dicts(path):
        out[row["season"].strip()] = row["legacy"].strip()
    return out


def block_hours(row):
    """The width in hours of each block of the 2020 day. See build_demand.py."""
    values = [v for v in row if v is not None]
    days = sum(values) / float(HOURS_PER_DAY)
    exact = [v / days for v in values]
    base = [int(v) for v in exact]
    order = sorted(range(len(exact)), key=lambda i: exact[i] - base[i], reverse=True)
    for i in order[:HOURS_PER_DAY - sum(base)]:
        base[i] += 1
    return base


def legacy_profile(reference):
    """(zone, tech, season) -> [24 factors], the 2020 profile spread over the day."""
    widths = {}
    _, rows = read_csv(os.path.join(reference, "pHours.csv"))
    for r in rows:
        widths[r[0].strip()] = block_hours([num(c) for c in r[2:]])

    out = {}
    header, rows = read_csv(os.path.join(reference, "supply", "pVREProfile.csv"))
    for r in rows:
        z, tech, q = r[0].strip(), r[1].strip(), r[2].strip()
        if q not in widths:
            continue
        day = []
        for width, value in zip(widths[q], [num(c) or 0.0 for c in r[4:]]):
            day += [value] * width
        out[(z, tech, q)] = day
    return out


def monthly_factors(workbook_code):
    """(name, [12 utilization factors]) of every row of the RenewableProfile sheet."""
    pattern = os.path.join(BOOKS, "*- {0}_V5.1_Clean.xlsx".format(workbook_code))
    matches = glob.glob(pattern)
    if not matches:
        raise IOError("no AssumptionBook for " + workbook_code)
    book = openpyxl.load_workbook(matches[0], read_only=True, data_only=True)
    sheets = [s for s in book.sheetnames
              if s.replace(" ", "").lower() == "renewableprofile"]
    if not sheets:
        raise IOError("no RenewableProfile sheet in " + os.path.basename(matches[0]))
    sheet = book[sheets[0]]

    rows = list(sheet.iter_rows(values_only=True))
    head = next(i for i, r in enumerate(rows)
                if r and any(str(c).strip() == "January" for c in r if c))
    columns = {str(c).strip(): j for j, c in enumerate(rows[head]) if c}
    name_col = columns["Assigned Name"]
    month_cols = [columns[m] for m in MONTHS]

    out = []
    for r in rows[head + 1:]:
        if len(r) <= name_col or not r[name_col]:
            continue
        values = [num(r[j]) if j < len(r) else None for j in month_cols]
        if any(v is None for v in values):
            continue
        out.append((str(r[name_col]).strip(), values))
    book.close()
    return out


def selected(rows, match, exclude=""):
    """The rows whose name carries any of the substrings and none of the exclusions.

    Both lists are read from the mapping, case insensitive, separated by a bar. The
    exclusion exists because one Kazakh solar plant is called Arm Wind: a name is a
    weak technology tag and the mapping is where that is repaired, not the code.
    """
    wanted = [w.strip().lower() for w in match.split("|") if w.strip()]
    barred = [w.strip().lower() for w in exclude.split("|") if w.strip()]
    return [(n, v) for n, v in rows
            if any(w in n.lower() for w in wanted)
            and not any(w in n.lower() for w in barred)]


def fit_scale(values, weights, target, limit, ceiling=1.0):
    """The multiplier bringing a series to a target mean without leaving [0, ceiling].

    Plain multiplication is the right model going down: soiling, outages and curtailment
    all scale output by a factor, which is what the solar gap is. It only breaks going up,
    because a series already at rated power cannot be doubled and a capacity factor above
    one is not a number the model can use. So the multiplier is fitted with the ceiling in
    place, which spends an increase on more hours AT rated rather than on hours above it,
    which is also what a stronger wind resource physically does.

    Returns None when the ceiling cannot reach the target however hard it is pushed, which
    happens when a series is zero for more of the year than the target leaves room for,
    or when the target is simply further away than the caller's limit allows.
    """
    total = float(sum(weights))

    def mean_at(scale):
        if not total:
            return 0.0
        return sum(min(v * scale, ceiling) * w
                   for v, w in zip(values, weights)) / total

    if mean_at(limit) < target:
        return None
    low, high = 0.0, limit
    for _ in range(60):
        middle = (low + high) / 2.0
        if mean_at(middle) < target:
            low = middle
        else:
            high = middle
    return (low + high) / 2.0


def representative_dates(path):
    """(season, day type) -> (month, day of month), the day each block stands for.

    build_demand.py chose these days on the load and wrote them down here. Reading the
    renewable year on the SAME days is the whole point: it is what puts the weather of a
    real evening behind the peak that was measured that evening, instead of a seasonal
    average behind every hour of the season.
    """
    out = {}
    for row in dicts(path):
        month, day = row["date"].split("/")
        out[(row["q"].strip(), row["d"].strip())] = (int(month), int(day))
    return out


def hourly_days(path):
    """(zone, month, day) -> [24 values], from a rescaled hourly year.

    A day is kept only if all 24 hours of it are there; a partial day would be read as a
    dark one and would quietly lower the season it fell in.
    """
    out = {}
    with io.open(path, encoding="utf-8-sig", newline="") as fh:
        for row in csv.DictReader(fh):
            value = (row.get("value") or "").strip()
            if not value:
                continue
            key = (row["zone"].strip(), int(row["month"]), int(row["day"]))
            day = out.setdefault(key, [None] * HOURS_PER_DAY)
            day[int(row["hour"]) - 1] = float(value)
    return dict((k, v) for k, v in out.items() if None not in v)


def real_days(zone, hourly, dates, days, weights, factors):
    """season -> ({day type: [24 values]}, multiplier), for the seasons that hold up.

    WHY THERE IS A SECOND RESCALE HERE when build_vre_hourly.py already put the whole
    year on the stated level. The three days kept out of a season are not a mean
    preserving sample of it: they were chosen on the load, so whether they happen to be
    sunny is luck, and the model integrates the season as those three days weighted by
    pHours and nothing else. Without this step a peak day that happened to be clear would
    raise the annual solar energy of the whole model by however much luck it had. The
    multiplier is ONE PER SEASON, so the three days keep their differences from each
    other, which is the reason for reading real days in the first place.

    A SEASON AT A TIME, not a zone at a time. A representative-days model has no
    continuity across a season boundary -- each season is its own set of blocks and
    already carries its own shape -- so a season that has to fall back costs the others
    nothing, and northern Kazakhstan keeps real spring, summer and autumn days instead of
    losing them all to one overcast February. Seasons missing from the returned map are
    the ones the caller must build the old way.
    """
    out = {}
    for season, types in days.items():
        picked = {}
        for day_type in types:
            day = hourly.get((zone,) + dates[(season, day_type)])
            if day is None:
                picked = None
                break
            picked[day_type] = day
        if picked is None:
            continue
        values = [v for t in types for v in picked[t]]
        hours = [h for t in types for h in weights[(season, t)]]
        scale = fit_scale(values, hours, factors[season], SEASON_STRETCH)
        if scale is None or scale < 1.0 / SEASON_STRETCH:
            continue
        out[season] = (dict((t, [min(v * scale, 1.0) for v in picked[t]])
                            for t in types), scale)
    return out


def main():
    ap = argparse.ArgumentParser(description="Build the solar and wind profiles.")
    ap.add_argument("--reference", default=os.path.join("epm", "input", "data_casa_2020"))
    ap.add_argument("--hourly", default=HOURLY,
                    help="where build_vre_hourly.py left the rescaled year")
    args = ap.parse_args()
    reference = os.path.join(REPO, args.reference)

    quarters = seasons(os.path.join(HERE, "mappings", "seasons_months.csv"))
    shape = legacy_profile(reference)
    inherits = legacy_of(os.path.join(HERE, "mappings", "seasons_months.csv"))
    plan = dicts(os.path.join(HERE, "mappings", "vre_profiles.csv"))

    header, hours = read_csv(os.path.join(HERE, "extracted", "pHours.csv"))
    days, weights = collections.OrderedDict(), {}
    for r in hours:
        key = (r[0].strip(), r[1].strip())
        days.setdefault(key[0], []).append(key[1])
        weights[key] = [num(c) or 0.0 for c in r[2:]]

    dates = representative_dates(os.path.join(HERE, "extracted", "demand_report.csv"))
    hourly = {}
    for tech in sorted(set(line["tech"].strip() for line in plan)):
        path = os.path.join(args.hourly, "{0}_casa.csv".format(tech))
        hourly[tech] = hourly_days(path) if os.path.exists(path) else {}
        if not hourly[tech]:
            print("no rescaled hourly year for {0} at {1}:".format(tech, path))
            print("  every {0} zone falls back to the 2020 block shape. Run "
                  "build_vre_hourly.py".format(tech))

    books, out, report, claimed = {}, [], [], {}
    for line in plan:
        z, tech = line["z"].strip(), line["tech"].strip()
        code, match = line["workbook"].strip(), line["match"].strip()
        if code and code not in books:
            books[code] = monthly_factors(code)
        picked = selected(books[code], match, line.get("exclude", "")) if code else []
        if code and not picked:
            raise SystemExit("no DeCA row matches {0} for {1} {2}".format(match, z, tech))
        if code:
            names = set(n for n, _ in picked)
            claimed.setdefault((code, z), {})[tech] = names
            for other, taken in claimed[(code, z)].items():
                if other != tech and names & taken:
                    raise SystemExit("{0}: {1} is claimed by both {2} and {3}".format(
                        code, sorted(names & taken)[0], tech, other))

        # The seasonal level first, because it is the same whatever the day shape is.
        means, factors = {}, {}
        for season, months in quarters.items():
            base = shape.get((z, tech, inherits[season]))
            if base is None:
                raise SystemExit("no 2020 shape for {0} {1} {2} (drawn from {3})"
                                 .format(z, tech, season, inherits[season]))
            means[season] = sum(base) / float(HOURS_PER_DAY)
            if picked:
                weight = sum(MONTH_LENGTH[m - 1] for m in months)
                factors[season] = sum(sum(v[m - 1] for _, v in picked) / len(picked)
                                      * MONTH_LENGTH[m - 1] for m in months) / weight
            else:
                factors[season] = means[season]

        built = real_days(z, hourly[tech], dates, days, weights, factors)
        for season in quarters:
            if season in built:
                day_of, scale = built[season]
            else:
                stretch = (factors[season] / means[season]) if means[season] else 0.0
                flat = [v * stretch for v in shape[(z, tech, inherits[season])]]
                day_of, scale = dict((d, flat) for d in days[season]), None
            for d in days[season]:
                out.append([z, tech, season, d]
                           + ["{0:.6g}".format(v) for v in day_of[d]])
            report.append([z, tech, season, "deca" if picked else "casa_2020",
                           "rninja hourly" if season in built else "casa_2020 blocks",
                           len(picked), "{0:.4f}".format(means[season]),
                           "{0:.4f}".format(factors[season]),
                           "{0:.3f}".format(scale) if scale else "",
                           "{0:.4f}".format(max(max(d) for d in day_of.values()))])

    write_csv(os.path.join(HERE, "extracted", "pVREProfile.csv"),
              ["z", "tech", "q", "d"] + ["t{0}".format(h) for h in
                                         range(1, HOURS_PER_DAY + 1)], out)
    write_csv(os.path.join(HERE, "extracted", "vre_report.csv"),
              ["z", "tech", "q", "level", "shape", "deca_rows", "cf_2020", "cf_new",
               "day_scale", "max"], report)

    moved = [r for r in report if r[3] == "deca"]
    real = [r for r in report if r[4] == "rninja hourly"]
    print("rows       {0} on {1} seasons x {2} days".format(
        len(out), len(quarters), len(days[list(quarters)[0]])))
    print("levels     {0} of {1} seasonal factors taken from DeCA".format(
        len(moved), len(report)))
    print("shapes     {0} of {1} zone-seasons read off their own representative days"
          .format(len(real), len(report)))
    for tech in sorted(set(r[1] for r in report)):
        before = [float(r[6]) for r in moved if r[1] == tech]
        after = [float(r[7]) for r in moved if r[1] == tech]
        if before:
            print("{0:<10} mean capacity factor {1:.3f} in 2020, {2:.3f} now".format(
                tech, sum(before) / len(before), sum(after) / len(after)))
        seen = set(r[0] for r in real if r[1] == tech)
        counted = collections.Counter(r[0] for r in real if r[1] == tech)
        flat = sorted(set(r[0] for r in report if r[1] == tech) - seen)
        partial = sorted(z for z in counted if counted[z] < len(quarters))
        print("{0:<10} {1} zone(s) on real days, {2} of them for part of the year"
              .format(tech, len(seen), len(partial)))
        if partial:
            print("{0:<10}   part of the year: {1}".format(tech, ", ".join(
                "{0} {1}/{2}".format(z, counted[z], len(quarters)) for z in partial)))
        if flat:
            print("{0:<10}   flat 2020 day all year: {1}".format(tech, ", ".join(flat)))


if __name__ == "__main__":
    main()
