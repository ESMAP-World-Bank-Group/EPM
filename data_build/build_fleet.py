# -*- coding: utf-8 -*-
"""Rebuild the generator table for a horizon the 2020 model never saw.

    python data_build/build_fleet.py

Reads:
    mappings/tech_lifetime.csv        operating life assumed per technology
    mappings/fleet_retirement.csv     the units whose retirement is a dated decision
    mappings/candidate_limits.csv     what a candidate may build, and how fast
    mappings/storage_candidates.csv   the shape of a battery candidate
    mappings/country_book.csv         which DeCA book speaks for which country
    <reference>/supply/pGenDataInput.csv          the 2020 table
    <reference>/supply/pStorageDataInput.csv      its column shape, it holds no rows
    <reference>/supply/pCapexTrajectoriesDefault.csv  idem
    <target>/y.csv                    the horizon of the new model
    <target>/zcmap.csv                zones to countries
    <target>/load/pDemandForecast.csv used only to share a national rate out by zone
    the DeCA AssumptionBooks          sheets RenewableMaxEntry and CAPEX

Writes:
    extracted/pGenDataInput.csv               taken as is by the build
    extracted/pStorageDataInput.csv           idem
    extracted/pCapexTrajectoriesDefault.csv   idem
    extracted/fleet_retirement_report.csv     what each unit was given, and why
    extracted/candidate_report.csv            idem for the candidates

STEP ONE, the retirement schedule. The schedule of the 2020 model is mechanical: it
is min(max(StYr + Life, 2023), 2031) for 208 of its 232 existing units. The 2031 is
not a retirement date, it is where that model stopped looking. Carried into a run
that goes to 2050 it wipes out 66 GW, Nurek and Toktogul included, in a single year.
The 2023 is the other end of the same convention: everything already past its nominal
life on the day the model was built was retired at once. Rebuilt here from a life
stated per technology, with a floor so that a plant running today does not vanish at
the first year, a horizon past which a unit simply never retires, and an exception
list for the units whose date is a real decision.

STEP TWO, the candidates. Three things were wrong for a 2050 run.

    the closed door   main.gms:754 makes the Capacity column the cumulative build
                      limit of a candidate, and it is 0 for every generic candidate
                      of Kyrgyzstan and Tajikistan and for four of Kazakhstan's. Those
                      countries could build nothing at all outside named projects.
    no speed limit    BuildLimitperYear was ten million MW everywhere, so anything
                      could appear at any rate at all: the run built tens of gigawatts
                      of gas in its first year, and ten gigawatts of Kazakh nuclear
                      was available from 2022. For solar and wind DeCA states a real
                      rate per country, and it steps up after 2030; EPM has one column
                      and cannot step, so the step is carried by a second candidate
                      that opens in the later period with the increment. For everything
                      else no source states a rate, so one is assumed: a share of the
                      peak the zone already carries, floored at one machine of the
                      technology, both written per technology in candidate_limits.csv.
                      The same file dates the first year a technology may be
                      commissioned where that date is a decision rather than a rate.
    no batteries      there were none, in a model that runs to 2050. DeCA prices them
                      at 2, 4, 6 and 8 hours, and the four prices are exactly linear
                      in duration, which separates the cost of the power from the cost
                      of the energy without an assumption having to be made.

The DeCA cost curves are one and the same in the five books, byte for byte, so they
are a technology curve and not a country one; they are applied in every zone.
Batteries are not: their perimeter is the countries named in country_book.csv, which
is what the DeCA data speaks for.
"""

import argparse
import csv
import io
import os

HERE = os.path.dirname(os.path.abspath(__file__))
REPO = os.path.dirname(HERE)

BOOK = "20250306_DeCA - AssumptionBook - {}_V5.1_Clean.xlsx"
# The 2020 table names its menu candidates <ZONE>_<TECH>_<FUEL>_cand and its named
# projects <PROJECT>_NEW_<year>. Only the menu is opened up here.
GENERIC = "_cand"
# What this dataset writes in a Capacity column that is meant to say "no cap".
OPEN = "10000"
# And in a BuildLimitperYear column that is meant to say the same.
NO_LIMIT = "10000000"


def read_csv(path):
    with io.open(path, encoding="utf-8-sig", newline="") as fh:
        rd = csv.reader(fh)
        return next(rd), [r for r in rd if any(c.strip() for c in r)]


def write_csv(path, header, rows):
    with io.open(path, "w", encoding="utf-8", newline="") as fh:
        w = csv.writer(fh, lineterminator="\n")
        w.writerow(header)
        w.writerows(rows)


def index(header):
    return dict((c.strip(), j) for j, c in enumerate(header))


def mapping(path, key):
    """Rows of a hand-edited mapping as dicts, indexed by one of their columns."""
    header, rows = read_csv(path)
    i = index(header)
    return dict((r[i[key]].strip(),
                 dict((c, r[j].strip()) for c, j in i.items())) for r in rows)


def pairs(path):
    """Rows of a mapping indexed by the (tech, fuel) pair it declares."""
    header, rows = read_csv(path)
    i = index(header)
    return dict(((r[i["tech"]].strip(), r[i["fuel"]].strip()),
                 dict((c, r[j].strip()) for c, j in i.items())) for r in rows)


def for_pair(table, tech, fuel):
    """The row that speaks for this pair, the exact one first, then the wildcard."""
    return table.get((tech, fuel)) or table.get((tech, "*"))


def years(path):
    _, rows = read_csv(path)
    return [int(r[0]) for r in rows]


def header_row(rows, label):
    """Index of the row that holds the column names, found by one of them."""
    for i, r in enumerate(rows):
        if any(isinstance(c, str) and c.strip() == label for c in r):
            return i
    raise KeyError("no header row carrying " + label)


def read_max_entry(path):
    """(country, DeCA technology) -> [(first year, last year, MW per year)].

    The sheet carries the technology on one row and the period on the next, so the
    technology has to be carried forward across the columns it spans.
    """
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    rows = list(wb["RenewableMaxEntry"].iter_rows(values_only=True))
    wb.close()

    h = header_row(rows, "Country")
    jc = [j for j, c in enumerate(rows[h])
          if isinstance(c, str) and c.strip() == "Country"][0]
    tech, last = {}, None
    for j in range(jc + 1, len(rows[h])):
        c = rows[h][j]
        if isinstance(c, str) and c.strip():
            last = c.strip()
        tech[j] = last
    period = {}
    for j, c in enumerate(rows[h + 1]):
        if isinstance(c, str) and c.count("-") == 1:
            a, b = c.split("-")
            if a.strip().isdigit() and b.strip().isdigit():
                period[j] = (int(a), int(b))
    if not period:
        raise ValueError("no period header under Country in " + os.path.basename(path))

    out = {}
    for r in rows[h + 2:]:
        cc = r[jc] if jc < len(r) else None
        if not isinstance(cc, str) or not cc.strip():
            continue
        for j, span in sorted(period.items()):
            if j < len(r) and isinstance(r[j], (int, float)):
                out.setdefault((cc.strip(), tech[j]), []).append(
                    (span[0], span[1], float(r[j])))
    for k in out:
        out[k].sort()
    return out


def read_capex(path):
    """Year -> the DeCA capital costs, with the battery split into power and energy.

    DeCA prices a battery at 2, 4, 6 and 8 hours. Those four prices are exactly
    linear in duration in every year, so the slope is the cost of one hour of energy
    and the intercept is the cost of the power. That is a reading of the sheet, not
    an assumption about it, and the check below is there to say so if it stops being
    true in a later version of the book.
    """
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    rows = list(wb["CAPEX"].iter_rows(values_only=True))
    wb.close()

    h = header_row(rows, "Year")
    col = {}
    for j, c in enumerate(rows[h]):
        if isinstance(c, str) and c.strip() and c.strip() not in col:
            col[c.strip()] = j
    out = {}
    for r in rows[h + 1:]:
        y = r[col["Year"]]
        if not isinstance(y, (int, float)):
            continue
        h2, h4, h6 = (float(r[col[k]]) for k in ("2h BESS", "4h BESS", "6h BESS"))
        step = (h4 - h2) / 2.0
        if abs((h6 - h4) / 2.0 - step) > 1.0:
            raise ValueError("the battery prices of {0} are no longer linear in "
                             "duration; the split into power and energy below is "
                             "not valid any more".format(int(y)))
        out[int(y)] = {"solar": float(r[col["Solar Power Plants"]]),
                       "wind": float(r[col["Wind Power Plants"]]),
                       "bess_mw": h2 - 2.0 * step,
                       "bess_mwh": step,
                       "bess_4h": h4}
    return out


def room(rate, first, last):
    """A cumulative cap wide enough that the yearly rate is what binds.

    The 2020 table wrote 10000 in this column to mean "no cap", which was true over a
    horizon that stopped in 2031 and is not true over one that runs to 2050: 10 GW is
    four years of Kazakh solar at the rate DeCA states. Where a rate is stated, the
    cap is derived from it instead of being a round number.
    """
    return "{0:.6g}".format(rate * (last - first + 1))


def implied_share(entry, books, zc, peak, deca_tech):
    """The rate DeCA states for a technology, read as a share of the peak it applies to.

    DeCA has a book for five of the seven countries and none for Pakistan and
    Afghanistan, whose candidates would otherwise keep the ten million MW a year of
    the 2020 table. Read as a share of the peak the country already carries, the rate
    DeCA states is steady enough across the five to stand in for the two: the median
    of those shares is what this returns, so the substitute is taken from the source
    rather than invented.
    """
    out = []
    for c, spec in sorted(books.items()):
        spans = entry.get((spec.get("book", ""), deca_tech))
        if not spans:
            continue
        national = sum(v for z, v in peak.items() if zc.get(z) == c)
        if national > 0:
            out.append(spans[0][2] / national)
    out.sort()
    return out[len(out) // 2] if out else 0.0


def peaks(path, year):
    """Zone -> its peak demand in that year."""
    header, rows = read_csv(path)
    i = index(header)
    if str(year) not in i:
        raise KeyError("pDemandForecast has no column for " + str(year))
    return dict((r[i["z"]].strip(), float(r[i[str(year)]]))
                for r in rows if r[i["type"]].strip().lower() == "peak")


def main():
    ap = argparse.ArgumentParser(description="Rebuild the generator table.")
    ap.add_argument("--reference", default=os.path.join("epm", "input", "data_casa_2020"))
    ap.add_argument("--target", default=os.path.join("epm", "input", "data_casa"))
    ap.add_argument("--books", default=os.path.join("..", "data_collection", "Mercados"))
    ap.add_argument("--floor", type=int, default=None,
                    help="earliest retirement allowed, default the start year plus four")
    args = ap.parse_args()

    maps = os.path.join(HERE, "mappings")
    ref = os.path.join(REPO, args.reference, "supply")
    tgt = os.path.join(REPO, args.target)

    run = years(os.path.join(tgt, "y.csv"))
    start, horizon = run[0], run[-1]
    floor = args.floor if args.floor is not None else start + 4

    lives = dict((k, int(v["life"])) for k, v in
                 pairs(os.path.join(maps, "tech_lifetime.csv")).items())
    fixed = mapping(os.path.join(maps, "fleet_retirement.csv"), "g")
    limits = pairs(os.path.join(maps, "candidate_limits.csv"))
    books = mapping(os.path.join(maps, "country_book.csv"), "c")

    zc = dict((r[0].strip(), r[1].strip())
              for r in read_csv(os.path.join(tgt, "zcmap.csv"))[1])
    peak = peaks(os.path.join(tgt, "load", "pDemandForecast.csv"), horizon)
    # The VRE rule shares a national rate out by zone, so it wants the weights of
    # the end of the run. A thermal rate is a share of the system as it stands
    # today, so it wants the first year: what a country can commission in a year
    # follows from the system it already operates, not from the one it forecasts.
    peak_now = peaks(os.path.join(tgt, "load", "pDemandForecast.csv"), start)

    entry, capex = {}, None
    for c in sorted(books):
        if not books[c]["book"]:
            continue
        path = os.path.join(REPO, args.books, BOOK.format(books[c]["book"]))
        if not os.path.isfile(path):
            raise IOError("AssumptionBook missing: " + path)
        entry.update(read_max_entry(path))
        if capex is None:
            capex = read_capex(path)
    if capex is None:
        raise IOError("no AssumptionBook was read, there is no cost curve to apply")

    # ---- the fleet, and the candidates that sit in the same table ---------------
    header, rows = read_csv(os.path.join(ref, "pGenDataInput.csv"))
    col = index(header)
    for need in ("g", "z", "tech", "fuel", "Status", "StYr", "RetrYr", "Life",
                 "Capacity", "BuildLimitperYear"):
        if need not in col:
            raise KeyError("the reference table has no column " + need)

    out, retired, cand, seen, medians = [], [], [], set(), {}
    for r in rows:
        r = list(r)
        g = r[col["g"]].strip()
        z = r[col["z"]].strip()
        tech = r[col["tech"]].strip()
        fuel = r[col["fuel"]].strip()
        styr = int(r[col["StYr"]])
        out.append(r)

        if r[col["Status"]].strip() == "1":
            was = int(r[col["RetrYr"]])
            life = lives.get((tech, fuel), lives.get((tech, "*")))
            if life is None:
                raise KeyError("no life declared for {0} on {1}".format(tech, fuel))
            if g in fixed:
                retr = int(fixed[g]["retire_year"])
                why = "dated decision: " + fixed[g]["reason"]
                seen.add(g)
                # A dated retirement and a life of 99 cannot both hold. Life is not a
                # duration to this model, it is a flag: main.gms:879 reads 99 as "this
                # unit may never be retired" and fixes vRetire to zero, while
                # main.gms:864 fixes the capacity of the same unit to zero from its
                # retirement year on. The capacity equation then has to drop by the
                # whole plant with nothing allowed to move, and the run stops on an
                # infeasible equation. Where a decision dates the stop, the life is
                # therefore written as the life that decision implies.
                if life >= 99 and retr <= horizon:
                    life = min(retr - styr, 98)
                    why += " (life set to {0} years, the stop being dated)".format(life)
            elif life >= 99:
                retr, why = horizon + 1, "never retires on age"
            else:
                retr = styr + life
                if retr < floor:
                    retr, why = floor, "past its life already, held to the floor year"
                elif retr > horizon:
                    retr, why = horizon + 1, "reaches the end of the run still standing"
                else:
                    why = "commissioned {0}, {1} years of life".format(styr, life)
            r[col["RetrYr"]], r[col["Life"]] = str(retr), str(life)
            retired.append([g, z, tech, fuel, r[col["Capacity"]], styr, was, retr,
                            life, why])
            continue

        # Candidates never retire in EPM: every retirement equation is written on the
        # existing set, base.gms:793 and main.gms:864. Only what they may build matters.
        if not g.endswith(GENERIC):
            continue
        rule = for_pair(limits, tech, fuel)
        if rule is None:
            continue
        why = []

        if rule.get("first_year"):
            first = int(rule["first_year"])
            if styr < first:
                r[col["StYr"]] = str(first)
                why.append("cannot be commissioned before {0}".format(first))
                styr = first

        if rule["cap_mw"] == "open":
            if float(r[col["Capacity"]] or 0) <= 0:
                r[col["Capacity"]] = OPEN
                why.append("was barred from building anything, opened")
        elif rule["cap_mw"] != "keep":
            r[col["Capacity"]] = rule["cap_mw"]
            why.append("capped at " + rule["cap_mw"] + " MW by the mapping")

        def set_rate(rate):
            """The yearly rate, and a cumulative cap derived from it rather than round."""
            r[col["BuildLimitperYear"]] = "{0:.6g}".format(rate)
            r[col["Capacity"]] = room(rate, max(styr, start), horizon)

        if rule["limit_per_year"] == "deca_max_entry":
            c = zc.get(z, "")
            spans = entry.get((books.get(c, {}).get("book", ""), rule["deca_tech"]))
            if not spans:
                med = medians.get(rule["deca_tech"])
                if med is None:
                    med = implied_share(entry, books, zc, peak_now, rule["deca_tech"])
                    medians[rule["deca_tech"]] = med
                rate = med * peak_now.get(z, 0.0)
                if rate <= 0:
                    why.append("DeCA states no rate for {0} and none can be inferred, "
                               "left without a limit".format(c))
                else:
                    set_rate(rate)
                    why.append("DeCA has no book for {0}; across the countries it does "
                               "cover it lets a median {1:.0%} of the national peak be "
                               "connected in {2} each year, and that share of the {3} "
                               "peak of this zone is {4:.6g} MW"
                               .format(c, med, rule["deca_tech"], start, rate))
            else:
                share = peak.get(z, 0.0)
                national = sum(v for zz, v in peak.items() if zc.get(zz) == c)
                part = share / national if national else 0.0
                rate = spans[0][2] * part
                set_rate(rate)
                why.append("DeCA lets {0} connect {1:.0f} MW of {2} a year from {3}, "
                           "and this zone carries {4:.0%} of the national peak in {5}; "
                           "the cumulative cap is set to what that rate can deliver "
                           "over the run, so the rate is what binds and not a round "
                           "number inherited from a shorter horizon"
                           .format(c, spans[0][2], rule["deca_tech"], spans[0][0],
                                   part, horizon))
                for span in spans[1:]:
                    step = (span[2] - spans[0][2]) * part
                    if step <= 0:
                        continue
                    extra = list(r)
                    extra[col["g"]] = "{0}_{1}".format(g, span[0])
                    extra[col["StYr"]] = str(span[0])
                    extra[col["BuildLimitperYear"]] = "{0:.6g}".format(step)
                    extra[col["Capacity"]] = room(step, span[0], horizon)
                    out.append(extra)
                    cand.append([extra[col["g"]], z, tech, fuel,
                                 extra[col["Capacity"]], span[0],
                                 extra[col["BuildLimitperYear"]],
                                 "the rate rises to {0:.0f} MW a year for {1} in {2}; "
                                 "EPM has one column and cannot step, so this second "
                                 "candidate carries the increment, and it carries it "
                                 "to the end of the run because DeCA stops stating a "
                                 "rate after {3}".format(span[2], c, span[0], span[1])])

        elif rule["limit_per_year"].endswith("%"):
            # No source states how fast a country can commission a thermal unit, the
            # way DeCA states it for solar and wind. What the 2020 table wrote instead
            # was ten million MW a year, which is not a limit at all and let the run
            # build tens of gigawatts of gas in its first year. The rate here is a
            # share of the peak the zone already carries, floored at one machine of
            # the technology, so that it scales with the system and never bars a small
            # country from building at all. It is an assumption, stated per technology
            # in candidate_limits.csv, and it is the number to change first.
            share = float(rule["limit_per_year"][:-1]) / 100.0
            top = peak_now.get(z, 0.0)
            rate = max(share * top, float(rule["floor_mw"] or 0))
            if rate <= 0:
                why.append("no peak stated for this zone, left without a limit")
            else:
                set_rate(rate)
                held = rate > share * top
                why.append("was allowed ten million MW a year; the rate is now {0:.6g} "
                           "MW, {1:.0%} of the {2} peak of the zone ({3:.0f} MW){4}, "
                           "and the cumulative cap follows from that rate over the run"
                           .format(rate, share, start, top,
                                   ", held up to the floor of one machine" if held
                                   else ""))

        cand.append([g, z, tech, fuel, r[col["Capacity"]], styr,
                     r[col["BuildLimitperYear"]], "; ".join(why) or "unchanged"])

    unused = sorted(set(fixed) - seen)
    if unused:
        raise KeyError("fleet_retirement.csv names units that are not existing "
                       "generators of the reference table: " + ", ".join(unused))

    # ---- the batteries ---------------------------------------------------------
    sheader, srows = read_csv(os.path.join(ref, "pStorageDataInput.csv"))
    scol = index(sheader)
    shape = pairs(os.path.join(maps, "storage_candidates.csv"))
    inside = sorted(z for z in peak if books.get(zc.get(z, ""), {}).get("book"))
    base = capex[start]
    for (tech, fuel), spec in sorted(shape.items()):
        dur = float(spec["duration_h"])
        for z in inside:
            row = [""] * len(sheader)
            row[scol["g"]] = "{0}_BESS_{1:g}h_cand".format(z, dur)
            row[scol["z"]] = z
            row[scol["tech"]] = tech
            row[scol["f"]] = fuel
            row[scol["Status"]] = "3"
            row[scol["StYr"]] = spec["first_year"]
            row[scol["Life"]] = spec["life"]
            row[scol["Capacity"]] = OPEN
            row[scol["StorageDuration"]] = "{0:g}".format(dur)
            row[scol["BuildLimitperYear"]] = NO_LIMIT
            # CapacityMWh is left empty on purpose: input_verification.py refuses a
            # storage unit that states both a duration and an energy, and the model
            # computes the energy as Capacity * StorageDuration anyway.
            row[scol["Efficiency"]] = spec["efficiency"]
            row[scol["Capex"]] = "{0:.6g}".format(base["bess_mw"] / 1e6)
            row[scol["CapexMWh"]] = "{0:.6g}".format(base["bess_mwh"] / 1e3)
            row[scol["FOMperMW"]] = "0"
            row[scol["VOM"]] = "0"
            row[scol["FixedOMMWh"]] = "0"
            row[scol["VOMMWh"]] = "0"
            srows.append(row)

    # ---- how those costs move ---------------------------------------------------
    cheader, _ = read_csv(os.path.join(ref, "pCapexTrajectoriesDefault.csv"))
    tyears = [int(c) for c in cheader[3:]]
    missing = [y for y in tyears if y not in capex]
    if missing:
        raise KeyError("the DeCA cost curve says nothing about " +
                       ", ".join(str(y) for y in missing))
    curve = {("PV", "Solar"): "solar", ("WT", "Wind"): "wind"}
    for (tech, fuel) in shape:
        curve[(tech, fuel)] = "bess_4h"
    traj, priced = [], set()
    for z in sorted(peak):
        for (tech, fuel), key in sorted(curve.items()):
            if key == "bess_4h" and z not in inside:
                continue
            ref_cost = capex[start][key]
            traj.append([z, tech, fuel] +
                        ["{0:.6g}".format(capex[y][key] / ref_cost) for y in tyears])
            priced.add((z, tech, fuel))

    # Every other triple of the fleet gets a flat curve, and it has to be written down
    # rather than left out: input_treatment.py joins this table on the (z, tech, fuel)
    # of every generator and refuses the run if one of them finds no row. A flat one is
    # not a filler value, it is the statement the sources support, DeCA giving a cost
    # trajectory for solar, wind and batteries alone; a gas turbine of 2050 is assumed
    # to cost what it costs today, in real terms.
    flat = ["1"] * len(tyears)
    for triple in sorted(set((r[col["z"]].strip(), r[col["tech"]].strip(),
                              r[col["fuel"]].strip()) for r in out)):
        if triple not in priced:
            traj.append(list(triple) + flat)

    ext = os.path.join(HERE, "extracted")
    write_csv(os.path.join(ext, "pGenDataInput.csv"), header, out)
    write_csv(os.path.join(ext, "pStorageDataInput.csv"), sheader, srows)
    write_csv(os.path.join(ext, "pCapexTrajectoriesDefault.csv"), cheader, traj)
    write_csv(os.path.join(ext, "fleet_retirement_report.csv"),
              ["g", "z", "tech", "fuel", "capacity_mw", "start_year",
               "retire_2020", "retire_new", "life", "why"], retired)
    write_csv(os.path.join(ext, "candidate_report.csv"),
              ["g", "z", "tech", "fuel", "cap_mw", "start_year",
               "limit_mw_per_year", "why"], cand)

    moved = sum(1 for r in retired if r[6] != r[7])
    opened = sum(1 for r in cand if "opened" in r[7])
    print("run          {0} to {1}, floor for retirement {2}"
          .format(start, horizon, floor))
    print("units        {0}, of which existing {1}".format(len(out), len(retired)))
    print("retirement   moved for {0}, kept for {1}"
          .format(moved, len(retired) - moved))
    print("candidates   {0} lines, {1} of them were barred from building"
          .format(len(cand), opened))
    print("batteries    {0} in {1} zones, {2:.0f} $/MW and {3:.0f} $/MWh in {4}"
          .format(len(srows), len(inside), base["bess_mw"], base["bess_mwh"], start))
    print("trajectories {0} rows, {1} to {2}, of which {3} on a DeCA cost curve"
          .format(len(traj), tyears[0], tyears[-1], len(priced)))


if __name__ == "__main__":
    main()
