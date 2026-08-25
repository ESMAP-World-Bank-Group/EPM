# -*- coding: utf-8 -*-
"""Rebuild the fuel prices and the thermal cost columns of the 2026 model.

    python data_build/build_costs.py

Reads:
    mappings/fuel_prices.csv          where each (country, fuel) price comes from
    mappings/thermal_plants.csv       which DeCA plant each existing unit is
    mappings/thermal_candidates.csv   which DeCA candidate line each (tech, fuel) reads
    <reference>/supply/pFuelPrice.csv the 2020 table, used where DeCA says nothing
    <target>/y.csv                    the horizon of the new model
    <target>/supply/pGenDataInput.csv the heat rates, to price a fuel quoted per MWh
    the DeCA AssumptionBooks          sheets Fuel Prices, Thermal PP and
                                      Thermal PP Candidates

Writes:
    extracted/pFuelPrice.csv           taken as is by the build
    extracted/fuel_price_report.csv    what each country and fuel was given, and why
    extracted/pGenDataInput.csv        rewritten in place, cost columns only
    extracted/thermal_cost_report.csv  what each existing unit was given, and why
    extracted/candidate_cost_report.csv what each candidate was given, and why

RUN ORDER. build_fleet.py first, this second: the thermal stage rewrites the file
build_fleet.py produces. It is idempotent, every value it writes being absolute and
the VOM baseline always being read from the 2020 reference, never from the file
under the pen.

WHY THIS TABLE HAD TO BE REBUILT. The 2020 table stops at 2030 and the model now runs
to 2050. main.gms:737 reads pFuelCost(g,f,y) = pFuelPrice(c,f,y) * pHeatRate(g,f) with
no interpolation and no default, so every year past 2030 was burning free fuel. On a
capacity expansion run that is not a small error, it is the arbitrage itself.

WHAT DeCA GIVES, AND WHICH SHEET SAYS WHAT. Three sheets are read here and they do
not carry the same authority.

    Fuel Prices              one per country, 2024 to 2060, real 2021 dollars. IT IS
                             HEADED International Prices, and that heading is the whole
                             of the difficulty: gas and nuclear there are genuine
                             forward economic prices, the actual 2024 price of each
                             country converging linearly by 2030 to the netback value
                             of the Chinese gas price, but coal there is one flat
                             33.53 $/mt for all five countries and all 37 years, which
                             is the seaborne benchmark and not what anyone in Central
                             Asia pays for lignite.
    Thermal PP               per plant: efficiency, non-fuel variable cost, fuel
                             transport cost, and in two books out of five the local
                             fuel cost actually adopted. That last column is where the
                             coal of this model comes from, and it carries a delivered
                             price, so nothing is added on top of it and NO CALORIFIC
                             VALUE IS ASSUMED ANYWHERE IN THIS FILE.
    Thermal PP Candidates    per technology and not per country: efficiency,
                             amortization period, CAPEX, fixed and variable O&M. The
                             five books state the same table, which is what makes it
                             usable in Pakistan and Afghanistan where DeCA does not
                             reach.

Nuclear is the one conversion left. DeCA quotes it at 2.92 $/MWh of electricity, and
the only way back to the $/MMBtu that EPM wants is to divide by the heat rate the model
carries for its own uranium units.

TRANSPORT. DeCA carries a per-plant fuel transport cost in $/MMBtu, and it is the piece
that turns a border price into a delivered one. It is not equally trustworthy
everywhere, so mappings/fuel_prices.csv says row by row where it goes:

    Uzbekistan and Turkmenistan   one national constant, 1.866 and 0.140. Nothing is
                                  lost by carrying it in the fuel price.
    Kazakh gas                    genuinely regional, 0 at Atyrau on the gas fields to
                                  13.3 in Akmola at the end of the pipeline, which is
                                  larger than the commodity itself. A national mean
                                  would be meaningless, so it is not in the price at
                                  all: each of the two Kazakh gas plants takes its own
                                  through its VOM.
    Kazakh coal                   not used. The column puts the Almaty CHPs at zero
                                  although they burn Ekibastuz coal hauled 1200 km, so
                                  it is not a haulage cost. The north/south differential
                                  is carried instead by the ratio of the 2020 model,
                                  which is the only sourced statement of it.

WHAT IS NOT REBUILT. Heavy fuel oil and LNG are priced nowhere in DeCA, Kazakh and Tajik
coal have no local price in their own books, and Pakistan and Afghanistan are outside
the DeCA perimeter altogether. Those rows keep their 2020 series with the last stated
year held flat, which is an assumption of no real change and is written as such in the
report rather than hidden in a number.
"""

import argparse
import csv
import io
import os

HERE = os.path.dirname(os.path.abspath(__file__))
REPO = os.path.dirname(HERE)

BOOK = "20250306_DeCA - AssumptionBook - {}_V5.1_Clean.xlsx"

# 1 GJ = 0.947817 MMBtu
GJ_TO_MMBTU = 1.0 / 1.055056


def read_csv(path):
    with io.open(path, encoding="utf-8-sig", newline="") as fh:
        rd = csv.reader(fh)
        return next(rd), [r for r in rd if any(c.strip() for c in r)]


def write_csv(path, header, rows):
    with io.open(path, "w", encoding="utf-8", newline="") as fh:
        w = csv.writer(fh, lineterminator="\n")
        w.writerow(header)
        w.writerows(rows)


def index(header):
    return {name.strip(): j for j, name in enumerate(header)}


def num(x):
    """A number, or None. Empty cells and text both come back as None."""
    try:
        return float(str(x).strip())
    except (TypeError, ValueError):
        return None


def years(path):
    """The horizon of the model, as integers, in file order."""
    _, rows = read_csv(path)
    return [int(r[0]) for r in rows if r[0].strip()]


def header_row(rows, test):
    """Index of the row the test accepts, searched from the top."""
    for i, r in enumerate(rows):
        if test(r):
            return i
    raise KeyError("no header row found")


def year_columns(row):
    """{year: column} for a header row that names years, tolerating a 2024* footnote."""
    out = {}
    for j, c in enumerate(row):
        if isinstance(c, int) and 1990 < c < 2200:
            out[c] = j
        elif isinstance(c, str):
            s = c.strip().rstrip("*")
            if s.isdigit() and 1990 < int(s) < 2200:
                out[int(s)] = j
    return out


def deca_fuel(name):
    """DeCA writes a fuel a dozen ways; only three of them are priced."""
    s = str(name).strip().lower()
    if s.startswith("biogas") or s.startswith("bio"):
        return None
    if "coal" in s:
        return "Coal"
    if "gas" in s:
        return "Gas"
    if "uranium" in s or "nuclear" in s:
        return "Nuclear"
    return None


def read_prices(path):
    """{fuel: {year: price}} from the Fuel Prices sheet, in the unit the sheet states."""
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    rows = list(wb["Fuel Prices"].iter_rows(values_only=True))
    wb.close()
    h = header_row(rows, lambda r: len(year_columns(r)) > 10)
    cols = year_columns(rows[h])
    out = {}
    for r in rows[h + 1:]:
        label = next((c for c in r[:3] if isinstance(c, str) and c.strip()), None)
        if not label:
            continue
        f = deca_fuel(label)
        if f is None:
            continue
        series = {y: num(r[j]) for y, j in cols.items() if num(r[j]) is not None}
        if series:
            out[f] = series
    return out


def read_transport(path):
    """Capacity-weighted mean transport cost per fuel, and the value of each plant.

    Returns ({fuel: mean}, {plant: (fuel, transport, capacity)}). The mean is what a
    country-level table can carry; the per-plant values are what the VOM step needs.
    """
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    rows = list(wb["Thermal PP"].iter_rows(values_only=True))
    wb.close()
    h = header_row(rows, lambda r: any(
        isinstance(c, str) and c.strip() == "Assigned Name" for c in r))
    col = {}
    for j, c in enumerate(rows[h]):
        if isinstance(c, str) and c.strip() and c.strip() not in col:
            col[c.strip()] = j
    c_name = col["Assigned Name"]
    c_fuel = col["Fuel Name"]
    c_cap = col["Installed Capacity (MW)"]
    c_tr = col["Transport Variable Costs (US$/MMBTU)"]

    plants, weighted = {}, {}
    for r in rows[h + 1:]:
        name = r[c_name]
        if not isinstance(name, str) or not name.strip():
            continue
        f = deca_fuel(r[c_fuel])
        tr = num(r[c_tr])
        cap = num(r[c_cap]) or 0.0
        if f is None or tr is None:
            continue
        plants[name.strip()] = (f, tr, cap)
        a = weighted.setdefault(f, [0.0, 0.0])
        a[0] += tr * cap
        a[1] += cap
    means = {f: (a[0] / a[1] if a[1] else 0.0) for f, a in weighted.items()}
    return means, plants



def read_local(path):
    """Capacity-weighted local fuel cost per fuel, from a DeCA Thermal PP sheet.

    THE Fuel Prices SHEET IS HEADED International Prices, and its coal row is the same
    33.53 $/mt in all five countries and in every year from 2024 to 2060. That is the
    seaborne benchmark, and DeCA says as much: the convergence method stated under the
    sheet is written for gas alone, netback of the Chinese gas price. Nobody burns
    Angren lignite at the international price, so the coal of this model is read here
    instead, from the price DeCA actually adopted for its own plants.

    That column is filled for two books only, UZ and KG, and not in the same unit in
    both: in UZ the value is $/MWh of electricity, the product of the value and the
    plant efficiency being constant at 9.0 to 9.4 across the 22 gas units, while in KG
    it is $/MMBtu per plant. The caller says which, the mapping carrying the unit next
    to the country.

    Returns {fuel: (value in the stated unit, efficiency)}, both capacity weighted. The
    efficiency comes back because converting a $/MWh figure needs the one DeCA used.
    """
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    rows = list(wb["Thermal PP"].iter_rows(values_only=True))
    wb.close()
    h = header_row(rows, lambda r: any(
        isinstance(c, str) and c.strip() == "Assigned Name" for c in r))
    col = {}
    for j, c in enumerate(rows[h]):
        if isinstance(c, str) and c.strip() and c.strip() not in col:
            col[c.strip()] = j
    c_name = col["Assigned Name"]
    c_fuel = col["Fuel Name"]
    c_cap = col["Installed Capacity (MW)"]
    c_eff = col["Efficiency (MWh produced / MWh fuel)"]
    c_local = next((v for k, v in col.items() if k.startswith("Local Fuel Costs")), None)
    if c_local is None:
        return {}

    acc = {}
    for r in rows[h + 1:]:
        if not isinstance(r[c_name], str) or not r[c_name].strip():
            continue
        f = deca_fuel(r[c_fuel])
        loc, eff, cap = num(r[c_local]), num(r[c_eff]), num(r[c_cap]) or 0.0
        if f is None or loc is None or not eff or not cap:
            continue
        a = acc.setdefault(f, [0.0, 0.0, 0.0])
        a[0] += loc * cap
        a[1] += eff * cap
        a[2] += cap
    return {f: (a[0] / a[2], a[1] / a[2]) for f, a in acc.items() if a[2]}


def held_flat(series, wanted):
    """A series read at each wanted year, the last stated year held flat past its end."""
    known = sorted(series)
    out = {}
    for y in wanted:
        if y in series:
            out[y] = series[y]
        elif y < known[0]:
            out[y] = series[known[0]]
        else:
            out[y] = series[max(k for k in known if k <= y)]
    return out



# --------------------------------------------------------------------------------
# The thermal cost columns.
# --------------------------------------------------------------------------------

BTU_PER_MWH = 3.412           # MMBtu of heat in one MWh of electricity at 100 per cent
DECA_DEFAULT_VOM = 3.4        # the value DeCA writes where it has nothing to say


def read_thermal(path):
    """{assigned name: (efficiency, vom, transport, derating)} from a Thermal PP sheet.

    THE DERATING IS THE SECOND CAPACITY COLUMN. DeCA states an installed capacity and,
    beside it, an available capacity, and the two differ for most of the fleet: Bishkek
    is 812 MW installed and 420 available, Turkmenbashi 420 and 180, Angren 393 and
    250. That is not the outage rate, which sits three columns further right in
    Forced and Scheduled Outage Rate and is 0.05 each for every unit of all five
    books; it is what an ageing plant can actually put on the bars. The two multiply:
    a unit is derated to its available capacity and then out of service a tenth of the
    time. Returned as one number per plant, available / installed x (1 - forced -
    scheduled), so that whoever uses it cannot apply half of it.

    It is a RATIO and that matters, the mapping from model units to DeCA plants not
    being one to one: 17 DeCA plants are split across several rows of this model, so
    an absolute capacity could not be carried across it without deciding how to share
    it, while a ratio applies to each row of a group unchanged.
    """
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    rows = list(wb["Thermal PP"].iter_rows(values_only=True))
    wb.close()
    h = header_row(rows, lambda r: any(
        isinstance(c, str) and c.strip() == "Assigned Name" for c in r))
    col = {}
    for j, c in enumerate(rows[h]):
        if isinstance(c, str) and c.strip() and c.strip() not in col:
            col[c.strip()] = j
    c_name = col["Assigned Name"]
    c_eff = col["Efficiency (MWh produced / MWh fuel)"]
    c_vom = col["Non-fuel Variable Costs (US$/MWh)"]
    c_tr = col["Transport Variable Costs (US$/MMBTU)"]
    c_inst = col["Installed Capacity (MW)"]
    c_avail = col["Available capacity (MW)"]
    c_fo = col["Forced Outage Rate"]
    c_so = col["Scheduled Outage Rate"]
    out = {}
    for r in rows[h + 1:]:
        name = r[c_name]
        if not isinstance(name, str) or not name.strip():
            continue
        inst, avail = num(r[c_inst]), num(r[c_avail])
        up = 1.0 - (num(r[c_fo]) or 0.0) - (num(r[c_so]) or 0.0)
        derate = None
        if inst and avail is not None:
            # Uzbekistan states an available capacity above the installed one for two
            # units, which is a book error and not a plant that exceeds its rating.
            derate = min(avail / inst, 1.0) * up
        out[name.strip()] = (num(r[c_eff]), num(r[c_vom]), num(r[c_tr]), derate)
    return out


def thermal(maps, books_dir, target, reference):
    """Rewrite HeatRate and VOM of the matched thermal units, in place.

    THE HEAT RATE is taken from DeCA wherever a unit is matched, because a stated
    per-plant efficiency beats the per-technology constant the 2020 model carried:
    that model gave every Kazakh coal CHP the same 11.23 where DeCA separates them
    from 0.208 at Astana-2 to 0.275 at Pavlodar-3.

    THE VOM IS TAKEN ONLY WHERE DeCA DEPARTS FROM ITS OWN DEFAULT. It writes 3.400
    $/MWh for 128 of its 136 thermal units, across five countries and every
    technology at once, which is a filler and not a measurement; adopting it would
    flatten the technology spread the 2020 model does carry. Where DeCA does depart,
    the departure is the information: the old Uzbek steam sets at 9.100, the modern
    combined cycles at 4.250, Navoi at 3.500, Takhiatash at 3.540.

    KAZAKH GAS TRANSPORT lands here rather than in pFuelPrice, because it is the one
    genuinely regional fuel cost in the region, 0 on the Atyrau gas fields to 13.3 at
    the end of the pipeline in Akmola. A national mean would be meaningless: the model
    carries two Kazakh gas plants and they sit at opposite ends of that range. Each
    takes its own, converted to $/MWh by its own heat rate.
    """
    m_header, m_rows = read_csv(os.path.join(maps, "thermal_plants.csv"))
    mi = index(m_header)

    decks = {}
    for cc in sorted(set(r[mi["book"]].strip() for r in m_rows if r[mi["book"]].strip())):
        decks[cc] = read_thermal(os.path.join(REPO, books_dir, BOOK.format(cc)))

    path = os.path.join(HERE, "extracted", "pGenDataInput.csv")
    if not os.path.exists(path):
        raise SystemExit("run build_fleet.py first: " + path + " is missing")
    header, rows = read_csv(path)
    gi = index(header)

    # The VOM baseline is the 2020 table and never the file being rewritten, so that
    # running this twice writes the same thing twice.
    _, ref_rows = read_csv(os.path.join(REPO, reference, "supply", "pGenDataInput.csv"))
    base_vom = {r[0].strip(): num(r[gi["VOM"]]) for r in ref_rows}
    base_hr = {r[0].strip(): num(r[gi["HeatRate"]]) for r in ref_rows}

    by_g = {}
    for r in m_rows:
        by_g[r[0].strip()] = (r[mi["book"]].strip(), r[mi["deca_plant"]].strip())

    report, touched = [], 0
    for r in rows:
        g = r[0].strip()
        if g not in by_g:
            continue
        cc, plant = by_g[g]
        # Read the before from the 2020 reference, so the report says the same
        # thing on a second run as on the first.
        hr_old = base_hr.get(g, num(r[gi["HeatRate"]]))
        vom_old = base_vom.get(g, num(r[gi["VOM"]]))
        if not cc:
            report.append([g, "", "",
                           "{0:.4f}".format(hr_old or 0), "{0:.4f}".format(hr_old or 0),
                           "{0:.3f}".format(vom_old or 0), "{0:.3f}".format(vom_old or 0),
                           "", "kept, the mapping says why"])
            continue
        if plant not in decks[cc]:
            raise KeyError("the mapping points at a plant the book does not carry: "
                           + cc + " " + plant)
        eff, vom_deca, tr, _derate = decks[cc][plant]
        if not eff:
            raise ValueError("no efficiency stated for " + cc + " " + plant)

        hr_new = BTU_PER_MWH / eff
        base = base_vom.get(g, vom_old) or 0.0
        why = []
        if vom_deca is not None and abs(vom_deca - DECA_DEFAULT_VOM) > 1e-9:
            base = vom_deca
            why.append("DeCA states {0:.3f}".format(vom_deca))
        else:
            why.append("2020 value kept, DeCA sits at its own default")
        vom_new = base
        term = ""
        if cc == "KZ" and r[gi["fuel"]].strip().lower().startswith("gas") and tr:
            add = tr * hr_new
            vom_new = base + add
            term = "{0:.3f}".format(add)
            why.append("transport {0:.3f} $/MMBtu over {1:.3f} MMBtu/MWh".format(tr, hr_new))

        r[gi["HeatRate"]] = "{0:.4f}".format(hr_new)
        r[gi["VOM"]] = "{0:.4f}".format(vom_new)
        touched += 1
        report.append([g, cc, plant,
                       "{0:.4f}".format(hr_old or 0), "{0:.4f}".format(hr_new),
                       "{0:.3f}".format(vom_old or 0), "{0:.3f}".format(vom_new),
                       term, "; ".join(why)])

    write_csv(path, header, rows)
    write_csv(os.path.join(HERE, "extracted", "thermal_cost_report.csv"),
              ["g", "book", "deca_plant", "heatrate_before", "heatrate_after",
               "vom_before", "vom_after", "transport_term", "why"], report)

    # ---- the derating, handed to build_hydro.py -------------------------------
    # Written here and not there because the sheet and the mapping are already open
    # here, and read there because availability is that file's subject. RUN ORDER IS
    # THEREFORE build_fleet.py, THIS, THEN build_hydro.py.
    every = sorted(d for deck in decks.values() for (_e, _v, _t, d) in deck.values()
                   if d is not None)
    if not every:
        raise ValueError("no book states an available capacity, the derating is empty")
    mid = every[len(every) // 2] if len(every) % 2 else (
        every[len(every) // 2 - 1] + every[len(every) // 2]) / 2.0
    # THE FILE IS THE PERIMETER. It carries a line for a unit of the five DeCA
    # countries and for no one else, so a Pakistani or Afghan plant is simply absent
    # and keeps whatever the 2020 model gave it. That is the rule of this build read
    # strictly: DeCA wins where DeCA and 2020 disagree, and outside the five countries
    # DeCA says nothing, so there is nothing to disagree with. Carrying the Central
    # Asian median across the border would be the opposite -- inventing a number for a
    # fleet no source in hand describes, and moving the answer of a study whose subject
    # is how much Central Asian power is worth to Pakistan.
    drows = []
    for g in sorted(by_g):
        cc, plant = by_g[g]
        d = decks[cc][plant][3] if cc and plant in decks.get(cc, {}) else None
        if d is not None:
            drows.append([g, "{0:.6g}".format(d), cc + " " + plant,
                          "available over installed capacity, times one less the "
                          "forced and scheduled outage rates"])
        else:
            # Inside the perimeter, unmatched: the country is described, this plant is
            # not. The median of the fleet DeCA does describe is the substitute, the
            # same move build_fleet.py makes for the RenewableMaxEntry rate.
            drows.append([g, "{0:.6g}".format(mid),
                          "median of the {0} thermal units of the five DeCA books"
                          .format(len(every)),
                          "DeCA carries no such plant, see thermal_plants.csv for why; "
                          "the median of the fleet it does carry stands in"])
    write_csv(os.path.join(HERE, "extracted", "thermal_derating.csv"),
              ["g", "availability", "source", "why"], drows)

    print("thermal    {0} units in the mapping, {1} rebuilt from DeCA".format(
        len(by_g), touched))
    own = sum(1 for r in drows if not r[2].startswith("median"))
    print("derating   {0} unit(s), {1} on their own DeCA plant, {2} on the median "
          "{3:.4f}".format(len(drows), own, len(drows) - own, mid))
    return report



# --------------------------------------------------------------------------------
# The candidates.
# --------------------------------------------------------------------------------

# Which column takes which shape once written back into pGenDataInput.
CAND_FORMAT = {"HeatRate": "{0:.4f}", "Capex": "{0:.6g}", "FOMperMW": "{0:g}",
               "VOM": "{0:g}", "Life": "{0:g}"}


def read_candidates(path):
    """{technology: (efficiency, life, capex $/MW, fom $/MW/yr, vom $/MWh, max MW)}.

    From a DeCA Thermal PP Candidates sheet. The five books do not lay that sheet out
    the same way: KZ inserts Region and Area columns and repeats the cost block to the
    right of the table, TM leaves the technology column unheaded, and TJ names its last
    column Max Cap. rather than Max Capacity. Columns are therefore found by their
    header text and never by position, the first cost block winning where a book has
    two, and the technology being read from the leftmost column that holds names to the
    left of the efficiency.
    """
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    rows = list(wb["Thermal PP Candidates"].iter_rows(values_only=True))
    wb.close()
    h = header_row(rows, lambda r: any(
        isinstance(c, str) and c.strip().startswith("Average Efficiency") for c in r))

    def first(prefix):
        for j, c in enumerate(rows[h]):
            if isinstance(c, str) and c.strip().startswith(prefix):
                return j
        raise KeyError(prefix + " is not on the candidate sheet of " + path)

    c_eff, c_life = first("Average Efficiency"), first("Amortization Period")
    c_capex, c_fom = first("CAPEX"), first("OMFix")
    c_vom, c_max = first("OMVar"), first("Max Cap")

    c_tech = None
    for j in range(c_eff):
        if any(isinstance(r[j], str) and r[j].strip() and num(r[c_eff]) is not None
               for r in rows[h + 1:]):
            c_tech = j
            break
    if c_tech is None:
        raise KeyError("no technology column on the candidate sheet of " + path)

    out = {}
    for r in rows[h + 1:]:
        name, eff = r[c_tech], num(r[c_eff])
        if not isinstance(name, str) or not name.strip() or eff is None:
            continue
        out[name.strip().rstrip("*").strip()] = (
            eff, num(r[c_life]), num(r[c_capex]), num(r[c_fom]), num(r[c_vom]),
            num(r[c_max]))
    return out


def candidate_costs(books_dir):
    """The candidate table read from every book in the directory, and reconciled.

    THE TABLE IS A TECHNOLOGY COST TABLE AND NOT A COUNTRY ONE. The five books state
    the same numbers to the dollar, which is what licenses using it in Pakistan and
    Afghanistan, where DeCA does not reach and where the alternative is a 2020 figure
    that was itself one global set applied to all sixteen zones. Where one book departs
    from the other four it is outvoted and the departure is printed, because a lone
    value in a table five books otherwise agree on is a typo and not a country
    difference: TJ prices new coal at 500000 $/MW against 2500000 everywhere else, and
    it is the only case.

    Returns (table, disagreements).
    """
    import glob
    paths = sorted(glob.glob(os.path.join(REPO, books_dir, BOOK.format("*"))))
    if not paths:
        raise IOError("no AssumptionBook in " + os.path.join(REPO, books_dir))

    seen = {}
    for path in paths:
        cc = os.path.basename(path).split(" - ")[-1].split("_")[0]
        for tech, vals in read_candidates(path).items():
            seen.setdefault(tech, {}).setdefault(vals, []).append(cc)

    table, quarrels = {}, []
    for tech in sorted(seen):
        counts = seen[tech]
        best = max(counts, key=lambda v: len(counts[v]))
        table[tech] = best
        for vals, ccs in counts.items():
            if vals != best:
                quarrels.append(
                    "{0} in {1}: efficiency {2}, capex {3}, fom {4}, vom {5}, against "
                    "{6}, {7}, {8}, {9} in the other books".format(
                        tech, "/".join(sorted(ccs)), vals[0], vals[2], vals[3], vals[4],
                        best[0], best[2], best[3], best[4]))
    return table, quarrels


def candidates(maps, books_dir, reference):
    """Rewrite the cost columns of the thermal candidates, in place.

    WHAT MOVES: HeatRate, Capex, FOMperMW, VOM and Life, from the DeCA candidate table
    through mappings/thermal_candidates.csv, which says technology by technology which
    DeCA line each (tech, fuel) pair of the model reads. LIFE MOVES WITH THE REST
    BECAUSE IT MUST: main.gms:742 builds the capital recovery factor out of it, so
    taking a CAPEX without the amortization period stated beside it would annualize the
    new number on the old horizon and misprice every candidate in the model.

    WHAT DOES NOT MOVE: Capacity, StYr and BuildLimitperYear, which are the shape of
    the menu and not its cost, and the Max Capacity column of DeCA, which is left in
    the report and out of the table. It reads like a unit size, 300 MW for a coal or
    combined cycle block and 1200 for a nuclear one, but DeCA never says so, and EPM
    reads Capacity as the cumulative build limit of a candidate: writing 300 there
    would silently cap all new coal in a country at one block.

    EVERY THERMAL CANDIDATE IS REBUILT, in Pakistan and Afghanistan as in the five DeCA
    countries. The 2020 model applied one global cost set to all sixteen zones, so
    holding the two outside countries there while the five move would not preserve a
    country difference, it would invent one, and the model would arbitrage on it.
    """
    m_header, m_rows = read_csv(os.path.join(maps, "thermal_candidates.csv"))
    mi = index(m_header)
    plan = {(r[mi["tech"]].strip(), r[mi["fuel"]].strip()):
            (r[mi["deca_tech"]].strip(), r[mi["note"]].strip()) for r in m_rows}

    table, quarrels = candidate_costs(books_dir)

    path = os.path.join(HERE, "extracted", "pGenDataInput.csv")
    if not os.path.exists(path):
        raise SystemExit("run build_fleet.py first: " + path + " is missing")
    header, rows = read_csv(path)
    gi = index(header)

    # The before column is read from the 2020 table and never from the file being
    # rewritten, so that running this twice reports the same thing twice.
    ref_header, ref_rows = read_csv(
        os.path.join(REPO, reference, "supply", "pGenDataInput.csv"))
    ri = index(ref_header)
    was = {r[0].strip(): r for r in ref_rows}

    report, touched, kept, missed = [], 0, 0, []
    for r in rows:
        if r[gi["Status"]].strip() != "3":
            continue
        key = (r[gi["tech"]].strip(), r[gi["fuel"]].strip())
        if key not in plan:
            if num(r[gi["HeatRate"]]):
                missed.append(" ".join(key))
            continue
        tech, note = plan[key]
        g = r[0].strip()
        b = was.get(g)
        before = [(num(b[ri[c]]) if b else num(r[gi[c]])) for c in CAND_FORMAT]

        if not tech:
            kept += 1
            row = [g, r[gi["z"]].strip(), key[0], key[1], ""]
            for c, v in zip(CAND_FORMAT, before):
                row += ["{0:g}".format(v or 0), "{0:g}".format(v or 0)]
            report.append(row + ["", note or "kept, DeCA carries no such candidate"])
            continue

        if tech not in table:
            raise KeyError("the mapping asks for a candidate line the books do not "
                           "carry: " + tech)
        eff, life, capex, fom, vom, cap = table[tech]
        after = {"HeatRate": BTU_PER_MWH / eff, "Capex": capex / 1e6,
                 "FOMperMW": fom, "VOM": vom, "Life": life}
        for c, fmt in CAND_FORMAT.items():
            r[gi[c]] = fmt.format(after[c])
        touched += 1

        row = [g, r[gi["z"]].strip(), key[0], key[1], tech]
        for c, v in zip(CAND_FORMAT, before):
            row += ["{0:g}".format(v or 0), "{0:g}".format(after[c])]
        report.append(row + ["{0:g}".format(cap or 0),
                             "DeCA {0}, efficiency {1:g}".format(tech, eff)])

    if missed:
        raise KeyError("thermal candidates the mapping does not cover: "
                       + ", ".join(sorted(set(missed))))

    write_csv(path, header, rows)
    cols = []
    for c in CAND_FORMAT:
        cols += [c.lower() + "_before", c.lower() + "_after"]
    write_csv(os.path.join(HERE, "extracted", "candidate_cost_report.csv"),
              ["g", "z", "tech", "fuel", "deca_tech"] + cols
              + ["deca_max_mw", "basis"], report)
    print("candidates {0} rebuilt from DeCA, {1} kept, DeCA carrying no such "
          "technology".format(touched, kept))
    for q in quarrels:
        print("  outvoted " + q)
    return report


def main():
    ap = argparse.ArgumentParser(description="Rebuild the fuel prices.")
    ap.add_argument("--reference", default=os.path.join("epm", "input", "data_casa_2020"))
    ap.add_argument("--target", default=os.path.join("epm", "input", "data_casa"))
    ap.add_argument("--books", default=os.path.join("..", "data_collection", "Mercados"))
    args = ap.parse_args()

    maps = os.path.join(HERE, "mappings")
    target = os.path.join(REPO, args.target)
    horizon = years(os.path.join(target, "y.csv"))

    # The 2020 table: the shape to reproduce, and the fallback where DeCA says nothing.
    ref_header, ref_rows = read_csv(
        os.path.join(REPO, args.reference, "supply", "pFuelPrice.csv"))
    ref_years = [int(c) for c in ref_header[2:] if c.strip().isdigit()]
    old = {}
    for r in ref_rows:
        series = {y: num(r[2 + i]) for i, y in enumerate(ref_years)
                  if num(r[2 + i]) is not None}
        old[(r[0].strip(), r[1].strip())] = series

    # The heat rate of each fuel, to price anything DeCA quotes per MWh of electricity.
    gen_header, gen_rows = read_csv(os.path.join(target, "supply", "pGenDataInput.csv"))
    gi = index(gen_header)
    heat = {}
    for r in gen_rows:
        hr = num(r[gi["HeatRate"]])
        if hr:
            heat.setdefault(r[gi["fuel"]].strip(), []).append(hr)
    heat = {f: sum(v) / len(v) for f, v in heat.items()}

    plan_header, plan_rows = read_csv(os.path.join(maps, "fuel_prices.csv"))
    pi = index(plan_header)

    wanted_books = sorted(set(r[pi["book"]].strip() for r in plan_rows
                              if r[pi["book"]].strip()))
    prices, transport, local = {}, {}, {}
    for cc in wanted_books:
        path = os.path.join(REPO, args.books, BOOK.format(cc))
        if not os.path.isfile(path):
            raise IOError("AssumptionBook missing: " + path)
        prices[cc] = read_prices(path)
        transport[cc] = read_transport(path)[0]
        local[cc] = read_local(path)

    out, report = [], []
    for r in plan_rows:
        c = r[pi["c"]].strip()
        fuel = r[pi["fuel"]].strip()
        book = r[pi["book"]].strip()
        dfuel = r[pi["deca_fuel"]].strip()
        unit = r[pi["unit"]].strip()
        gj = num(r[pi["heat_content_gj_t"]])
        ratio = num(r[pi["price_ratio"]]) or 1.0
        use_tr = (r[pi["transport"]].strip() == "1")
        note = r[pi["note"]].strip()

        adder = 0.0
        if book and dfuel:
            if unit in ("local_mmbtu", "local_mwh_e"):
                # The price DeCA adopted for its own plants rather than the seaborne
                # benchmark of the Fuel Prices sheet. One 2020 figure, so the series is
                # flat, which is all the source has and what DeCA does with coal anyway.
                if dfuel not in local.get(book, {}):
                    raise KeyError("{0} states no local cost for {1}".format(book, dfuel))
                loc, eff = local[book][dfuel]
                if unit == "local_mmbtu":
                    value = loc
                    basis = "DeCA local adopted, {0:.3f} $/MMBtu".format(loc)
                else:
                    value = loc * eff / BTU_PER_MWH
                    basis = ("DeCA local adopted, {0:.3f} $/MWh over an efficiency of "
                             "{1:.3f}".format(loc, eff))
                series, factor = dict((y, value) for y in horizon), 1.0
            else:
                if dfuel not in prices[book]:
                    raise KeyError("{0} prices no {1}".format(book, dfuel))
                series = held_flat(prices[book][dfuel], horizon)
                if unit == "usd_per_mmbtu":
                    factor, basis = 1.0, "DeCA international $/MMBtu"
                elif unit == "usd_per_t":
                    if not gj:
                        raise ValueError("no heat content for " + c + " " + fuel)
                    factor = 1.0 / (gj * GJ_TO_MMBTU)
                    basis = "DeCA international $/t over {0:.1f} GJ/t".format(gj)
                elif unit == "usd_per_mwh":
                    hr = heat.get(fuel)
                    if not hr:
                        raise ValueError("no heat rate in the fleet for " + fuel)
                    factor = 1.0 / hr
                    basis = "DeCA $/MWh over {0:.3f} MMBtu/MWh".format(hr)
                else:
                    raise ValueError("unknown unit " + unit)
            if use_tr:
                adder = transport[book].get(dfuel, 0.0)
            values = {y: series[y] * factor * ratio + adder for y in horizon}
            source = basis + (", transport +{0:.3f}".format(adder) if adder else "")
            if ratio != 1.0:
                source += ", ratio {0:g}".format(ratio)
        else:
            if (c, fuel) not in old:
                raise KeyError("no 2020 series to fall back on for " + c + " " + fuel)
            values = held_flat(old[(c, fuel)], horizon)
            source = "2020 model, held flat from {0}".format(max(old[(c, fuel)]))

        out.append([c, fuel] + ["{0:.6g}".format(values[y]) for y in horizon])
        was = old.get((c, fuel), {})
        report.append([
            c, fuel, source,
            "{0:.4f}".format(values[horizon[0]]),
            "{0:.4f}".format(values[horizon[-1]]),
            "{0:.4f}".format(was[max(was)]) if was else "",
            "{0:.2f}".format(values[horizon[0]] / was[max(was)]) if was and was[max(was)] else "",
            note,
        ])

    if set((r[0], r[1]) for r in out) != set(old):
        raise ValueError("the mapping does not cover the same country and fuel pairs "
                         "as the 2020 table")

    write_csv(os.path.join(HERE, "extracted", "pFuelPrice.csv"),
              ["c", "fuel"] + [str(y) for y in horizon], out)
    write_csv(os.path.join(HERE, "extracted", "fuel_price_report.csv"),
              ["c", "fuel", "basis", "first_year", "last_year",
               "last_2020_value", "ratio_to_2020", "note"], report)

    from_deca = sum(1 for r in report if r[2].startswith("DeCA"))
    print("horizon    {0} to {1}, {2} years".format(horizon[0], horizon[-1], len(horizon)))
    print("rows       {0}, of which {1} rebuilt from DeCA and {2} held from 2020".format(
        len(out), from_deca, len(out) - from_deca))
    print("transport  " + ", ".join(
        "{0} {1} {2:.3f}".format(cc, f, v)
        for cc in sorted(transport) for f, v in sorted(transport[cc].items())))

    thermal(maps, args.books, target, args.reference)
    candidates(maps, args.books, args.reference)


if __name__ == "__main__":
    main()
