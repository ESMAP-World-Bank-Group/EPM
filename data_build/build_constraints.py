# -*- coding: utf-8 -*-
"""Rebuild the fuel availability limits, and prepare the emission caps.

    python data_build/build_constraints.py

Reads:
    mappings/fuel_limits.csv          row by row, where each limit comes from
    mappings/country_book.csv         which DeCA book speaks for which country
    <reference>/constraint/pMaxFuellimit.csv      the 2020 tables, kept where DeCA
    <reference>/constraint/pMaxFuellimitZone.csv  says nothing
    <target>/y.csv                    the horizon of the new model
    <target>/zcmap.csv                zones to countries
    the DeCA AssumptionBooks          sheets Fuel Availability and Emissions

Writes:
    extracted/pMaxFuellimit.csv          taken as is by the build
    extracted/pMaxFuellimitZone.csv      taken as is by the build
    extracted/fuel_limit_report.csv      what each row was given, and why
    extracted/pEmissionsCountry_netzero.csv  NOT taken by the build, see below
    extracted/emissions_report.csv       the same path in the units DeCA states it in

WHY THE FUEL LIMITS HAD TO BE REBUILT, and it is the same defect as the fuel prices.
The 2020 tables run 2017 to 2030 against a horizon that runs 2026 to 2050, and
base.gms:817 writes the constraint only where the value is strictly positive:

    eFuelLimit(c,f,y)$(fApplyFuelConstraint and pMaxFuelLimit(c,f,y) > 0)..

so past 2030 EVERY FUEL LIMIT SIMPLY DISAPPEARS. Both switches are on in this model,
fApplyFuelConstraint and fApplyZonalFuelLimit, so for half the horizon Uzbekistan could
burn any quantity of gas it liked and northern Kazakhstan was no longer held to the
0.1 that shapes its whole expansion. Nothing crashes, which is what makes it dangerous.

WHAT DeCA GIVES. One sheet, in one book: UZ carries Fuel Availability, and it is the
only one of the five that does. It states 12.1 bcm of gas available for electricity
production, sourced to the Ministry Concept Note for ensuring electricity supply in
Uzbekistan in 2020-2030, and says in as many words that it holds that limit to 2060.
Its coal line is not usable: 8.5 M m3 is a volume, not a heat content, and turning it
into MMBtu would need a density and a calorific value the book never states, which is
exactly the assumption phase 6 spent its time removing. The Uzbek coal limit therefore
stays where it was.

THE ONE CONVERSION. 1 bcm of natural gas is 36 trillion Btu, hence 36 million MMBtu,
which is the unit these tables are written in. 12.1 bcm is 435.6 against the 667.6 the
2020 model allowed, a THIRD LESS GAS FOR UZBEK POWER, and it will structure the Uzbek
expansion. It is a deliberate assumption of DeCA and not an accident of arithmetic: gas
production rises to 2030 but domestic demand rises faster, and the government has ended
gas exports to keep the molecules for higher value uses.

THE EMISSIONS PATH IS NOT WIRED INTO THE BASELINE, and that is a decision rather than
an omission. Every book carries a Net Zero path built off the country NDC, and it is a
SCENARIO: DeCA states it as such. Three things have to be settled before it can be a
constraint, and all three belong to phase 9:

    scope        the path covers ELECTRICITY AND HEAT, and EPM carries no heat. In
                 Kazakhstan heat was 87.1 of the 112.4 MtCO2-eq the sector emitted in
                 1990, so a heat and power cap read onto a power only model is loose by
                 a factor nobody has measured.
    coverage     base.gms:1253 writes eEmissionsCountry for EVERY country at once when
                 fApplyCountryCo2Constraint is on, and a country with no value in the
                 table gets a cap of zero. Pakistan and Afghanistan have no NDC path in
                 these books, so switching the constraint on with this file as it stands
                 would order them to emit nothing at all.
    before 2030  the path states nothing before 2030. Held flat backwards here, which
                 is the mildest thing to do and still binds four years earlier than the
                 NDC ever intended.

The path is written out all the same, on the horizon and in the tons EPM wants, so that
phase 9 has it ready.
"""

import argparse
import csv
import io
import os

HERE = os.path.dirname(os.path.abspath(__file__))
REPO = os.path.dirname(HERE)

BOOK = "20250306_DeCA - AssumptionBook - {}_V5.1_Clean.xlsx"

# 1 bcm of natural gas = 36 trillion Btu = 36 million MMBtu, and these tables are
# written in million MMBtu: base.gms:818 multiplies them by 1e6 to face vFuel.
MMBTU_PER_BCM = 36.0


def read_csv(path):
    with io.open(path, encoding="utf-8-sig", newline="") as fh:
        rd = csv.reader(fh)
        return next(rd), [r for r in rd if any(c.strip() for c in r)]


def write_csv(path, header, rows):
    with io.open(path, "w", encoding="utf-8", newline="") as fh:
        w = csv.writer(fh, lineterminator="\n")
        w.writerow(header)
        w.writerows(rows)


def index(header):
    return {name.strip(): j for j, name in enumerate(header)}


def num(x):
    """A number, or None. Empty cells and text both come back as None."""
    try:
        return float(str(x).strip())
    except (TypeError, ValueError):
        return None


def years(path):
    """The horizon of the model, as integers, in file order."""
    _, rows = read_csv(path)
    return [int(r[0]) for r in rows if r[0].strip()]


def year_columns(row):
    """{year: column} for a header row that names years."""
    out = {}
    for j, c in enumerate(row):
        if isinstance(c, int) and 1990 < c < 2200:
            out[c] = j
        elif isinstance(c, str):
            s = c.strip().rstrip("*")
            if s.isdigit() and 1990 < int(s) < 2200:
                out[int(s)] = j
    return out


def labelled_row(rows, prefix):
    """(row index, column of the label) of the first row whose label starts so."""
    for i, r in enumerate(rows):
        for j, c in enumerate(r):
            if isinstance(c, str) and c.strip().startswith(prefix):
                return i, j
    raise KeyError("no row labelled " + prefix)


def years_above(rows, i):
    """The nearest year header above row i, which is how these sheets are laid out."""
    for k in range(i, -1, -1):
        cols = year_columns(rows[k])
        if len(cols) >= 2:
            return cols
    raise KeyError("no year header above row " + str(i))


def held_flat(series, wanted):
    """A series on the years wanted: interpolated between what is stated, flat outside.

    DeCA states a handful of milestone years and EPM wants every year of the horizon.
    Outside the stated range the last value is repeated, which is an assumption of no
    further change and is written as such in the report rather than hidden in a number.
    """
    stated = sorted(series)
    out = {}
    for y in wanted:
        if y <= stated[0]:
            out[y] = series[stated[0]]
        elif y >= stated[-1]:
            out[y] = series[stated[-1]]
        else:
            lo = max(s for s in stated if s <= y)
            hi = min(s for s in stated if s >= y)
            if lo == hi:
                out[y] = series[lo]
            else:
                w = float(y - lo) / (hi - lo)
                out[y] = series[lo] + w * (series[hi] - series[lo])
    return out


def read_gas_for_power(path):
    """{year: bcm} of gas available for electricity, from a Fuel Availability sheet."""
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if "Fuel Availability" not in wb.sheetnames:
        wb.close()
        raise KeyError("no Fuel Availability sheet in " + path)
    rows = list(wb["Fuel Availability"].iter_rows(values_only=True))
    wb.close()
    i, _ = labelled_row(rows, "Gas Consumption for Power")
    cols = years_above(rows, i)
    got = {y: num(rows[i][j]) for y, j in cols.items() if num(rows[i][j]) is not None}
    if not got:
        raise ValueError("the Gas Consumption for Power row of " + path + " is empty")
    return got


def read_netzero(path):
    """{year: MtCO2-eq} of the Net Zero path, from an Emissions sheet.

    Two books state the path twice, once for the whole economy and once for electricity
    and heat. The first Net Zero row is the sector one in both, and it is the sector one
    this model would ever be held to, so the first is what is read.
    """
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    rows = list(wb["Emissions"].iter_rows(values_only=True))
    wb.close()
    i, j = labelled_row(rows, "Net Zero")
    unit = next((c for c in rows[i][j + 1:j + 3]
                 if isinstance(c, str) and c.strip()), "")
    cols = years_above(rows, i)
    got = {y: num(rows[i][c]) for y, c in cols.items() if num(rows[i][c]) is not None}
    if not got:
        raise ValueError("the Net Zero row of " + path + " is empty")
    return got, unit.strip()


def to_tons(unit):
    """The multiple that turns a stated emission into tons, read off its own unit."""
    u = unit.lower().replace(" ", "")
    if u.startswith("mt"):
        return 1e6
    if u.startswith("kt"):
        return 1e3
    raise ValueError("unknown emission unit " + repr(unit))


def old_table(path):
    """{(key, fuel): {year: value}} of a 2020 fuel limit table."""
    header, rows = read_csv(path)
    cols = [(int(c), j) for j, c in enumerate(header) if c.strip().isdigit()]
    out = {}
    for r in rows:
        series = {y: num(r[j]) for y, j in cols if num(r[j]) is not None}
        if series:
            out[(r[0].strip(), r[1].strip())] = series
    return out


def limits(maps, books_dir, target, reference, horizon):
    """Rebuild both fuel limit tables on the horizon."""
    m_header, m_rows = read_csv(os.path.join(maps, "fuel_limits.csv"))
    mi = index(m_header)

    b_header, b_rows = read_csv(os.path.join(maps, "country_book.csv"))
    bi = index(b_header)
    book_of = {r[0].strip(): r[bi["book"]].strip() for r in b_rows}

    _, z_rows = read_csv(os.path.join(target, "zcmap.csv"))
    country_of = {r[0].strip(): r[1].strip() for r in z_rows}

    old = {
        "country": old_table(os.path.join(REPO, reference, "constraint",
                                          "pMaxFuellimit.csv")),
        "zone": old_table(os.path.join(REPO, reference, "constraint",
                                       "pMaxFuellimitZone.csv")),
    }

    gas = {}
    out = {"country": [], "zone": []}
    report = []
    for r in m_rows:
        table = r[mi["table"]].strip()
        key = r[mi["key"]].strip()
        fuel = r[mi["fuel"]].strip()
        basis = r[mi["basis"]].strip()
        note = r[mi["note"]].strip()
        if table not in out:
            raise ValueError("unknown table " + table + " in fuel_limits.csv")

        was = old[table].get((key, fuel))
        if basis == "held_flat":
            if not was:
                raise KeyError("no 2020 series to hold for " + key + " " + fuel)
            values = held_flat(was, horizon)
            why = "2020 model, held flat from {0}".format(max(was))
        elif basis == "deca_gas_power":
            cc = country_of.get(key, key)
            cb = book_of.get(cc)
            if not cb:
                raise KeyError("no DeCA book for " + cc)
            if cb not in gas:
                gas[cb] = read_gas_for_power(
                    os.path.join(REPO, books_dir, BOOK.format(cb)))
            bcm = held_flat(gas[cb], horizon)
            values = {y: bcm[y] * MMBTU_PER_BCM for y in horizon}
            why = ("DeCA gas for power, {0:.1f} bcm in {1} at {2:g} million MMBtu per "
                   "bcm".format(bcm[horizon[-1]], horizon[-1], MMBTU_PER_BCM))
        else:
            raise ValueError("unknown basis " + basis + " for " + key + " " + fuel)

        out[table].append([key, fuel] + ["{0:.6g}".format(values[y]) for y in horizon])
        report.append([
            table, key, fuel, why,
            "{0:.3f}".format(values[horizon[0]]),
            "{0:.3f}".format(values[horizon[-1]]),
            "{0:.3f}".format(was[max(was)]) if was else "",
            "{0:.2f}".format(values[horizon[-1]] / was[max(was)])
            if was and was[max(was)] else "",
            note])

    for table, rows_out in out.items():
        if set((r[0], r[1]) for r in rows_out) != set(old[table]):
            raise ValueError("the mapping does not cover the same rows as the 2020 "
                             + table + " table")

    write_csv(os.path.join(HERE, "extracted", "pMaxFuellimit.csv"),
              ["c", "fuel"] + [str(y) for y in horizon], out["country"])
    write_csv(os.path.join(HERE, "extracted", "pMaxFuellimitZone.csv"),
              ["z", "fuel"] + [str(y) for y in horizon], out["zone"])
    write_csv(os.path.join(HERE, "extracted", "fuel_limit_report.csv"),
              ["table", "key", "fuel", "basis", "first_year", "last_year",
               "last_2020_value", "ratio_to_2020", "note"], report)
    print("fuel limits {0} country rows and {1} zone rows, on {2} to {3}".format(
        len(out["country"]), len(out["zone"]), horizon[0], horizon[-1]))
    return report


def emissions(maps, books_dir, horizon):
    """Write the Net Zero path on the horizon, for phase 9 to pick up."""
    b_header, b_rows = read_csv(os.path.join(maps, "country_book.csv"))
    bi = index(b_header)

    out, report = [], []
    for r in b_rows:
        c, cb = r[0].strip(), r[bi["book"]].strip()
        if not cb:
            report.append([c, "", "", "", "no DeCA book, no NDC path in this source"])
            continue
        series, unit = read_netzero(os.path.join(REPO, books_dir, BOOK.format(cb)))
        values = held_flat(series, horizon)
        # EPM counts country caps in tons and the books do not all use the same
        # multiple: four state MtCO2-eq and Kyrgyzstan states kT, a thousandfold
        # difference that would have given it a cap it could never break.
        out.append([c] + ["{0:.6g}".format(values[y] * to_tons(unit))
                          for y in horizon])
        report.append([c, unit,
                       "{0:.6g}".format(values[horizon[0]] * to_tons(unit)),
                       "{0:.6g}".format(values[horizon[-1]] * to_tons(unit)),
                       "stated on " + ", ".join(str(y) for y in sorted(series))])

    write_csv(os.path.join(HERE, "extracted", "pEmissionsCountry_netzero.csv"),
              ["c"] + [str(y) for y in horizon], out)
    write_csv(os.path.join(HERE, "extracted", "emissions_report.csv"),
              ["c", "stated_unit", "first_year_tons", "last_year_tons", "stated_years"], report)
    print("emissions  {0} countries on the Net Zero path, {1} without one; NOT wired "
          "into the baseline".format(len(out), len(report) - len(out)))
    return report


def main():
    ap = argparse.ArgumentParser(description="Rebuild the constraint tables.")
    ap.add_argument("--reference", default=os.path.join("epm", "input", "data_casa_2020"))
    ap.add_argument("--target", default=os.path.join("epm", "input", "data_casa"))
    ap.add_argument("--books", default=os.path.join("..", "data_collection", "Mercados"))
    args = ap.parse_args()

    maps = os.path.join(HERE, "mappings")
    target = os.path.join(REPO, args.target)
    horizon = years(os.path.join(target, "y.csv"))

    limits(maps, args.books, target, args.reference, horizon)
    emissions(maps, args.books, horizon)


if __name__ == "__main__":
    main()
