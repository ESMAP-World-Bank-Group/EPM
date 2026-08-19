# -*- coding: utf-8 -*-
"""Extrait la demande annuelle des carnets DeCA V5.1 vers la forme EPM.

    python data_build/extract_demand.py [--deca DOSSIER] [--scenario Reference]

Produit trois fichiers dans data_build/extracted/ :

    deca_demand_hourly.csv     série horaire agrégée aux zones EPM (intermédiaire)
    deca_demand_annual.csv     énergie et pointe par pays, les deux scénarios (audit)
    pDemandForecast.csv        la ressource EPM, reprise telle quelle par le build

Méthode, en trois temps.

1. Demand Evolution donne l'énergie (GWh) et la pointe (MW) PAR PAYS, 2023-2060,
   en deux scénarios. Aucune extrapolation n'est nécessaire : l'horizon 2026-2050
   est entièrement couvert.

2. Demand Profile donne 8760 h par région. On en tire, pour chaque zone EPM, sa
   part d'énergie et sa pointe propre. Le découpage intra-pays est donc celui de
   l'année de base du profil et reste constant dans le temps : DeCA ne fournit pas
   de trajectoire régionale.

3. Énergie zonale = énergie pays x part. Pointe zonale = pointe de la zone dans le
   profil, remise à l'échelle par la croissance de la pointe pays. On passe par le
   facteur de coïncidence observé dans le profil, de sorte que la pointe zonale
   reste NON COÏNCIDENTE, ce qu'attend generate_demand.gms : il multiplie le profil
   normalisé de la zone par cette pointe.

L'Afghanistan et le Pakistan n'ont aucune source 2026. Ils sont reconduits depuis
le pDemandData du modèle 2020 et prolongés à son propre taux de croissance.
"""

import argparse
import collections
import csv
import glob
import io
import os
import sys

import openpyxl

HERE = os.path.dirname(os.path.abspath(__file__))
REPO = os.path.dirname(HERE)
DEFAULT_DECA = os.path.join(os.path.dirname(REPO), "data_collection", "Mercados")

HEADER_ROW = 9          # les carnets DeCA ont un cartouche de 8 lignes
FIRST_DATA_ROW = 10


# ── Lecture des carnets ───────────────────────────────────────────────────────

def workbook(deca_dir, code):
    hits = glob.glob(os.path.join(deca_dir, "*{}_V5.1_Clean.xlsx".format(code)))
    if not hits:
        raise IOError("carnet DeCA introuvable pour {} dans {}".format(code, deca_dir))
    return openpyxl.load_workbook(hits[0], read_only=True, data_only=True)


def header_index(ws):
    hdr = next(ws.iter_rows(min_row=HEADER_ROW, max_row=HEADER_ROW, values_only=True))
    return {str(c).strip(): i for i, c in enumerate(hdr) if c not in (None, "")}


def read_profile(ws, columns):
    """Séries horaires demandées. La colonne d'index d'heure filtre les lignes.

    Les carnets portent des lignes de total sous le tableau (UZ ligne 8770 vaut
    76,7 TWh dans la colonne Central) : exiger un index d'heure numérique les
    écarte proprement, sans coder en dur un nombre de lignes.
    """
    idx = header_index(ws)
    hour_col = next(i for name, i in idx.items() if name.lower().replace(" ", "") in
                    ("year-hour", "hour-year"))
    missing = [c for c in columns if c not in idx]
    if missing:
        raise KeyError("colonnes absentes du profil : {} (disponibles : {})"
                       .format(missing, sorted(idx)))
    out = {c: [] for c in columns}
    for row in ws.iter_rows(min_row=FIRST_DATA_ROW, values_only=True):
        if not isinstance(row[hour_col], (int, float)):
            continue
        for c in columns:
            v = row[idx[c]]
            out[c].append(float(v) if isinstance(v, (int, float)) else 0.0)
    return out


def read_evolution(ws):
    """{(scenario, 'Energy'|'Peak'): {annee: valeur}}.

    Les cinq carnets ne présentent pas le bloc de la même façon : KZ et TM
    préfixent la ligne par le code pays, KG, TJ et UZ non ; les libellés portent
    des astérisques de renvoi. On repère donc les blocs par leur unité, GWh ou MW,
    et le scénario par le mot Reference ou Net Zero présent dans la ligne.
    """
    years, out = None, {}
    for row in ws.iter_rows(min_row=HEADER_ROW, values_only=True):
        cells = list(row)
        nums = [(i, c) for i, c in enumerate(cells)
                if isinstance(c, (int, float)) and 2000 < c < 2100 and float(c).is_integer()]
        labels = " ".join(str(c) for c in cells if isinstance(c, str))

        # ligne d'en-tête d'un bloc : une suite d'années
        if len(nums) >= 20:
            years = {i: int(c) for i, c in nums}
            continue
        if years is None:
            continue

        unit = "Energy" if " GWh" in " " + labels else ("Peak" if " MW" in " " + labels else None)
        if unit is None:
            continue
        scen = "Net Zero" if "net zero" in labels.lower() else (
               "Reference" if "reference" in labels.lower() else None)
        if scen is None:
            continue
        series = {y: float(cells[i]) for i, y in years.items()
                  if isinstance(cells[i], (int, float))}
        if series:
            out[(scen, unit)] = series
    return out


# ── Reconduction Afghanistan / Pakistan depuis le modèle 2020 ─────────────────

def legacy_demand(ref_dir):
    """Énergie (GWh) et pointe (MW) par zone et par année, lues dans le modèle 2020."""
    hours = {}
    with io.open(os.path.join(ref_dir, "pHours.csv"), encoding="utf-8-sig", newline="") as fh:
        rd = csv.reader(fh)
        hdr = next(rd)
        for r in rd:                       # q, d, puis une colonne par bloc horaire
            for i, v in enumerate(r[2:], 2):
                if v not in ("", None):
                    hours[(r[0], r[1], hdr[i])] = float(v)

    energy = collections.defaultdict(float)
    peak = collections.defaultdict(float)
    with io.open(os.path.join(ref_dir, "load", "pDemandData.csv"),
                 encoding="utf-8-sig", newline="") as fh:
        rd = csv.reader(fh)
        hdr = next(rd)                     # z, q, d, y, puis une colonne par bloc
        for r in rd:
            z, q, d, y = r[0], r[1], r[2], int(float(r[3]))
            for i, v in enumerate(r[4:], 4):
                if v in ("", None):
                    continue
                mw = float(v)
                energy[(z, y)] += mw * hours.get((q, d, hdr[i]), 0.0) / 1e3
                peak[(z, y)] = max(peak[(z, y)], mw)
    return energy, peak


def extend(series, years):
    """Prolonge une série au taux composé de ses cinq dernières années."""
    known = sorted(series)
    if not known:
        return {y: 0.0 for y in years}
    last, first = known[-1], known[max(0, len(known) - 6)]
    span = last - first
    rate = ((series[last] / series[first]) ** (1.0 / span) - 1.0) if span and series[first] else 0.0
    out = {}
    for y in years:
        if y in series:
            out[y] = series[y]
        elif y < known[0]:
            out[y] = series[known[0]]
        else:
            out[y] = series[last] * (1.0 + rate) ** (y - last)
    return out


# ── Assemblage ────────────────────────────────────────────────────────────────

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--deca", default=DEFAULT_DECA)
    ap.add_argument("--scenario", default="Reference",
                    help="scenario DeCA retenu comme reference (Reference | Net Zero)")
    ap.add_argument("--target", default=os.path.join(REPO, "epm", "input", "data_casa"))
    ap.add_argument("--reference", default=os.path.join(REPO, "epm", "input", "data_casa_2020"))
    args = ap.parse_args()

    out_dir = os.path.join(HERE, "extracted")
    if not os.path.isdir(out_dir):
        os.makedirs(out_dir)

    # années du modèle, lues dans y.csv : rien n'est codé en dur
    with io.open(os.path.join(args.target, "y.csv"), encoding="utf-8-sig", newline="") as fh:
        years = [int(r[0]) for r in list(csv.reader(fh))[1:] if r and r[0].strip()]

    # table de correspondance zone EPM -> colonnes DeCA
    with io.open(os.path.join(HERE, "mappings", "zones_demand.csv"),
                 encoding="utf-8-sig", newline="") as fh:
        mapping = list(csv.DictReader(fh))

    books = sorted({m["workbook"] for m in mapping if m["workbook"]})
    evolution, profiles = {}, {}
    for code in books:
        wb = workbook(args.deca, code)
        evolution[code] = read_evolution(wb["Demand Evolution"])
        cols = sorted({c for m in mapping if m["workbook"] == code
                       for c in m["profile_columns"].split(";") if c})
        profiles[code] = read_profile(wb["Demand Profile"], cols)
        wb.close()

    # ---- journal d'audit : les deux scénarios, par pays
    with io.open(os.path.join(out_dir, "deca_demand_annual.csv"), "w",
                 encoding="utf-8", newline="\n") as fh:
        w = csv.writer(fh, lineterminator="\n")
        w.writerow(["workbook", "scenario", "type", "year", "value"])
        for code, ev in sorted(evolution.items()):
            for (scen, unit), series in sorted(ev.items()):
                for y in sorted(series):
                    w.writerow([code, scen, unit, y, round(series[y], 3)])

    # ---- série horaire agrégée aux zones EPM (livrable intermédiaire)
    hourly = {}
    for m in mapping:
        if not m["workbook"]:
            continue
        cols = [c for c in m["profile_columns"].split(";") if c]
        series = profiles[m["workbook"]]
        hourly[m["z"]] = [sum(series[c][h] for c in cols) for h in range(len(series[cols[0]]))]
    with io.open(os.path.join(out_dir, "deca_demand_hourly.csv"), "w",
                 encoding="utf-8", newline="\n") as fh:
        w = csv.writer(fh, lineterminator="\n")
        w.writerow(["z", "hour", "MW"])
        for z in sorted(hourly):
            for h, v in enumerate(hourly[z], 1):
                w.writerow([z, h, round(v, 3)])

    # ---- pDemandForecast
    legacy_energy, legacy_peak = legacy_demand(args.reference)
    rows, diag, warnings = {}, [], []

    for code in books:
        zones = [m for m in mapping if m["workbook"] == code]
        ev = evolution[code]
        try:
            e_country = ev[(args.scenario, "Energy")]
            p_country = ev[(args.scenario, "Peak")]
        except KeyError:
            raise KeyError("carnet {} : scenario '{}' absent (disponibles : {})"
                           .format(code, args.scenario, sorted({s for s, _ in ev})))

        # Garde-fou d'unité. Une pointe ne peut pas être inférieure à la puissance
        # moyenne : quand c'est le cas, le bloc est en GW sous une étiquette MW.
        # C'est le cas du carnet TJ, qui annonce 3,72 MW de pointe pour 19,6 TWh.
        mean_mw = max(e_country.values()) * 1e3 / 8760.0
        if max(p_country.values()) < mean_mw:
            p_country = {y: v * 1e3 for y, v in p_country.items()}
            warnings.append("   {} : pointes multipliees par 1000, le bloc etait en GW "
                            "sous une etiquette MW".format(code))

        tot = {z["z"]: sum(hourly[z["z"]]) for z in zones}
        grand = sum(tot.values())
        pk = {z["z"]: max(hourly[z["z"]]) for z in zones}
        n = len(hourly[zones[0]["z"]])
        coincident = max(sum(hourly[z["z"]][h] for z in zones) for h in range(n))

        # la pointe pays du profil sert de repère ; la trajectoire vient de DeCA
        base = min(y for y in p_country if y in e_country)
        for z in zones:
            share = tot[z["z"]] / grand
            ratio = pk[z["z"]] / coincident        # pointe propre / pointe coincidente
            rows[(z["z"], "Energy")] = {y: e_country.get(y, 0.0) * share for y in years}
            rows[(z["z"], "Peak")] = {y: p_country.get(y, 0.0) * ratio for y in years}
        diag.append("   {}  part {} | coincidence {:.2f} | base profil {:.1f} TWh vs DeCA {} {:.1f} TWh"
                    .format(code, " / ".join("{} {:.1f}%".format(z["z"], 100 * tot[z["z"]] / grand)
                                             for z in zones),
                            sum(pk.values()) / coincident, grand / 1e6, base,
                            e_country[base] / 1e3))

    for m in mapping:
        if m["workbook"]:
            continue
        e = extend({y: v for (zz, y), v in legacy_energy.items() if zz == m["z"]}, years)
        p = extend({y: v for (zz, y), v in legacy_peak.items() if zz == m["z"]}, years)
        rows[(m["z"], "Energy")], rows[(m["z"], "Peak")] = e, p

    order = [m["z"] for m in mapping]
    path = os.path.join(out_dir, "pDemandForecast.csv")
    with io.open(path, "w", encoding="utf-8", newline="\n") as fh:
        w = csv.writer(fh, lineterminator="\n")
        w.writerow(["z", "type"] + [str(y) for y in years])
        for z in order:
            for kind in ("Energy", "Peak"):
                w.writerow([z, kind] + [round(rows[(z, kind)][y], 1) for y in years])

    print("Scenario retenu : {}".format(args.scenario))
    print("\n".join(diag))
    print("Ecrit : {}".format(path))
    for y in (years[0], years[-1]):
        tot_e = sum(rows[(z, "Energy")][y] for z in order)
        tot_p = sum(rows[(z, "Peak")][y] for z in order)
        print("   {} : {:.1f} TWh, somme des pointes zonales {:.0f} MW"
              .format(y, tot_e / 1e3, tot_p))


if __name__ == "__main__":
    main()
