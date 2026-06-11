"""Generate Georgia Data Index Excel file."""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Georgia Data Index"

# ── Styles ─────────────────────────────────────────────────────────────────
hdr_font     = Font(bold=True, color="FFFFFF", size=10)
hdr_fill     = PatternFill("solid", fgColor="1F4E79")
used_fill    = PatternFill("solid", fgColor="E2EFDA")
partial_fill = PatternFill("solid", fgColor="FFF2CC")
unused_fill  = PatternFill("solid", fgColor="FCE4D6")
priority_fill = PatternFill("solid", fgColor="DDEBF7")
na_fill      = PatternFill("solid", fgColor="F2F2F2")
section_fill = PatternFill("solid", fgColor="D6E4F0")
section_font = Font(bold=True, size=10, color="1F4E79")
wrap   = Alignment(wrap_text=True, vertical="top")
center = Alignment(horizontal="center", vertical="top", wrap_text=True)
thin   = Side(border_style="thin", color="BFBFBF")
bdr    = Border(left=thin, right=thin, top=thin, bottom=thin)

# ── Columns ─────────────────────────────────────────────────────────────────
headers    = ["#", "File", "Subfolder", "Type", "Content", "EPM Parameters", "Status", "Priority / Notes"]
col_widths = [4,    42,     28,          7,      58,        28,               12,       42]

for col, (h, w) in enumerate(zip(headers, col_widths), 1):
    c = ws.cell(row=1, column=col, value=h)
    c.font = hdr_font; c.fill = hdr_fill; c.alignment = center; c.border = bdr
    ws.column_dimensions[get_column_letter(col)].width = w
ws.row_dimensions[1].height = 30
ws.freeze_panes = "A2"

# ── Data ────────────────────────────────────────────────────────────────────
# (id, file, subfolder, type, content, epm_params, status, notes)
# id="SECTION" -> section header row
S = "SECTION"
U = "USED"; P = "PARTIAL"; N = "NOT USED"; NA = "N/A"

rows = [
    # ── TEAM ROOT ──────────────────────────────────────────────────────────
    (S, "Team/ (root)", "", "", "", "", "", ""),
    (1,  "Av. 3% Load growth (hourly profiles) 2021-2040.xlsx",
         "Team/", "XLSX",
         "Hourly load MW, 2021-2040 (175,320 rows). 3%/yr growth baked in. Single column: datetime + MW load.",
         "pDemandForecast (annual peak + energy)\npDemandProfile (seasonal mean hourly profile)",
         U, "Growth rate undocumented. Confirm source and methodology with GSE/GNERC. No electrification scenario."),

    (2,  "GE_Power Sector_Data Repository.xlsx",
         "Team/", "XLSX",
         "4 sheets incl. Power Plant List: 113 plants with type, capacity MW, commissioning year, public/private flag.",
         "pGenDataInput (StYr, tech type cross-check)",
         U, ""),

    (3,  "Generation_1.07.2022_shared by Genex.xlsx",
         "Team/", "XLSX",
         "2 sheets: (1) GENERATION_2022 — ~140 plants with capacity, PSO/PPA status, tariffs (tetri/kWh + GEL/day capacity fee); (2) Balance 2007-2021 — historical electricity balance by category. Source: GENEX market operator.",
         "pFuelPrice (validate implied gas price from PSO tariffs)",
         N, "Contains actual PSO tariffs and balance 2007-2021. Useful to validate gas price assumption and historical dispatch."),

    (4,  "List of PPAs and MOUs_May 2022 English.xlsx",
         "Team/", "XLSX",
         "3 sheets (HPP, Wind, Solar). All PPAs/MOUs as of May 2022: status (construction/signed/MOU), capacity MW, company, region.",
         "pGenDataInput (committed plants: Khudoni 702 MW, Namakhvani 433 MW, Nenskra 280 MW, SHP aggregate)",
         U, ""),

    (5,  "List of PPAs and MOUs_May 2022 Georgian.xlsx",
         "Team/", "XLSX",
         "Same content in Georgian script.",
         "-", NA, "Duplicate — use English version."),

    (6,  "221007 Georgia RE Pipeline.xlsx",
         "Team/", "XLSX",
         "RE pipeline summary by stage (feasibility, licensing, construction, PPA) and type. Totals: Hydro ~500 MW, Wind ~94 MW, Solar ~150 MW.",
         "pGenDataInput (RE candidates: Wind 300 MW, PV 200 MW)",
         U, ""),

    (7,  "230110 DPO Results Indicator on RE.xlsx",
         "Team/", "XLSX",
         "5 sheets: RE installed capacity with PSO/PPA tariffs per plant (tetri/kWh); RE project pipeline; balance 2007-2021; installed capacity 2022. WB DPO source.",
         "pFuelPrice (tariff validation)\npGenDataInput (cross-check capacity)",
         N, "Contains actual tariffs per plant — useful to validate implied gas/fuel prices."),

    (8,  "RoR_Hourly Historic Operating Data.xlsx",
         "Team/", "XLSX",
         "Single sheet: hour (1-8760) + P/Pn (normalized aggregate RoR capacity factor). Actual historic operating data.",
         "pAvailabilityCustom (RoR aggregate: Q1=0.45, Q2=0.81, Q3=0.54, Q4=0.40)",
         U, ""),

    (9,  "Georgia_fuel-subsidies_2022.xlsx",
         "Team/", "XLSX",
         "IMF Energy Subsidies DB, Georgia 2021. Key sheet Single_Country_Fuel: consumer price, supply cost, climate cost for each fuel (gas-power, coal-power, diesel etc.) in USD/GJ.",
         "pFuelPrice (Tkibuli coal: 3.82 $/MMBtu)\n[gas-power supply cost available but not yet extracted]",
         P, "Gas-power supply cost is available in the file — should be extracted directly instead of using the current proxy."),

    (10, "Copy of ALL ONGOING PROJECTS - August 2022.xlsx",
         "Team/", "XLSX",
         "3 sheets (HPP, WPP, SPP). Detailed project list: company, region, municipality, river, type, status.",
         "pGenDataInput (complement RE candidates)",
         N, "More granular than RE Pipeline summary."),

    (11, "Electricity - reliability and price.xlsx",
         "Team/", "XLSX",
         "Sheet 1: network reliability indicators (SAIDI, SAIFI, DB scores, 2015-2019). Sheet 2: electricity prices globally.",
         "-", NA, "Context only — no direct EPM input."),

    (12, "Generation Mix (LCGDP).xlsx",
         "Team/", "XLSX",
         "Aggregated generation mix (% fossil, nuclear, RE). No time series.",
         "-", NA, "Context only."),

    (13, "2024 GNERC Annual Report.pdf",
         "Team/", "PDF",
         "GNERC regulator annual report, 2024 data. Likely: actual demand, installed capacity, generation by source, tariffs, reliability metrics.",
         "pDemandForecast (2024 baseline validation)",
         N, "PRIORITY 2 — official 2024 statistics. Read before team meeting."),

    (14, "2022_10 Notes on GE-AM Cross-Border Interconnection.msg",
         "Team/", "MSG",
         "Email/technical note on Georgia-Armenia interconnection (Oct 2022). Likely: actual line capacity, technical constraints, CTN project status.",
         "pTransferLimit (GE-ARM corridor)",
         N, "Relevant before team meeting on interconnection discussion."),

    # ── GENERATION BALANCE ─────────────────────────────────────────────────
    (S, "Team/Generation Balance/", "", "", "", "", "", ""),
    (15, "Generation hourly profiles 2019-2022.xlsx",
         "Team/Generation Balance/", "XLSX",
         "4 sheets (2019-2022), 8,760 hourly rows each. Columns: Reservoir HPP (MW), RoR (MW), TPP (MW), Wind (MW). Actual dispatch by technology type.",
         "pAvailabilityCustom (cross-validate reservoir and RoR seasonal CFs)",
         N, "Can cross-check WB EPM v8.5 calibration. Derive actual seasonal capacity factors per technology type 2019-2022."),

    (16, "Georgia Electricity Balance_2020-22.xlsx",
         "Team/Generation Balance/", "XLSX",
         "Monthly electricity balance 2020-2022: hydro, gas, imports from Azerbaijan, imports from Russia, exports, total consumption.",
         "pTransferLimit (actual cross-border trade volumes by month)",
         N, "Actual monthly import flows from Russia and Azerbaijan. Useful for seasonal import constraints calibration."),

    (17, "230703 Georgia Electricity Balance.pptx",
         "Team/Generation Balance/", "PPTX",
         "Presentation (July 2023) with electricity balance data, context 2022-2023.",
         "-", N, "More recent balance context — read before team meeting."),

    (18, "230703 Georgia Electricity Balance_Eurostat Data.xlsx",
         "Team/Generation Balance/", "XLSX",
         "Eurostat data series for Georgia (historical statistics).",
         "-", N, "Historical validation reference."),

    (19, "generation_2020.pdf",
         "Team/Generation Balance/", "PDF",
         "Official GSE annual generation balance 2020 by plant/technology.",
         "-", N, "Official source."),

    (20, "generation_2021.pdf",
         "Team/Generation Balance/", "PDF",
         "Official GSE annual generation balance 2021 by plant/technology.",
         "-", N, ""),

    (21, "generation_2022.pdf",
         "Team/Generation Balance/", "PDF",
         "Official GSE annual generation balance 2022 by plant/technology.",
         "-", N, ""),

    (22, "generation_2023.pdf",
         "Team/Generation Balance/", "PDF",
         "Official GSE annual generation balance 2023 by plant/technology. Most recent PDF available.",
         "pGenDataInput (cross-check actual capacity in service 2023)",
         N, "PRIORITY — most authoritative source for current fleet status."),

    # ── TRANSMISSION & GENERATION PROJECT TIMELINES ────────────────────────
    (S, "Team/Transmission & Generation Project Timelines/", "", "", "", "", "", ""),
    (23, "GenerationPipeline_GSE_TYNDP.xlsx",
         "Team/Transmission & Gen. Timelines/", "XLSX",
         "~100 hydro projects (mostly RoR <15 MW) with: capacity MW, annual energy GWh, TYNDP COD date, estimated actual COD.",
         "pGenDataInput (refine SmallHydro candidates beyond current 300 MW aggregate)",
         N, "Much more detailed than current AGG_SmallHydro. Worth exploiting to improve candidate pipeline."),

    (24, "Financial Model_Draft Version_2024.xlsx",
         "Team/Transmission & Gen. Timelines/", "XLSX",
         "Financial model for generation projects (March 2024). Likely: CAPEX, OPEX, financing structure by project.",
         "pGenDataInput (CAPEX for candidate plants)",
         N, "More recent CAPEX estimates than EPM generic defaults."),

    # ── GENEX MARKET SIMULATIONS ───────────────────────────────────────────
    (S, "Team/GENEX Market Simulations/", "", "", "", "", "", ""),
    (25, "Price and Volume (JULY).xlsx",
         "Team/GENEX Market Simulations/", "XLSX",
         "Hourly spot market prices (USD/MWh) and volumes (MW), July 2022. Prices highly variable: 50-300 $/MWh — hydro abundant in summer, low prices.",
         "-", N, "Context: validates that hydro surplus drives low prices in summer. Useful for dispatch result validation, not a direct EPM input."),

    (26, "Price and Volume (October).xlsx",
         "Team/GENEX Market Simulations/", "XLSX",
         "Hourly spot market prices and volumes, October 2022. Prices near-constant at 300 $/MWh (price cap hit continuously) — strong system stress signal in Q4.",
         "-", N, "Confirms Q4 system stress (low hydro + imports near capacity limit). Useful to validate model dispatch results."),

    # ── GSE SCADA & METERING ───────────────────────────────────────────────
    (S, "Team/GSE SCADA & Metering Systems/", "", "", "", "", "", ""),
    (27, "2020-11-09_GSE SCADA-EMS Upgrade_AT report.pdf",
         "Team/GSE SCADA & Metering/", "PDF",
         "Technical audit report of GSE SCADA/EMS modernisation project (Nov 2020). System architecture, measurements, data flows.",
         "-", NA, "Infrastructure context only — no EPM input."),

    (28, "2020-11-24_GSE SCADA-EMS Upgrade_Operational Acceptance Certificate.pdf",
         "Team/GSE SCADA & Metering/", "PDF",
         "Acceptance certificate for SCADA/EMS upgrade project.",
         "-", NA, "Administrative."),

    (29, "AW Procurement Package - GSE metering assessment.msg",
         "Team/GSE SCADA & Metering/", "MSG",
         "Email: procurement package for GSE metering system assessment.",
         "-", NA, "Administrative."),

    (30, "HESI.RAR",
         "Team/GSE SCADA & Metering/", "RAR",
         "Archive (~9 MB). Contents unknown — possibly SCADA data or HESI system files.",
         "-", N, "Unopened — may contain metering/SCADA data worth exploring if hourly load data needed."),

    (31, "MeteringSystem-Georgia.pdf",
         "Team/GSE SCADA & Metering/", "PDF",
         "Detailed study of Georgia metering system architecture (~18 MB).",
         "-", NA, "Technical network documentation — no EPM input."),

    (32, "SAMAST.pdf",
         "Team/GSE SCADA & Metering/", "PDF",
         "Description of SAMAST metering/metrology system (GSE).",
         "-", NA, "Technical."),

    (33, "SCADA Upgrade Project Pictures...docx",
         "Team/GSE SCADA & Metering/", "DOCX",
         "Photos of SCADA hardware/software upgrade (Dec 2020).",
         "-", NA, "Documentation only."),

    # ── SOE DATABASE ───────────────────────────────────────────────────────
    (S, "Team/SOE Database/", "", "", "", "", "", ""),
    (34, "GE_SOE_global_database_for_validation.xlsx",
         "Team/SOE Database/", "XLSX",
         "WB Global SOE Database, Georgia subset (under embargo). Sheets: company list (BvD ID, name, sector NACE), ownership structure. Covers energy SOEs (GSE, Engurhesi etc.).",
         "-", N, "Useful to confirm public/private ownership of generators — relevant for privatisation or restructuring scenarios."),

    (35, "GE_Georgia SOE DB.pdf",
         "Team/SOE Database/", "PDF",
         "PDF version of SOE database.",
         "-", NA, "Duplicate of xlsx."),

    # ── ENTERPRISE SURVEY ─────────────────────────────────────────────────
    (S, "Team/Enterprise Survey/", "", "", "", "", "", ""),
    (36, "GE_Enterprise Survey_Infrastructure Indicators.xlsx",
         "Team/Enterprise Survey/", "XLSX",
         "WB Enterprise Survey, Georgia, 2008/2013/2019/2023. Indicators: outage frequency, duration, economic losses, grid access.",
         "-", NA, "Context: reliability of supply for demand-side assumptions."),

    # ── SZILVIA FOLDER ────────────────────────────────────────────────────
    (S, "Team/Szilvia's Folder/", "", "", "", "", "", ""),
    (37, "Georgia Electricity Market Reform Project.docx",
         "Team/Szilvia's Folder/", "DOCX",
         "Market reform project document.",
         "-", N, "Context: market structure and reform trajectory — relevant for scenario design."),

    (38, "GeorgiaMarketAnalysisResultsWorkshopApril2023.pdf",
         "Team/Szilvia's Folder/", "PDF",
         "Market analysis results workshop, April 2023.",
         "-", N, "Recent market analysis — useful context before team meeting."),

    (39, "Georgia_CCDR_QER SD energy.docx",
         "Team/Szilvia's Folder/", "DOCX",
         "Energy chapter for Georgia CCDR (Country Climate & Development Report). Decarbonisation scenarios and policy context.",
         "-", N, "Strategic context — relevant for scenario framing."),

    (40, "Proposed Next Steps Toward a Wholesale Power Market.pptx",
         "Team/Szilvia's Folder/", "PPTX",
         "Roadmap for wholesale market opening in Georgia.",
         "-", N, "Context: market reform timeline."),

    (41, "Notes of Georgia market opening.docx",
         "Team/Szilvia's Folder/", "DOCX",
         "Notes on market opening discussions.",
         "-", NA, "Administrative notes."),

    (42, "Agenda Georgia wholesale market opening.docx",
         "Team/Szilvia's Folder/", "DOCX",
         "Meeting agenda.",
         "-", NA, "Administrative."),

    (43, "Villpiac_2015osz_ea_merged.pdf",
         "Team/Szilvia's Folder/", "PDF",
         "Hungarian document (~17 MB) — likely comparative market analysis.",
         "-", NA, "Not directly relevant."),

    # ── EPM GEORGIA MODELS ────────────────────────────────────────────────
    (S, "EPM_Georgia/ (WB model files)", "", "", "", "", "", ""),
    (44, "WB_EPM_v8_5.xlsb",
         "EPM_Georgia2022/Baseline/", "XLSB",
         "Full WB EPM v8.5 model, Georgia 2022 (Excel binary). Key sheets: GenAvailability (monthly CFs per plant), generator parameters (heat rate, VOM, FOM).",
         "pGenDataInput (heat rates: Mtkvari 10.3, Gardabani CCGT 6.93 MMBtu/MWh)\npAvailabilityCustom (7 reservoir hydro plants, monthly to quarterly CFs)",
         U, "Main calibration source for hydro availability and thermal heat rates."),

    (45, "Timeseries all data.xlsx",
         "EPM_Georgia/2022/1. Data/", "XLSX",
         "Hourly typical-year time series (~8,760 h) for RoR, OnshoreWind, PV as capacity factor profiles.",
         "pVREProfile (RoR, Wind, PV seasonal mean hourly profiles)",
         U, "Single typical year only. Wind CF ~0.27 vs Qartli actual ~0.46 — likely underestimates Qartli dispatch. To replace with Renewables Ninja multi-year."),

    (46, "RoR_Hourly.xlsx",
         "EPM_Georgia/2022/1. Data/", "XLSX",
         "Hourly RoR data from EPM 2022 study. Likely same source as Team/RoR_Hourly Historic Operating Data.xlsx.",
         "pAvailabilityCustom / pVREProfile",
         N, "Probable duplicate of Team/ RoR file — verify before using."),

    (47, "Wind_Hourly.xlsx",
         "EPM_Georgia/2022/1. Data/", "XLSX",
         "Hourly wind data from EPM 2022 study.",
         "pVREProfile (Wind)",
         N, "Probable duplicate of Timeseries data — verify."),

    (48, "Georgia EPM Modeling.pptx",
         "EPM_Georgia/2022/1. Data/", "PPTX",
         "Methodology presentation for WB EPM Georgia 2022 model.",
         "-", N, "Useful to understand assumptions behind v8.5 calibration."),

    (49, "WB_EPM_v8_2.xlsb",
         "EPM_Georgia/2021/Model/", "XLSB",
         "Earlier version of WB EPM model (v8.2, 2021).",
         "-", NA, "Superseded by v8.5."),

    # ── GEORGIA ROOT ─────────────────────────────────────────────────────
    (S, "Georgia/ (root)", "", "", "", "", "", ""),
    (50, "Georgian Power Market report March 2026.pdf",
         "Georgia/", "PDF",
         "Georgia electricity market report, March 2026 — most recent document in the entire folder. Likely: 2025 actual demand, generation mix, prices, ongoing projects, interconnection flows.",
         "All parameters — most recent comprehensive overview",
         N, "PRIORITY 1 — read before team meeting. Most valuable document for updating all assumptions."),

    (51, "Draft Proposal_Georgia Clean Energy...",
         "EPM_Georgia/", "DOCX",
         "Draft proposal related to Georgia clean energy (WB project document).",
         "-", N, "Policy context — relevant for scenario framing."),
]

# ── Write rows ───────────────────────────────────────────────────────────────
row_num = 2
for entry in rows:
    if entry[0] == S:
        ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=8)
        c = ws.cell(row=row_num, column=1, value=entry[1])
        c.font = section_font; c.fill = section_fill
        c.alignment = Alignment(horizontal="left", vertical="center")
        c.border = bdr
        ws.row_dimensions[row_num].height = 18
        row_num += 1
        continue

    num, fname, subfolder, ftype, content, epm_params, status, notes = entry

    if status == U:
        fill = used_fill
    elif status == P:
        fill = partial_fill
    elif status == N:
        fill = priority_fill if "PRIORITY" in notes else unused_fill
    else:
        fill = na_fill

    vals = [num, fname, subfolder, ftype, content, epm_params, status, notes]
    for col, val in enumerate(vals, 1):
        c = ws.cell(row=row_num, column=col, value=val)
        c.fill = fill
        c.alignment = center if col in (1, 3, 4, 7) else wrap
        c.border = bdr
        c.font = Font(size=9)
    ws.row_dimensions[row_num].height = 60
    row_num += 1

# ── Legend ────────────────────────────────────────────────────────────────────
row_num += 1
legend_items = [
    ("Legend", hdr_fill, Font(bold=True, color="FFFFFF", size=9)),
    ("USED — used in current EPM build", used_fill, Font(size=9)),
    ("PARTIAL — partially used, more data available in the file", partial_fill, Font(size=9)),
    ("NOT USED (PRIORITY) — not yet used, flagged for reading before team meeting", priority_fill, Font(size=9)),
    ("NOT USED — available but not needed for current build", unused_fill, Font(size=9)),
    ("N/A — not relevant for EPM inputs", na_fill, Font(size=9)),
]
for i, (label, fill, font) in enumerate(legend_items):
    ws.merge_cells(start_row=row_num+i, start_column=1, end_row=row_num+i, end_column=4)
    c = ws.cell(row=row_num+i, column=1, value=label)
    c.fill = fill; c.font = font
    c.alignment = Alignment(wrap_text=True, vertical="center")
    c.border = bdr
    ws.row_dimensions[row_num+i].height = 18

out_path = r"C:\Users\wb590892\Documents\EPM_Models\black_sea_2026\Data\Georgia\Georgia_Data_Index.xlsx"
wb.save(out_path)
print("Saved:", out_path)
