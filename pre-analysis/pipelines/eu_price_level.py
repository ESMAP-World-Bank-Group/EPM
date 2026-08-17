# -*- coding: utf-8 -*-
"""P3 - the annual level L(zext, y) of the EU price hypothesis, 2024-2053.

Companion to eu_price.py, which builds the hourly shape S. Together they give
P = L(zext,y) x S(q,d,t) x (1 - lambda) - W - C, with L carrying every scenario
and S carrying none.

Everything is assembled in *real EUR of 2024* and converted to USD 2024 once, at
the very end. Keeping the currency conversion to a single multiplication at the
final step means one number to audit rather than one per series.

Three facts about the source data drove the design, all established empirically
and all invisible if you simply read the parquet files:

  1. The ENTSO-E long series are not in a single currency. Romanian day-ahead
     prices are quoted in RON until 2021-05 and in EUR from 2021-07; Bulgarian
     prices are in BGN until 2021-12 and in EUR from 2022-02. Ratios against
     Hungary, which is native EUR throughout, sit at 4.6-4.9 and 1.75-2.00
     respectively before the switch and at ~1.0 after. Left uncorrected, the
     Romanian pre-crisis level comes out 4.7 times too high.
  2. The files named ...2024-12-31 carry a handful of 2025 rows (3 for Romania).
     They must be dropped or the last year is a three-hour average.
  3. The money year of the TYNDP marginal costs appears nowhere in the workbook,
     its readme, the implementation guidelines or ERAA. It is stated only in the
     TYNDP 2024 Scenarios Methodology Report (Final Version, January 2025), p.73:
     "The fuel and CO2 prices [...] with prices in real terms (in EUR 2022)",
     sourced from the IEA World Energy Outlook 2022 Announced Pledges Scenario.
     Hence the default below. It stays a parameter because the choice moves the
     2030 level by about 20 % and a future TYNDP edition will move the year.
"""
from __future__ import annotations

import argparse
import json
from pathlib import Path
from typing import Dict

import numpy as np
import pandas as pd

_PRE_ANALYSIS = Path(__file__).resolve().parents[1]

# Price zones of the EU hypothesis. The five non-EU external zones of
# pTradePrice.csv (Iran, Iraq, Kazakhstan, Russia, Syria) have no liquid hub, no
# TYNDP entry and no ETS; they need a cost-based basis and are handled apart.
ZONES: Dict[str, str] = {"RO": "Romania", "BG": "Bulgaria", "GR": "Greece"}

# TYNDP 2024 zone codes. GR00 is mainland Greece; GR03 is the island system and
# is not the interconnection-relevant hub.
TYNDP_CODES = {"Romania": "RO00", "Bulgaria": "BG00", "Greece": "GR00"}

# The three published TYNDP 2024 scenarios and the horizons each one covers.
# NT+ is the national-policies baseline; DE (Distributed Energy) and GA (Global
# Ambition) are the two decarbonisation variants. Both sit *below* NT on price,
# so the published family gives a downside spread and no upside case at all.
TYNDP_SCENARIOS = {"NT": (2030, 2040), "DE": (2035, 2040, 2050),
                   "GA": (2035, 2040, 2050)}

# Read the level off the hourly sheet, never off the 'Yearly Outputs' summary.
# In the DE and GA workbooks that summary row is populated for only 38 of 224
# columns and carries sentinel values in pseudo-zone columns; the hourly sheet is
# complete for all 222 zones. National Trends read here reproduces its own
# Yearly Outputs figures exactly, which is what proves the three scenarios are
# the same quantity.
HOURLY_SHEET = "Hourly Market Data emarket"
HOURLY_MC_ROW = 10        # 'Avg MC (excl. 3 000 Euro/MWh) [MWh]:'
HOURLY_COUNTRY_ROW = 12   # zone code under each value

# Every TYNDP release ships four workbooks per scenario and horizon: electricity,
# hydrogen, heat and synthetic fuels, and offshore. They share the file name stem
# and the sheet layout, so a loose glob picks up a hydrogen marginal cost without
# complaining. Only the unqualified electricity workbook is wanted.
OTHER_COMMODITIES = ("h2", "heat", "synthfuels", "offshore")

REFERENCE_ZONE = "HU"    # native EUR over the whole period, used to detect currency

# Local currency before the switch, detected against REFERENCE_ZONE.
FOREIGN_CCY = {"RO": ("eur_ron", 2.5), "BG": ("eur_bgn", 1.4)}

HORIZON = range(2024, 2054)


def _fail(msg: str) -> None:
    raise SystemExit(f"[eu_price_level] FAIL: {msg}")


# ── Deflators and rates ─────────────────────────────────────────────────────
def load_config(path: Path) -> pd.DataFrame:
    df = pd.read_csv(path, comment="#")
    need = {"series", "year", "value"}
    if not need.issubset(df.columns):
        _fail(f"{path.name}: expected columns {need}")
    return df


def _get(cfg: pd.DataFrame, series: str, year: int) -> float:
    m = cfg[(cfg["series"] == series) & (cfg["year"] == year)]
    if m.empty:
        _fail(f"{series} for {year} missing from the deflator config")
    return float(m["value"].iloc[0])


def to_real_eur(cfg: pd.DataFrame, value: float, from_year: int,
                base: int = 2024) -> float:
    """Nominal EUR of `from_year` -> real EUR of `base`, via euro-area HICP."""
    return value * _get(cfg, "eur_hicp", base) / _get(cfg, "eur_hicp", from_year)


# ── Observed history, currency-normalised ───────────────────────────────────
def load_long_series(cache_dir: Path, code: str) -> pd.Series:
    hits = sorted(cache_dir.glob(f"da_{code}_lt_*.parquet"))
    if not hits:
        _fail(f"no long series da_{code}_lt_*.parquet in {cache_dir}")
    df = pd.read_parquet(hits[0])
    s = df.iloc[:, 0]
    if not isinstance(s.index, pd.DatetimeIndex):
        _fail(f"{hits[0].name}: index is not a DatetimeIndex")
    if s.index.tz is None:
        _fail(f"{hits[0].name}: index is tz-naive, the market clock is unknown")
    s = s.tz_convert("UTC").sort_index()
    s.attrs["source"] = hits[0].name
    return s


def drop_partial_years(s: pd.Series, min_hours: int = 8000) -> pd.Series:
    """Drop any year that is not substantially complete (the stray 2025 rows)."""
    n = s.groupby(s.index.year).size()
    keep = set(n[n >= min_hours].index)
    dropped = {int(y): int(c) for y, c in n.items() if y not in keep}
    out = s[s.index.year.isin(keep)]
    out.attrs = dict(s.attrs)
    out.attrs["dropped_years"] = dropped
    return out


def normalise_currency(s: pd.Series, ref: pd.Series, code: str,
                       cfg: pd.DataFrame) -> pd.Series:
    """Convert the pre-switch stretch of a series from local currency to EUR.

    Classification is by the daily ratio to the reference market, smoothed with a
    7-day centred median so a single cheap day cannot flip the currency. The
    ratios are far apart - about 4.7 against 1.0 for RON, 1.96 against 1.0 for
    BGN - so the threshold sits in a wide empty gap rather than on a knife edge.
    """
    if code not in FOREIGN_CCY:
        s = s.copy()
        s.attrs["currency_note"] = "native EUR throughout"
        return s

    series_name, threshold = FOREIGN_CCY[code]
    d_loc = s.resample("D").mean()
    d_ref = ref.resample("D").mean()
    both = pd.concat([d_loc, d_ref], axis=1, keys=["loc", "ref"]).dropna()
    both = both[both["ref"] > 5.0]                     # ratio is meaningless near zero
    ratio = (both["loc"] / both["ref"]).rolling(7, center=True, min_periods=3).median()
    is_local = ratio > threshold

    # Days with no reference price inherit the neighbouring classification;
    # anything still unresolved is treated as euro, the safe default. Carried as
    # float rather than object so the fills stay numeric.
    flags = (is_local.reindex(d_loc.index).astype(float)
             .ffill().bfill().fillna(0.0) > 0.5)
    day_key = pd.Series(s.index.normalize(), index=s.index)
    mask = day_key.map(flags).fillna(False).to_numpy().astype(bool)

    out = s.astype(float).copy()
    years = pd.Index(out.index.year)
    if series_name == "eur_bgn":
        # A fixed peg: one rate, no year dimension.
        rate = np.full(out.shape, _get(cfg, "eur_bgn", 0))
    else:
        # Look the rate up only for the years actually converted. Requiring one
        # for every year in the file would demand a RON rate for 2024, long
        # after the quotation moved to euro.
        by_year = {int(y): _get(cfg, series_name, int(y))
                   for y in np.unique(years[mask])}
        rate = np.array([by_year.get(int(y), np.nan) for y in years])
    out.iloc[mask] = out.to_numpy()[mask] / rate[mask]

    conv = pd.Series(mask, index=s.index).groupby(s.index.year).mean()
    out.attrs = dict(s.attrs)
    out.attrs["currency_note"] = (
        f"{series_name} applied to {mask.sum()} of {mask.size} hours")
    out.attrs["share_converted_by_year"] = {int(k): round(float(v), 3)
                                            for k, v in conv.items()}
    last = pd.Series(mask, index=s.index)
    switched = last[last].index.max() if mask.any() else None
    out.attrs["last_local_currency_hour"] = str(switched) if switched else None
    return out


def gate_g3bis(annual: pd.DataFrame, ref_name: str) -> dict:
    """After conversion every zone must sit within a credible band of the reference.

    A residual factor of ~2 or ~4.7 is the signature of a missed currency stretch,
    and it is exactly the failure this whole step exists to prevent.
    """
    out, bad = {}, []
    for z in annual.columns:
        if z == ref_name:
            continue
        r = (annual[z] / annual[ref_name]).dropna()
        out[z] = {int(y): round(float(v), 3) for y, v in r.items()}
        off = {int(y): round(float(v), 2) for y, v in r.items()
               if not (0.5 <= v <= 2.0)}
        if off:
            bad.append(f"{z} {off}")
    out["pass"] = not bad
    if bad:
        out["offending"] = bad
    return out


# ── TYNDP forward points ────────────────────────────────────────────────────
def _hourly_header_rows(path: Path) -> Dict[int, Dict[int, object]]:
    """Rows 10 and 12 of the hourly sheet, keyed by their real column index.

    National Trends ships as .xlsx and the two decarbonisation scenarios as
    .xlsb, so the reader dispatches on the suffix. Both engines return the rows
    sparse, hence the explicit column index: reading them positionally silently
    misaligns every zone.
    """
    want = (HOURLY_MC_ROW, HOURLY_COUNTRY_ROW)
    rows: Dict[int, Dict[int, object]] = {}

    if path.suffix.lower() == ".xlsb":
        try:
            from pyxlsb import open_workbook
        except ImportError:
            _fail("pyxlsb is needed to read the DE and GA workbooks")
        with open_workbook(str(path)) as wb:
            sheet = HOURLY_SHEET if HOURLY_SHEET in wb.sheets else next(
                (s for s in wb.sheets if "hourly" in s.lower()), None)
            if sheet is None:
                _fail(f"{path.name}: no hourly sheet among {wb.sheets}")
            with wb.get_sheet(sheet) as ws:
                for i, row in enumerate(ws.rows(), 1):
                    if i in want:
                        rows[i] = {c.c: c.v for c in row}
                    if i >= max(want):
                        break
    else:
        try:
            import openpyxl
        except ImportError:
            _fail("openpyxl is needed to read the National Trends workbooks")
        wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
        sheet = HOURLY_SHEET if HOURLY_SHEET in wb.sheetnames else next(
            (s for s in wb.sheetnames if "hourly" in s.lower()), None)
        if sheet is None:
            wb.close()
            _fail(f"{path.name}: no hourly sheet among {wb.sheetnames}")
        for i, row in enumerate(wb[sheet].iter_rows(min_row=1, max_row=max(want),
                                                    values_only=True), 1):
            if i in want:
                rows[i] = dict(enumerate(row))
        wb.close()

    label = str(rows.get(HOURLY_MC_ROW, {}).get(1, ""))
    if "avg mc" not in label.lower():
        _fail(f"{path.name}: row {HOURLY_MC_ROW} reads {label!r}, not the "
              "average marginal cost line")
    return rows


def read_tyndp(tyndp_dir: Path, cache_csv: Path, log=print) -> pd.DataFrame:
    """Average marginal cost per zone for every published scenario and horizon.

    Values come out in TYNDP money (EUR 2022), one row per scenario x horizon x
    zone. Missing workbooks are reported and skipped rather than fatal: the
    scenarios actually present decide which trajectories can be built.
    """
    books, missing = {}, []
    for scen, horizons in TYNDP_SCENARIOS.items():
        for horizon in horizons:
            hits = [p for p in
                    sorted(tyndp_dir.glob(f"MMStandardOutputFile_{scen}{horizon}_*.xls*"))
                    if not any(t in p.stem.lower().split("_")
                               for t in OTHER_COMMODITIES)]
            if not hits:
                missing.append(f"{scen}{horizon}")
            elif len(hits) > 1:
                _fail(f"{scen}{horizon} is ambiguous: {[p.name for p in hits]}")
            else:
                books[(scen, horizon)] = hits[0]

    if not books:
        _fail(f"no TYNDP workbook found in {tyndp_dir}")
    if missing:
        log(f"  !! workbooks absent, scenarios degraded: {', '.join(missing)}")

    # A cache built before a workbook was downloaded would silently freeze the
    # ladder at its old shape, so it is only reused when it covers exactly the
    # scenarios and horizons now on disk.
    if cache_csv.exists():
        cached = pd.read_csv(cache_csv)
        have = {(str(s), int(y)) for s, y in
                zip(cached.get("scenario", []), cached.get("year", []))}
        if have == set(books):
            log(f"  TYNDP from cache {cache_csv.name}")
            return cached
        log(f"  cache {cache_csv.name} covers {sorted(have)} but the directory "
            f"now holds {sorted(books)} - rereading")

    rows = []
    for (scen, horizon), path in sorted(books.items()):
        grid = _hourly_header_rows(path)
        mc, ctry = grid[HOURLY_MC_ROW], grid[HOURLY_COUNTRY_ROW]
        by_code = {}
        for c, v in mc.items():
            if c < 2 or not isinstance(v, (int, float)):
                continue
            code = str(ctry.get(c, "") or "").strip()
            if code:
                by_code.setdefault(code, float(v))
        for zname, code in TYNDP_CODES.items():
            if code not in by_code:
                _fail(f"{code} absent from {path.name}")
            rows.append({"scenario": scen, "zone": zname, "year": horizon,
                         "tyndp_code": code, "value": by_code[code],
                         "source": path.name})
        log(f"  read {path.name} ({len(by_code)} zones)")

    df = pd.DataFrame(rows)
    cache_csv.parent.mkdir(parents=True, exist_ok=True)
    df.to_csv(cache_csv, index=False, float_format="%.4f")
    log(f"  cached {cache_csv}")
    return df


# ── Scenario construction ───────────────────────────────────────────────────
def _path_from_points(points: Dict[int, float]) -> pd.Series:
    """Linear between given years, flat beyond the last one."""
    ys = sorted(points)
    out = {}
    for y in HORIZON:
        if y <= ys[0]:
            out[y] = points[ys[0]]
        elif y >= ys[-1]:
            out[y] = points[ys[-1]]
        else:
            hi = next(k for k in ys if k >= y)
            lo = max(k for k in ys if k <= y)
            f = 0.0 if hi == lo else (y - lo) / (hi - lo)
            out[y] = points[lo] + f * (points[hi] - points[lo])
    return pd.Series(out, name="value")


# Which published TYNDP scenario carries which trajectory. Three of the five are
# traceable to an ENTSO-E deliverable; the two upper ones cannot be, because the
# published family holds no scenario above National Trends.
PUBLISHED_TRAJECTORIES = {"CENTRAL": "NT", "LOW": "DE", "VERY_LOW": "GA"}


def build_scenarios(anchor: float, published: Dict[str, Dict[int, float]],
                    crisis: float) -> Dict[str, pd.Series]:
    """Five trajectories, each defined by observed or published points only.

    CENTRAL   National Trends: the anchor normalises onto the published 2030 and
              2040 marginal costs of the national-policies baseline.
    LOW       Distributed Energy, 2035/2040/2050.
    VERY_LOW  Global Ambition, 2035/2040/2050 - the cheapest published future.
    HIGH      the geometric mirror of LOW about CENTRAL, year by year.
    EU_HIGH   the crisis regime persists: convergence by 2030 to the 2021-2023
              observed average, held to the horizon.

    The two upper trajectories have to be built here because ENTSO-E publishes
    nothing above National Trends: DE and GA are both decarbonisation routes, so
    the published family only opens downwards. Mirroring is the least arbitrary
    way to close it. Price uncertainty is symmetric in the log, not in the level,
    so if a halving is deemed plausible on the way down then a doubling is
    equally plausible on the way up; taking CENTRAL^2 / LOW makes the upper rung
    inherit the published spread instead of inventing one. EU_HIGH keeps a
    separate, empirical logic on purpose - see mirror_check for what happens when
    the two are compared.

    `published` holds the TYNDP points already in real EUR 2024, keyed by
    scenario then year. A scenario whose workbook was missing is skipped, so a
    partial download degrades the ladder instead of corrupting it.
    """
    start = min(HORIZON)
    out: Dict[str, pd.Series] = {}
    for name, scen in PUBLISHED_TRAJECTORIES.items():
        pts = published.get(scen) or {}
        if pts:
            out[name] = _path_from_points({start: anchor, **pts})
    if "CENTRAL" in out and "LOW" in out:
        out["HIGH"] = out["CENTRAL"] ** 2 / out["LOW"]
    else:
        # Without both published rungs there is nothing to mirror; fall back to
        # holding the anchor, which at least stays above CENTRAL.
        out["HIGH"] = _path_from_points({start: anchor})
    out["EU_HIGH"] = _path_from_points({start: anchor, 2030: crisis})
    return out


def mirror_check(scen: Dict[str, pd.Series], years=(2040, 2050)) -> dict:
    """Compare EU_HIGH against the mirror of VERY_LOW, which it is not built on.

    EU_HIGH comes from the observed 2021-2023 crisis; mirroring the cheapest
    published scenario is a wholly independent route to the same rung. Landing in
    the same place is a sanity check on the magnitude of the upper bound, not a
    proof of it, so this is reported and never gated.
    """
    if not {"CENTRAL", "VERY_LOW", "EU_HIGH"} <= set(scen):
        return {"available": False}
    out = {"available": True, "years": {}}
    for y in years:
        mirrored = float(scen["CENTRAL"][y] ** 2 / scen["VERY_LOW"][y])
        actual = float(scen["EU_HIGH"][y])
        out["years"][int(y)] = {
            "eu_high": actual, "mirror_of_very_low": mirrored,
            "dev_pct": (mirrored - actual) / actual * 100.0}
    return out


def gate_g3(central: pd.Series, observed_2023_real: float,
            tol_pct: float = 2.0) -> dict:
    """Back-cast 2023 from the 2024-2030 slope and compare to the observed year.

    The CENTRAL slope out of the anchor is nearly flat, so this is in substance a
    test of whether the anchor year is representative of the current regime: a
    market that moved sharply between 2023 and 2024 cannot pass it under any
    single-year anchor. That is the intended reading, not a defect of the test.
    """
    slope = (central[2030] - central[2024]) / 6.0
    back = central[2024] - slope
    dev = (back - observed_2023_real) / observed_2023_real * 100.0
    return {"back_cast_2023": back, "observed_2023_real_eur2024": observed_2023_real,
            "dev_pct": dev, "tolerance_pct": tol_pct,
            "pass": bool(abs(dev) <= tol_pct)}


# What the design actually claims, and therefore all that can be tested. The
# three published rungs are claimed to be ordered. The two upper ones are each
# claimed to sit above CENTRAL, but *not* to be ordered against each other: they
# are two unrelated upside stories - a mirrored published spread and a return of
# the observed crisis - and nothing makes one systematically the more extreme.
# Ranking them would be a claim about the world that neither construction backs.
ORDERED_LADDER = ["VERY_LOW", "LOW", "CENTRAL"]
UPPER_TRAJECTORIES = ["HIGH", "EU_HIGH"]


def gate_g3ter(scen: Dict[str, pd.Series], tol_pct: float = 1.0) -> dict:
    """Check the trajectory names actually rank the way they say they do.

    DE and GA are two decarbonisation routes, not a low and a lower case, so
    nothing upstream guarantees one stays below the other. A crossing worth less
    than `tol_pct` is a tie and the labels survive it; a larger one means the
    names are lying and the mapping has to be revisited.
    """
    chain = [n for n in ORDERED_LADDER if n in scen]
    pairs = list(zip(chain, chain[1:]))
    pairs += [(chain[-1], u) for u in UPPER_TRAJECTORIES
              if u in scen and chain]

    worst, breaches = 0.0, []
    for lo, hi in pairs:
        d = (scen[lo] - scen[hi]) / scen[hi] * 100.0   # >0 means lo is above hi
        bad = d[d > tol_pct]
        if len(bad):
            breaches.append({"below": lo, "above": hi,
                             "years": [int(y) for y in bad.index],
                             "max_excess_pct": float(bad.max())})
        worst = max(worst, float(d.max()))
    return {"ordered_chain": chain, "upper_unranked": UPPER_TRAJECTORIES,
            "tested_pairs": [f"{lo}<={hi}" for lo, hi in pairs],
            "tolerance_pct": tol_pct, "max_inversion_pct": worst,
            "breaches": breaches, "pass": not breaches}


def anchor_sensitivity(col: pd.Series, alt_years,
                       published: Dict[str, Dict[int, float]], crisis: float,
                       base: Dict[str, pd.Series]) -> dict:
    """How much of the answer rests on the choice of anchor year.

    G3 measures the anchor against one past year; this measures the trajectories
    against an alternative anchor. The two say different things: CENTRAL forgets
    its anchor once it reaches the TYNDP points, whereas HIGH carries it to the
    horizon, so the same anchor doubt is transient in one scenario and permanent
    in another.
    """
    yrs = [y for y in alt_years if y in col.index]
    if not yrs:
        return {"available": False}
    alt = float(col[yrs].mean())
    other = build_scenarios(alt, published, crisis)
    out = {"available": True, "alt_anchor_years": yrs, "alt_anchor": alt,
           "scenarios": {}}
    for name, path in base.items():
        d = (other[name] - path) / path * 100.0
        out["scenarios"][name] = {
            "dev_2026_pct": float(d[2026]), "dev_2030_pct": float(d[2030]),
            "dev_2040_pct": float(d[2040]),
            "mean_abs_dev_pct": float(d.abs().mean())}
    return out


# ── Driver ──────────────────────────────────────────────────────────────────
def run(cache_dir: Path, tyndp_dir: Path, config: Path, out_dir: Path,
        tyndp_money_year: int, anchor_years, alt_anchor_years, log=print) -> dict:
    cfg = load_config(config)
    out_dir.mkdir(parents=True, exist_ok=True)

    ref_raw = drop_partial_years(load_long_series(cache_dir, REFERENCE_ZONE))

    nominal, qc_hist = {}, {}
    for code, zname in ZONES.items():
        s = drop_partial_years(load_long_series(cache_dir, code))
        s = normalise_currency(s, ref_raw, code, cfg)
        nominal[zname] = s.groupby(s.index.year).mean()
        qc_hist[zname] = {k: s.attrs.get(k) for k in
                          ("source", "currency_note", "share_converted_by_year",
                           "last_local_currency_hour", "dropped_years")}
    nominal[REFERENCE_ZONE] = ref_raw.groupby(ref_raw.index.year).mean()
    annual_nom = pd.DataFrame(nominal)

    g3bis = gate_g3bis(annual_nom, REFERENCE_ZONE)
    log("  currency check " + ("PASS" if g3bis["pass"] else "FAIL " + str(g3bis["offending"])))

    real = annual_nom.apply(lambda col: pd.Series(
        {y: to_real_eur(cfg, v, int(y)) for y, v in col.dropna().items()}))
    log("\n  observed annual means, real EUR 2024")
    log(real[list(ZONES.values())].round(1).to_string())

    tyndp = read_tyndp(tyndp_dir, out_dir / "tyndp2024_marginal_cost.csv", log)

    rows, qc = [], {"tyndp_money_year": tyndp_money_year,
                    "anchor_years": list(anchor_years),
                    "currency_gate_G3bis": g3bis, "history": qc_hist, "zones": {}}

    for zname in ZONES.values():
        col = real[zname].dropna()
        anchor = float(col[list(anchor_years)].mean())
        obs23 = float(col[2023])
        pre = [y for y in (2015, 2016, 2017, 2018, 2019) if y in col.index]
        precrisis = float(col[pre].mean())
        crisis = float(col[[y for y in (2021, 2022, 2023) if y in col.index]].mean())

        z_tyndp = tyndp[tyndp["zone"] == zname]
        published, as_published = {}, {}
        for scen, g in z_tyndp.groupby("scenario"):
            published[scen] = {int(r.year): to_real_eur(cfg, float(r.value),
                                                        tyndp_money_year)
                               for r in g.itertuples()}
            as_published[scen] = {int(r.year): float(r.value) for r in g.itertuples()}

        scen = build_scenarios(anchor, published, crisis)
        g3 = gate_g3(scen["CENTRAL"], obs23)
        g3ter = gate_g3ter(scen)
        mirror = mirror_check(scen)
        sens = anchor_sensitivity(col, alt_anchor_years, published, crisis, scen)

        qc["zones"][zname] = {
            "anchor_real_eur2024": anchor,
            "observed_2023_real_eur2024": obs23,
            "observed_2024_real_eur2024": float(col[2024]),
            # No longer a trajectory - kept as the benchmark the published LOW
            # and VERY_LOW are worth checking against.
            "precrisis_2015_2019_real": precrisis if pre else None,
            "crisis_2021_2023_real": crisis,
            "tyndp_real_eur2024": published,
            "tyndp_as_published": as_published,
            "G3_backcast": g3,
            "G3ter_ladder_order": g3ter,
            "upper_bound_mirror_check": mirror,
            "anchor_sensitivity": sens,
            "scenario_2030": {k: float(v[2030]) for k, v in scen.items()},
            "scenario_2050": {k: float(v[2050]) for k, v in scen.items()},
        }
        log(f"\n  {zname}: anchor {anchor:6.1f} | pre-crisis {precrisis:6.1f} | "
            f"crisis {crisis:6.1f} | G3 {'PASS' if g3['pass'] else 'FAIL'} "
            f"{g3['dev_pct']:+.1f} % | G3ter "
            f"{'PASS' if g3ter['pass'] else 'FAIL'} "
            f"(worst inversion {g3ter['max_inversion_pct']:+.2f} %)")
        for b in g3ter["breaches"]:
            log(f"      !! {b['below']} sits above {b['above']} by up to "
                f"{b['max_excess_pct']:.1f} % in {b['years'][0]}-{b['years'][-1]}")
        if mirror.get("available"):
            log("      upper bound cross-check: " + "  ".join(
                f"{y} EU_HIGH {v['eu_high']:.0f} vs mirror {v['mirror_of_very_low']:.0f}"
                f" ({v['dev_pct']:+.0f} %)" for y, v in mirror["years"].items()))
        for s in sorted(published):
            log("      TYNDP %-3s real EUR2024: %s" % (s, "  ".join(
                f"{y} {v:6.1f}" for y, v in sorted(published[s].items()))))
        if sens.get("available"):
            log(f"      anchor sensitivity vs {sens['alt_anchor_years']} "
                f"({sens['alt_anchor']:.1f}): " + " ".join(
                    f"{k} {v['mean_abs_dev_pct']:.1f}%" for k, v in
                    sens["scenarios"].items()))
        for name, path in scen.items():
            for y, v in path.items():
                rows.append({"zone": zname, "scenario": name, "year": int(y),
                             "L_eur2024": float(v)})

    lvl = pd.DataFrame(rows)
    rate = _get(cfg, "eur_usd", 2024)
    lvl["L_usd2024"] = lvl["L_eur2024"] * rate
    qc["eur_usd_2024"] = rate

    lvl.to_csv(out_dir / "level_L.csv", index=False, float_format="%.3f")
    real.to_csv(out_dir / "observed_annual_real.csv", float_format="%.3f")
    (out_dir / "qc_level_L.json").write_text(json.dumps(qc, indent=2, default=str),
                                             encoding="utf-8")
    log(f"\n  wrote {out_dir / 'level_L.csv'} ({len(lvl)} rows)")
    log(f"  wrote {out_dir / 'observed_annual_real.csv'}")
    log(f"  wrote {out_dir / 'qc_level_L.json'}")

    qc["failed"] = ([] if g3bis["pass"] else ["G3bis currency"]) + \
                   [f"{z} G3" for z, v in qc["zones"].items()
                    if not v["G3_backcast"]["pass"]] + \
                   [f"{z} G3ter" for z, v in qc["zones"].items()
                    if not v["G3ter_ladder_order"]["pass"]]
    return qc


def plot_qc(out_dir: Path, log=print) -> None:
    try:
        import matplotlib
        matplotlib.use("Agg")
        import matplotlib.pyplot as plt
    except ImportError:
        log("  (matplotlib absent - skipping the figure)")
        return

    lvl = pd.read_csv(out_dir / "level_L.csv")
    hist = pd.read_csv(out_dir / "observed_annual_real.csv", index_col=0)
    zones = list(ZONES.values())
    fig, axes = plt.subplots(1, len(zones), figsize=(4.4 * len(zones), 3.6), sharey=True)
    for ax, z in zip(np.atleast_1d(axes), zones):
        if z in hist.columns:
            h = hist[z].dropna()
            ax.plot(h.index, h.values, color="0.25", lw=1.4, label="observed (real EUR 2024)")
        for name, g in lvl[lvl["zone"] == z].groupby("scenario"):
            g = g.sort_values("year")
            ax.plot(g["year"], g["L_eur2024"], lw=1.3, ls="--", label=name)
        ax.set_title(z)
        ax.set_xlabel("year")
        ax.grid(alpha=.3)
    np.atleast_1d(axes)[0].set_ylabel("L, real EUR 2024 per MWh")
    np.atleast_1d(axes)[0].legend(fontsize=7)
    fig.tight_layout()
    fig.savefig(out_dir / "qc_level_L.png", dpi=140)
    log(f"  wrote {out_dir / 'qc_level_L.png'}")


def main(argv=None) -> int:
    ap = argparse.ArgumentParser(description="Build L(zext,y), the EU price level.")
    ap.add_argument("--cache-dir", type=Path,
                    default=_PRE_ANALYSIS.parents[1] / "Data" / "cache_entso_e")
    ap.add_argument("--tyndp-dir", type=Path,
                    default=_PRE_ANALYSIS.parents[1] / "Data" / "TYNDP" / "2024")
    ap.add_argument("--config", type=Path,
                    default=_PRE_ANALYSIS / "config" / "price_deflators.csv")
    ap.add_argument("--out-dir", type=Path, default=_PRE_ANALYSIS / "output_prices")
    ap.add_argument("--tyndp-money-year", type=int, default=2022,
                    help="money year of the TYNDP marginal costs; EUR2022 per "
                         "the TYNDP 2024 Methodology Report p.73")
    ap.add_argument("--anchor-years", default="2024",
                    help="observed years averaged to set the starting level")
    ap.add_argument("--alt-anchor-years", default="2023,2024",
                    help="alternative anchor, reported as a sensitivity only")
    ap.add_argument("--no-plot", action="store_true")
    args = ap.parse_args(argv)

    years = [int(y) for y in args.anchor_years.split(",") if y.strip()]
    alt = [int(y) for y in args.alt_anchor_years.split(",") if y.strip()]

    print(f"[eu_price_level] TYNDP money year assumed {args.tyndp_money_year}; "
          f"anchor {years}")
    qc = run(args.cache_dir, args.tyndp_dir, args.config, args.out_dir,
             args.tyndp_money_year, years, alt)
    if not args.no_plot:
        plot_qc(args.out_dir)
    if qc["failed"]:
        print("[eu_price_level] gates not met: " + ", ".join(qc["failed"]))
        return 1
    print("[eu_price_level] all gates passed")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
