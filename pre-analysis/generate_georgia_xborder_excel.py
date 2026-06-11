"""
Generate Georgia_CrossBorder_Trade.xlsx
  Tab 1: Data  — structured table (imports seasonal, exports annual, transit, net position)
  Tab 2: Charts — 3 charts side by side
"""
import openpyxl
from openpyxl.chart import BarChart, LineChart, AreaChart, Reference
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

OUT = r'C:\Users\wb590892\Documents\EPM_Models\black_sea_2026\Data\Georgia\Georgia_CrossBorder_Trade.xlsx'

# ── palette ──────────────────────────────────────────────────────────────────
C = {
    'AZ':    'FF4472C4',  # blue
    'RU':    'FFC00000',  # dark red
    'TR':    'FFED7D31',  # orange
    'ARM':   'FF70AD47',  # green
    'TOTAL': 'FF767171',  # grey
    'hdr':   'FF1F4E79',  # navy (section headers)
    'sub':   'FF2E75B6',  # medium blue (column headers)
    'imp':   'FFD6E4F0',  # light blue tint (imports rows)
    'exp':   'FFFFF2CC',  # light yellow (exports rows)
    'tra':   'FFE2EFDA',  # light green (transit rows)
    'net':   'FFFCE4D6',  # light orange (net rows)
    'total_row': 'FFBDD7EE',
    'annuel_row': 'FFB8CCE4',
    'note':  'FF808080',
}

def fill(hex_color):
    return PatternFill(start_color=hex_color, end_color=hex_color, fill_type='solid')

def hdr_font(size=11, bold=True, color='FFFFFFFF'):
    return Font(bold=bold, color=color, size=size, name='Calibri')

def normal_font(size=10, bold=False, color='FF000000'):
    return Font(bold=bold, color=color, size=size, name='Calibri')

def center():
    return Alignment(horizontal='center', vertical='center', wrap_text=True)

def left():
    return Alignment(horizontal='left', vertical='center', wrap_text=False)

thin = Side(style='thin', color='FFB0B0B0')
med  = Side(style='medium', color='FF808080')
thin_border = Border(left=thin, right=thin, top=thin, bottom=thin)
med_border  = Border(left=med,  right=med,  top=med,  bottom=med)

def set_cell(ws, row, col, value, bg=None, font=None, align=None, border=None, num_fmt=None):
    c = ws.cell(row=row, column=col, value=value)
    if bg:     c.fill    = fill(bg)
    if font:   c.font    = font
    if align:  c.alignment = align
    if border: c.border  = border
    if num_fmt: c.number_format = num_fmt
    return c

# ── data ─────────────────────────────────────────────────────────────────────
YEARS_5 = [2020, 2021, 2022, 2023, 2024]   # import/export/transit cols
YEARS_7 = [2018, 2019, 2020, 2021, 2022, 2023, 2024]  # net position

IMPORTS = {
    # country: {season: [2020, 2021, 2022, 2023, 2024]}
    'Azerbaijan': {
        'Summer (Apr–Sep)': [47,   52,   10,    1,  None],
        'Winter (Oct–Mar)': [679,  558,  120,    2,  None],
        'Annual':           [726,  610,  130,  790,   184],
    },
    'Russia': {
        'Summer (Apr–Sep)': [121,  177,   70,    0,  None],
        'Winter (Oct–Mar)': [450, 1059,  977,  599,  None],
        'Annual':           [571, 1236, 1047,  787,   970],
    },
    'Turkey': {
        'Summer (Apr–Sep)': [14,   50,    0,    0,  None],
        'Winter (Oct–Mar)': [299,  111,    0,    0,  None],
        'Annual':           [313,  161,    0,    0,    74],
    },
    'Armenia': {
        'Summer (Apr–Sep)': [0,     0,    0,    0,  None],
        'Winter (Oct–Mar)': [0,     0,  139,    0,  None],
        'Annual':           [0,     0,  139,    0,     0],
    },
    'TOTAL': {
        'Summer (Apr–Sep)': [182,  279,   80,    1,  None],
        'Winter (Oct–Mar)': [1428, 1728, 1236,  789,  None],
        'Annual':           [1610, 2006, 1533,  790,  1228],
    },
}

EXPORTS = {
    # country: [2020, 2021, 2022, 2023, 2024]  — annual only
    'Turkey':             [None, None,  865, None,   42],
    'Armenia':            [None, None,   97, None,    2],
    'Azerbaijan':         [None, None,    9, None,  136],
    'Abkhazia (Russia)¹': [None, None,    0, None,  869],
    'TOTAL':              [ 154,  391,  971, 1468,  1047],
}

TRANSIT = {
    'Azerbaijan → Turkey': [102,  721, 2219, 3105, None],
    'Russia → Turkey':     [  0,  322,  269,  194, None],
    'Russia → Armenia':    [  0,  141,  156,    0, None],
    'Armenia → Turkey':    [  0,    0,    0,  145, None],
    'TOTAL':               [102, 1184, 2644, 3444, None],
}

NET_IMPORTS = [1509, 1627, 1610, 2006, 1533,  790, 1228]
NET_EXPORTS = [ 589,  243,  154,  391,  971, 1468, 1047]
NET_BALANCE = [i - e for i, e in zip(NET_IMPORTS, NET_EXPORTS)]

# ─────────────────────────────────────────────────────────────────────────────
# TAB 1: DATA
# ─────────────────────────────────────────────────────────────────────────────
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Data"

# column widths
ws.column_dimensions['A'].width = 26
ws.column_dimensions['B'].width = 18
for col in ['C','D','E','F','G']:
    ws.column_dimensions[col].width = 10

def write_section(ws, start_row, title, years, data_dict, bg_tint, country_colors,
                  show_season=True, note=None):
    """
    Write a section block.  Returns next free row.
    data_dict:  {label: [values]}  OR {country: {season: [values]}}  (when show_season=True)
    """
    r = start_row
    ncols = 2 + len(years)

    # section title
    ws.row_dimensions[r].height = 18
    c = ws.cell(row=r, column=1, value=title)
    c.fill = fill(C['hdr']); c.font = hdr_font(12); c.alignment = left()
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncols)
    r += 1

    # column headers
    ws.row_dimensions[r].height = 16
    set_cell(ws, r, 1, 'Border / Corridor', bg=C['sub'],
             font=hdr_font(10), align=center(), border=thin_border)
    if show_season:
        set_cell(ws, r, 2, 'Season', bg=C['sub'],
                 font=hdr_font(10), align=center(), border=thin_border)
    else:
        set_cell(ws, r, 2, '', bg=C['sub'],
                 font=hdr_font(10), align=center(), border=thin_border)

    for i, yr in enumerate(years):
        label = str(yr)
        if yr == 2022 and show_season:
            label = '2022*'
        set_cell(ws, r, 3+i, label, bg=C['sub'],
                 font=hdr_font(10), align=center(), border=thin_border)
    r += 1

    for label, data in data_dict.items():
        is_total = label == 'TOTAL'
        country_bg = country_colors.get(label, bg_tint)

        if show_season and isinstance(data, dict):
            seasons = data
            for si, (season, vals) in enumerate(seasons.items()):
                is_annuel = season == 'Annual'
                row_bg = C['annuel_row'] if is_total and is_annuel else \
                         C['total_row'] if is_total else \
                         bg_tint
                ws.row_dimensions[r].height = 15
                # Country label — merge across season rows? Just write on first season row
                if si == 0:
                    ws.cell(row=r, column=1).value = label
                    # merge the 3 season rows for country label
                    ws.merge_cells(start_row=r, start_column=1,
                                   end_row=r+2, end_column=1)
                    ws.cell(row=r, column=1).alignment = Alignment(
                        horizontal='center', vertical='center')
                    ws.cell(row=r, column=1).font = normal_font(10, bold=is_total)
                    ws.cell(row=r, column=1).fill = fill(country_bg)
                    ws.cell(row=r, column=1).border = thin_border

                # Season label
                set_cell(ws, r, 2, season,
                         bg='FFEFEFEF' if is_annuel else row_bg,
                         font=normal_font(9, bold=is_annuel, color='FF404040' if is_annuel else 'FF000000'),
                         align=left(), border=thin_border)
                # Values
                for i, val in enumerate(vals):
                    cell_bg = 'FFEFEFEF' if is_annuel else row_bg
                    if val is None:
                        disp = '—'
                        set_cell(ws, r, 3+i, disp, bg='FFF2F2F2',
                                 font=normal_font(9, color='FFAAAAAA'),
                                 align=center(), border=thin_border)
                    else:
                        set_cell(ws, r, 3+i, val, bg=cell_bg,
                                 font=normal_font(10, bold=is_total and is_annuel),
                                 align=center(), border=thin_border,
                                 num_fmt='#,##0')
                r += 1
        else:
            # flat row
            ws.row_dimensions[r].height = 15
            row_bg = C['total_row'] if is_total else bg_tint
            set_cell(ws, r, 1, label, bg=country_bg,
                     font=normal_font(10, bold=is_total), align=left(), border=thin_border)
            set_cell(ws, r, 2, '', bg=row_bg, border=thin_border)
            for i, val in enumerate(data):
                if val is None:
                    set_cell(ws, r, 3+i, '—', bg='FFF2F2F2',
                             font=normal_font(9, color='FFAAAAAA'),
                             align=center(), border=thin_border)
                else:
                    set_cell(ws, r, 3+i, val, bg=row_bg,
                             font=normal_font(10, bold=is_total),
                             align=center(), border=thin_border,
                             num_fmt='#,##0')
            r += 1

    if note:
        ws.cell(row=r, column=1).value = note
        ws.cell(row=r, column=1).font = Font(italic=True, size=8, color='FF808080')
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncols)
        r += 1

    r += 1  # blank row
    return r

# ── Title ────────────────────────────────────────────────────────────────────
ws.row_dimensions[1].height = 26
ws.merge_cells('A1:G1')
c = ws['A1']
c.value = "Georgia — Cross-Border Electricity Trade (GWh)"
c.fill = fill(C['hdr']); c.font = hdr_font(14); c.alignment = center()

ws.row_dimensions[2].height = 13
ws.merge_cells('A2:G2')
ws['A2'].value = ("Sources: GSE/GENEX Generation Balance PDFs 2020–2023 (imports) · "
                  "GNERC Annual Report 2024 (exports, net position)")
ws['A2'].font = Font(italic=True, size=8, color='FF595959')
ws['A2'].alignment = left()

# ── IMPORTS section ──────────────────────────────────────────────────────────
country_import_colors = {
    'Azerbaijan': 'FFDAE6F5',
    'Russia':     'FFFCE4D6',
    'Turkey':     'FFFDEBD6',
    'Armenia':    'FFE9F5E3',
    'TOTAL':      'FFBDD7EE',
}

next_row = write_section(
    ws, start_row=4,
    title='IMPORTS  (flows into Georgia)',
    years=YEARS_5,
    data_dict=IMPORTS,
    bg_tint=C['imp'],
    country_colors=country_import_colors,
    show_season=True,
    note='* 2022: data Jan–Nov only; GNERC annual = 1,533 GWh. 2024: seasonal data unavailable (— = n/a)'
)

# ── EXPORTS section ──────────────────────────────────────────────────────────
country_export_colors = {
    'Turkey':             'FFFDEBD6',
    'Armenia':            'FFE9F5E3',
    'Azerbaijan':         'FFDAE6F5',
    'Abkhazia (Russia)¹': 'FFFCE4D6',
    'TOTAL':              C['total_row'],
}
next_row = write_section(
    ws, start_row=next_row,
    title='EXPORTS  (flows out of Georgia, annual only)',
    years=YEARS_5,
    data_dict=EXPORTS,
    bg_tint=C['exp'],
    country_colors=country_export_colors,
    show_season=False,
    note=('¹ GNERC 2024 classifies 83% of exports as "Russia" = Georgian power supplied to '
          'Abkhazia (occupied territory). Listed separately from Russian imports.')
)

# ── TRANSIT section ──────────────────────────────────────────────────────────
transit_colors = {
    'Azerbaijan → Turkey': 'FFDAE6F5',
    'Russia → Turkey':     'FFFCE4D6',
    'Russia → Armenia':    'FFFCE4D6',
    'Armenia → Turkey':    'FFE9F5E3',
    'TOTAL':               C['total_row'],
}
next_row = write_section(
    ws, start_row=next_row,
    title='TRANSIT  (flows through Georgia — does not affect national balance)',
    years=YEARS_5,
    data_dict=TRANSIT,
    bg_tint=C['tra'],
    country_colors=transit_colors,
    show_season=False,
    note='2024: transit data not yet available.'
)

# ── NET POSITION section ─────────────────────────────────────────────────────
r = next_row
ws.row_dimensions[r].height = 18
c = ws.cell(row=r, column=1,
            value='NET POSITION  (imports – exports, annual  ·  source GNERC 2024)')
c.fill = fill(C['hdr']); c.font = hdr_font(12); c.alignment = left()
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=9)
r += 1

ws.row_dimensions[r].height = 16
set_cell(ws, r, 1, 'Indicator', bg=C['sub'], font=hdr_font(10), align=center(), border=thin_border)
set_cell(ws, r, 2, '',           bg=C['sub'], font=hdr_font(10), align=center(), border=thin_border)
for i, yr in enumerate(YEARS_7):
    set_cell(ws, r, 3+i, str(yr), bg=C['sub'], font=hdr_font(10), align=center(), border=thin_border)
r += 1

net_rows = [
    ('Total Imports',              NET_IMPORTS, C['imp']),
    ('Total Exports',              NET_EXPORTS, C['exp']),
    ('Net Position  (+ = deficit)', NET_BALANCE, C['net']),
]
for label, vals, bg in net_rows:
    is_net = 'Net' in label
    ws.row_dimensions[r].height = 15
    set_cell(ws, r, 1, label, bg=bg, font=normal_font(10, bold=is_net), align=left(),  border=thin_border)
    set_cell(ws, r, 2, '',    bg=bg, border=thin_border)
    for i, val in enumerate(vals):
        # highlight negative net in green, large positive in red
        cell_bg = bg
        font_color = '00000000'
        if is_net:
            if val < 0:
                cell_bg = 'FF70AD47'; font_color = 'FFFFFFFF'  # green = net exporter
            elif val > 1000:
                cell_bg = 'FFFFD966'  # yellow = large deficit
        set_cell(ws, r, 3+i, val, bg=cell_bg,
                 font=normal_font(10, bold=is_net, color=font_color),
                 align=center(), border=thin_border, num_fmt='#,##0')
    r += 1

r += 1  # blank

# freeze panes
ws.freeze_panes = 'C5'

# ─────────────────────────────────────────────────────────────────────────────
# TAB 2: CHARTS
# ─────────────────────────────────────────────────────────────────────────────
ws2 = wb.create_sheet("Charts")
ws2.sheet_view.showGridLines = False

# ── Helper data tables ────────────────────────────────────────────────────────

# ---- Chart 1 helper: Seasonal imports 2020–2023 ----
# Rows 1-10: category labels + 4 series (AZ, RU, TR, ARM)
# Categories: 2020 Été | 2020 Hiver | 2021 Été | 2021 Hiver | ... (8 rows)
C1_START = 1  # row
CATEGORIES_1 = ['2020 Summer','2020 Winter','2021 Summer','2021 Winter',
                '2022 Summer','2022 Winter','2023 Summer','2023 Winter']
C1_DATA = {  # country: [2020S, 2020W, 2021S, 2021W, 2022S, 2022W, 2023S, 2023W]
    'AZ':  [47,  679,  52,  558, 10,  120,  1,   2  ],
    'RU':  [121, 450, 177, 1059, 70,  977,  0, 599  ],
    'TR':  [14,  299,  50,  111,  0,    0,  0,   0  ],
    'ARM': [0,     0,   0,    0,  0,  139,  0,   0  ],
}

ws2.cell(row=C1_START, column=1).value = 'Cat'
for i, cat in enumerate(CATEGORIES_1):
    ws2.cell(row=C1_START+1+i, column=1).value = cat
ws2.cell(row=C1_START, column=2).value = 'AZ'
ws2.cell(row=C1_START, column=3).value = 'RU'
ws2.cell(row=C1_START, column=4).value = 'TR'
ws2.cell(row=C1_START, column=5).value = 'ARM'
for ci, (country, vals) in enumerate(C1_DATA.items()):
    for ri, val in enumerate(vals):
        ws2.cell(row=C1_START+1+ri, column=2+ci).value = val

# ---- Chart 2 helper: Annual balance 2018–2024 ----
C2_START = 12
ws2.cell(row=C2_START,   column=1).value = 'Year'
ws2.cell(row=C2_START,   column=2).value = 'Imports'
ws2.cell(row=C2_START,   column=3).value = 'Exports'
ws2.cell(row=C2_START,   column=4).value = 'Net Position'
for i, yr in enumerate(YEARS_7):
    ws2.cell(row=C2_START+1+i, column=1).value = yr
    ws2.cell(row=C2_START+1+i, column=2).value = NET_IMPORTS[i]
    ws2.cell(row=C2_START+1+i, column=3).value = NET_EXPORTS[i]
    ws2.cell(row=C2_START+1+i, column=4).value = NET_BALANCE[i]

# ---- Chart 3 helper: Transit 2020–2023 ----
C3_START = 22
TRANSIT_SERIES = {
    'AZ→TR':  [102,  721, 2219, 3105],
    'RU→TR':  [  0,  322,  269,  194],
    'RU→ARM': [  0,  141,  156,    0],
    'ARM→TR': [  0,    0,    0,  145],
}
ws2.cell(row=C3_START, column=1).value = 'Year'
for i, s in enumerate(TRANSIT_SERIES.keys()):
    ws2.cell(row=C3_START, column=2+i).value = s
for i, yr in enumerate([2020, 2021, 2022, 2023]):
    ws2.cell(row=C3_START+1+i, column=1).value = yr
    for j, vals in enumerate(TRANSIT_SERIES.values()):
        ws2.cell(row=C3_START+1+i, column=2+j).value = vals[i]

# ── Chart 1: Seasonal imports stacked bar ────────────────────────────────────
chart1 = BarChart()
chart1.type    = 'col'
chart1.grouping = 'stacked'
chart1.overlap = 100
chart1.title   = 'Seasonal Imports by Country (2020–2023)'
chart1.y_axis.title = 'GWh'
chart1.x_axis.title = ''
chart1.legend.position = 'b'
chart1.style   = 10
chart1.width   = 16
chart1.height  = 12

cats1 = Reference(ws2, min_col=1, min_row=C1_START+1, max_row=C1_START+8)
series_colors = ['FF4472C4', 'FFC00000', 'FFED7D31', 'FF70AD47']
for ci, (label, col_idx, color) in enumerate([
    ('Azerbaijan', 2, 'FF4472C4'),
    ('Russia',     3, 'FFC00000'),
    ('Turkey',     4, 'FFED7D31'),
    ('Armenia',    5, 'FF70AD47'),
]):
    data_ref = Reference(ws2, min_col=col_idx, min_row=C1_START, max_row=C1_START+8)
    series = openpyxl.chart.Series(data_ref, title_from_data=True)
    series.graphicalProperties.solidFill = color[2:]  # strip 'FF'
    series.graphicalProperties.line.solidFill = color[2:]
    chart1.series.append(series)
chart1.set_categories(cats1)

# ── Chart 2: Annual balance bar + line ───────────────────────────────────────
chart2 = BarChart()
chart2.type     = 'col'
chart2.grouping = 'clustered'
chart2.title    = 'Annual Import / Export Balance (2018–2024)'
chart2.y_axis.title = 'GWh'
chart2.x_axis.title = ''
chart2.legend.position = 'b'
chart2.style    = 10
chart2.width    = 16
chart2.height   = 12

cats2 = Reference(ws2, min_col=1, min_row=C2_START+1, max_row=C2_START+7)
imp_data = Reference(ws2, min_col=2, min_row=C2_START, max_row=C2_START+7)
exp_data = Reference(ws2, min_col=3, min_row=C2_START, max_row=C2_START+7)

s_imp = openpyxl.chart.Series(imp_data, title_from_data=True)
s_imp.graphicalProperties.solidFill = '4472C4'
s_exp = openpyxl.chart.Series(exp_data, title_from_data=True)
s_exp.graphicalProperties.solidFill = '70AD47'
chart2.series.append(s_imp)
chart2.series.append(s_exp)
chart2.set_categories(cats2)

# Add net line on secondary axis
line2 = LineChart()
net_data = Reference(ws2, min_col=4, min_row=C2_START, max_row=C2_START+7)
s_net = openpyxl.chart.Series(net_data, title_from_data=True)
s_net.graphicalProperties.line.solidFill = 'ED7D31'
s_net.graphicalProperties.line.width = 20000
s_net.smooth = False
line2.series.append(s_net)
line2.y_axis.axId = 200
line2.y_axis.crosses = 'max'

chart2 += line2
chart2.y_axis.title = 'GWh'

# ── Chart 3: Transit stacked area ────────────────────────────────────────────
chart3 = AreaChart()
chart3.grouping = 'stacked'
chart3.title    = 'Transit Through Georgia (2020–2023)'
chart3.y_axis.title = 'GWh'
chart3.x_axis.title = ''
chart3.legend.position = 'b'
chart3.style    = 10
chart3.width    = 16
chart3.height   = 12

cats3 = Reference(ws2, min_col=1, min_row=C3_START+1, max_row=C3_START+4)
transit_colors = ['FF4472C4', 'FFC00000', 'FFFFC000', 'FF70AD47']
for ci, (label, color) in enumerate([
    ('AZ→TR', 'FF4472C4'), ('RU→TR', 'FFC00000'),
    ('RU→ARM', 'FFFFC000'), ('ARM→TR', 'FF70AD47')
]):
    data_ref = Reference(ws2, min_col=2+ci, min_row=C3_START, max_row=C3_START+4)
    s = openpyxl.chart.Series(data_ref, title_from_data=True)
    s.graphicalProperties.solidFill = color[2:]
    s.graphicalProperties.line.solidFill = color[2:]
    chart3.series.append(s)
chart3.set_categories(cats3)

# ── Place charts on Charts tab ────────────────────────────────────────────────
ws2.add_chart(chart1, 'A32')
ws2.add_chart(chart2, 'J32')
ws2.add_chart(chart3, 'S32')

# ── Charts tab title ─────────────────────────────────────────────────────────
ws2.merge_cells('A28:AA28')
t = ws2['A28']
t.value = "Georgia — Cross-Border Electricity Trade Charts"
t.font  = Font(bold=True, size=14, color='FF1F4E79', name='Calibri')
t.alignment = Alignment(horizontal='center', vertical='center')

# subtitle
ws2.merge_cells('A29:AA29')
s = ws2['A29']
s.value = ("Left: seasonal imports by country (2020–2023)  ·  "
           "Centre: annual import vs export balance + net position (2018–2024)  ·  "
           "Right: transit by corridor (2020–2023)")
s.font = Font(italic=True, size=9, color='FF595959', name='Calibri')
s.alignment = Alignment(horizontal='center', vertical='center')
ws2.row_dimensions[28].height = 22
ws2.row_dimensions[29].height = 14

# ─────────────────────────────────────────────────────────────────────────────
wb.save(OUT)
print(f'Saved: {OUT}')
