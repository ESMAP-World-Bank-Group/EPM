"""
build_reference_lines_from_v7.py
================================
One-shot / idempotent converter: enrich the canonical reference_lines.csv from
the curated Excel workbook blacksea_crossborder_lines_v7.xlsx.

reference_lines.csv is the SOURCE OF TRUTH consumed by
pipelines/transmission_capacity.py. This script (re)generates it from v7 so the
analyst can keep editing the comfortable Excel and re-run to propagate, OR edit
reference_lines.csv directly afterwards.

What it does
------------
1. Reads the detailed "Internal lines" + "External lines" sheets of v7.
2. Parses the free-text "NTC / Capacity (MW)" into mw_fwd / mw_rev (best effort),
   ALWAYS keeping the verbatim text in `ntc_raw` and raising `parse_flag` when a
   human should double-check.
3. Maps v7 status {Existing, Committed, Candidate} -> the 3-value vocabulary EPM
   needs (existing / committed / candidate).
4. Assigns each endpoint to a model zone (from_zone / to_zone):
     - Türkiye      -> via data/substation_zone_map.csv (substring match)
     - Georgia/Armenia -> single zone (= country)
     - Azerbaijan   -> AzerbaijanMain ; label "Nakhchivan" -> Nakhchivan
     - external countries -> the external-node name (Bulgaria, Greece, ...)
5. Writes reference_lines.csv (backing up the previous file to *.bak) and prints
   a reconciliation diff against the previous version.

Usage
-----
    conda run -n gams_env python pre-analysis/build_reference_lines_from_v7.py \
        --v7 ../Data/blacksea_crossborder_lines_v7.xlsx

Add --dry-run to write to data/reference_lines.NEW.csv instead of overwriting.
"""
from __future__ import annotations

import argparse
import re
import sys
from pathlib import Path

import pandas as pd

_BASE = Path(__file__).resolve().parent
_DATA_DIR = _BASE.parent.parent / "Data"
_DEFAULT_V7 = _DATA_DIR / "blacksea_crossborder_lines_v7.xlsx"
_EPM_INCLUDED = _DATA_DIR / "blacksea_crossborder_lines_v8.xlsx"
_ZONE_MAP = _BASE / "data" / "substation_zone_map.csv"
_OUT = _BASE / "data" / "reference_lines.csv"

# Single master sheet that holds every line (internal + external).
_SHEET = "All"

# Capacity "menu" columns inserted right after the existing "Max Capacity (MW)".
CAP_MENU = ["NTC (MW)", "Commercial NTC (MW)"]

# EPM decision columns appended at the end of the "All" sheet (analyst-editable).
# Display names match the workbook as adjusted by the analyst; _load_sheet maps
# any EPM-ish header to a canonical name so build stays robust to renames.
EPM_HEADERS = ["EPM Type", "EPM zone from", "EPM zone to",
               "EPM capa fwd (MW)", "EPM capa rev (MW)", "EPM loss (%)",
               "EPM cap — basis"]
_DEFAULT_LOSS = "0.0198"   # default internal corridor loss factor (per-unit)

# Second-line explanations added under headers (matches v7's "(for EPM)" style).
_SUBTITLES = {
    "Max Capacity (MW)":    "(physical / thermal rating)",
    "NTC (MW)":             "(net transfer capacity — operational, N-1 secure)",
    "Commercial NTC (MW)":  "(firm allocated / auctioned, TEİAŞ)",
    "EPM capa fwd (MW)":    "(used by EPM — default = NTC)",
    "EPM capa rev (MW)":    "(used by EPM — default = NTC)",
    "EPM loss (%)":         "(corridor loss factor, per-unit)",
    "EPM cap — basis":      "(why this value)",
}


def _canon_epm(name: str) -> str:
    """Map an EPM-ish header (whatever the analyst named it) to a canonical key
    so the build reader is robust to renames like 'cap'→'capa', '(p.u.)'→'(%)'."""
    cl = str(name).lower()
    if not cl.startswith("epm"):
        return name
    if "basis" in cl:      return "EPM cap — basis"
    if "type" in cl:       return "EPM Type"
    if "zone from" in cl:  return "EPM zone from"
    if "zone to" in cl:    return "EPM zone to"
    if "loss" in cl:       return "EPM loss"
    if "fwd" in cl:        return "EPM cap fwd (MW)"
    if "rev" in cl:        return "EPM cap rev (MW)"
    return name


def _basis(status: str, typ: str, fwd: str, rev: str, src_used: str, doubt: bool) -> str:
    """Succinct justification of the EPM cap choice."""
    fwd, rev = _clean(fwd), _clean(rev)
    status, typ = (status or "").lower(), (typ or "").lower()
    if not fwd and not rev:
        return "to fill (no capacity available)"
    if fwd in ("0", "0.0") and rev in ("0", "0.0", ""):
        return "0 — closed border / no transfer"
    if status == "candidate":
        return "candidate → pNewTransmission"
    if any(k in typ for k in ("b2b", "back-to-back", "hvdc")):
        return "NTC = converter rating (HVDC B2B)"
    if src_used == "MAX":
        return "physical capacity (NTC unavailable — verify)"
    if src_used == "COMM":
        return "commercial NTC (operational NTC unavailable — verify)"
    if doubt:
        return "operational NTC (verify: seasonal/range text)"
    return "operational NTC"

# TEİAŞ "Mevcut Enterkonneksiyon Hatlarının NTK" announcement (12.01.2026), EK-1.
# Keyed by a distinctive substring of the v7 Route (Türkiye-side substation).
# value = (max_cap_thermal, comm_fwd, comm_rev, note).
#   max_cap → Max Capacity (MW) column (physical/converter ceiling)
#   comm    → Commercial NTC (MW) column (firm yearly NTK)
#   NTC itself comes from v7's own "NTC / Capacity (MW) (for EPM)" text.
_TEIAS: dict[str, tuple] = {
    "Hamitabat": (1350, 100, 100, "TEİAŞ EK-1 (12.01.2026): firm 100/100 yr (+monthly/daily auctions); 2×400kV AC"),
    "Babaeski":  (600, 50, 50,    "TEİAŞ EK-1 (12.01.2026): firm 50/50 yr; 400kV Babaeski-Nea Santa AC"),
    "Borçka":    (700, 400, 0,    "TEİAŞ EK-1 (12.01.2026): Jan-26 firm export 400 / import 0; HVDC B2B 2×350=700"),
    "Van (B2B":  (600, 0, 0,      "TEİAŞ EK-1 (12.01.2026): firm NTC bilateral-TBD; HVDC B2B 1×600"),
    "Silopi":    (600, 600, 0,    "TEİAŞ EK-1 (12.01.2026, Cizre-Kasek 400kV): firm export up to 600 (isolated-feed)"),
    "Birecik":   (300, 300, 0,    "TEİAŞ EK-1 (12.01.2026, Reyhanlı/Elbeyli 66kV): firm export up to ~300 combined"),
    "Sederek":   (150, 0, 0,      "TEİAŞ EK-1 (12.01.2026, Iğdır-Arpaçay/Aralık-Babek 154kV): firm 0, isolated-feed TBD"),
    "Kars":      (0, 0, 0,        "TEİAŞ EK-1 (12.01.2026, Kars-Gumri 220kV): 0 — missing 154/220kV transformer"),
}


def _teias_match(route) -> tuple | None:
    r = str(route).lower()
    for key, vals in _TEIAS.items():
        if key.lower() in r:
            return vals
    return None


def _capstr(fwd: str, rev: str) -> str:
    """Compact capacity cell: 'X' if symmetric, 'fwd/rev' otherwise, '' if empty."""
    fwd, rev = _clean(fwd), _clean(rev)
    if not fwd and not rev:
        return ""
    return fwd if fwd == rev else f"{fwd}/{rev}"


_TEIAS_TAG = ("Specs (kV, length), Max Capacity & Commercial NTC: "
              "TEİAŞ EK-1 (12.01.2026)")


def _tag_source(base: str, is_teias: bool) -> str:
    """Field-tagged provenance for the Source column (what source documents what).
    `base` is the original v7 source list (operational NTC, status, project...)."""
    base = _clean(base)
    if is_teias:
        return f"{_TEIAS_TAG}. NTC, status & project: {base}" if base else _TEIAS_TAG
    return f"NTC, specs, status & project: {base}" if base else ""

# ── Country metadata ──────────────────────────────────────────────────────────
# label (as it appears in v7 From/To country) -> (iso3, default_zone, internal?)
# default_zone == None means "resolve from substation map" (Türkiye, multi-zone).
_COUNTRY: dict[str, tuple[str, str | None, bool]] = {
    "Türkiye":         ("TUR", None,             True),
    "Turkiye":         ("TUR", None,             True),
    "Georgia":         ("GEO", "Georgia",        True),
    "Armenia":         ("ARM", "Armenia",        True),
    "Azerbaijan":      ("AZE", "AzerbaijanMain", True),
    "Azerbaijan main": ("AZE", "AzerbaijanMain", True),
    "Nakhchivan":      ("AZE", "Nakhchivan",     True),
    # External neighbours (zext) — zone == external-node name
    "Bulgaria": ("BGR", "Bulgaria", False),
    "Romania":  ("ROU", "Romania",  False),
    "Greece":   ("GRC", "Greece",   False),
    "Russia":   ("RUS", "Russia",   False),
    "Iran":     ("IRN", "Iran",     False),
    "Iraq":     ("IRQ", "Iraq",     False),
    "Syria":    ("SYR", "Syria",    False),
    "Serbia":   ("SRB", "Serbia",   False),
    "Hungary":  ("HUN", "Hungary",  False),
    "Moldova":  ("MDA", "Moldova",  False),
    "Ukraine":  ("UKR", "Ukraine",  False),
    "Kazakhstan": ("KAZ", "Kazakhstan", False),
}

_STATUS = {"existing": "existing", "committed": "committed", "candidate": "candidate"}


def _load_sheet(xl: pd.ExcelFile, sheet: str = _SHEET) -> pd.DataFrame:
    raw = xl.parse(sheet, header=None)
    hrow = next(
        i for i in range(min(10, len(raw)))
        if raw.iloc[i].astype(str).str.contains("From country").any()
    )
    df = raw.iloc[hrow + 1:].copy()
    # header may carry a "\n(subtitle)" second line — key on the first line only
    df.columns = [str(c).split("\n")[0].strip() for c in raw.iloc[hrow].tolist()]
    df = df.loc[:, ~df.columns.duplicated()]          # keep first "Border" (= scope flag)
    df = df[df["From country"].notna()].reset_index(drop=True)
    # Normalise v7's source NTC text column (carries "Capacity"; not the inserted
    # "NTC (MW)" menu column) to a stable name "NTC".
    ntc = next((c for c in df.columns
                if str(c).startswith("NTC") and "Capacity" in str(c)), None)
    if ntc and ntc != "NTC":
        df = df.rename(columns={ntc: "NTC"})
    # robust to analyst renames of the EPM block (cap↔capa, p.u.↔%, …)
    df.columns = [_canon_epm(c) for c in df.columns]
    return df


def _split_route(route: str) -> tuple[str, str]:
    """First segment -> from_substation, last segment -> to_substation."""
    parts = [p.strip() for p in re.split(r"→|->", str(route)) if p.strip()]
    if not parts:
        return "", ""
    return parts[0], parts[-1]


def _load_zone_map() -> list[tuple[str, str, str]]:
    zm = pd.read_csv(_ZONE_MAP, comment="#")
    return [(r.country.strip(), r.match_key.strip(), r.zone.strip())
            for r in zm.itertuples()]


def _resolve_zone(label: str, substation: str, zone_map) -> tuple[str, str, bool]:
    iso, default_zone, internal = _COUNTRY.get(label, ("???", label, False))
    if default_zone is not None:
        return iso, default_zone, internal
    # Türkiye: substring match on substation
    for country, key, zone in zone_map:
        if country == iso and key.lower() in substation.lower():
            return iso, zone, internal
    return iso, "UNMAPPED", internal


# ── NTC text parser ─────────────────────────────────────────────────────────--
_NUM = r"\d[\d,]*\.?\d*"


def _to_float(tok: str) -> float:
    return float(tok.replace(",", ""))


def _last_num_before(text: str, keyword: str) -> str:
    """Largest/last number appearing just before `keyword` (handles ranges -> upper)."""
    m = re.search(rf"({_NUM}(?:\s*[-–]\s*{_NUM})?)\s*(?:mw\s*)?{keyword}", text.lower())
    if not m:
        return ""
    nums = re.findall(_NUM, m.group(1))
    return str(int(_to_float(nums[-1]))) if nums else ""


def parse_ntc(text: str) -> tuple[str, str, str]:
    """Return (mw_fwd, mw_rev, parse_flag). fwd = from_country -> to_country."""
    t = str(text).strip()
    low = t.lower()
    nums = re.findall(_NUM, t)

    # No digits at all -> nothing to parse (TBD, —, On hold, Limited, Combined...)
    if not nums:
        return "", "", "manual: no numeric NTC"

    # Directional "X import / Y export" -> export = fwd (from→to), import = rev
    if "import" in low or "export" in low:
        fwd = _last_num_before(t, "export")
        rev = _last_num_before(t, "import")
        return fwd, rev, "review: import/export parsed from text"

    flag = "review: value given in MVA" if "mva" in low else ""

    # Directional slash "1400 BG→GR / 1700 GR→BG" or "570 / 650"
    if "/" in t and len(nums) >= 2:
        return str(int(_to_float(nums[0]))), str(int(_to_float(nums[1]))), flag

    # Range "250-700" -> upper bound (symmetric), flagged
    if re.search(rf"{_NUM}\s*[-–]\s*{_NUM}", t) and len(nums) >= 2:
        v = str(int(_to_float(nums[1])))
        return v, v, "review: range, took upper bound"

    if len(nums) == 1:
        v = str(int(_to_float(nums[0])))
        return v, v, flag

    v = str(int(_to_float(nums[0])))
    return v, v, flag or "review: multiple numbers, took first"


def _clean(v) -> str:
    """Stringify a cell, dropping NaN/None and trailing '.0' on whole numbers."""
    if v is None:
        return ""
    s = str(v).strip()
    if s.lower() in ("nan", "none", "nat"):
        return ""
    if re.fullmatch(r"-?\d+\.0", s):
        s = s[:-2]
    return s


def build(v7_path: Path) -> pd.DataFrame:
    xl = pd.ExcelFile(v7_path)
    zone_map = _load_zone_map()
    df = _load_sheet(xl, _SHEET)
    has_epm = "EPM zone from" in df.columns
    rows: list[dict] = []
    for r in df.itertuples():
        d = df.loc[r.Index]
        from_sub, to_sub = _split_route(d["Route"])
        f_iso, f_zone_auto, f_int = _resolve_zone(str(d["From country"]).strip(), from_sub, zone_map)
        t_iso, t_zone_auto, t_int = _resolve_zone(str(d["To country"]).strip(), to_sub, zone_map)
        status = _STATUS.get(str(d["Status"]).strip().lower(), str(d["Status"]).strip().lower())
        # Operational NTC from the clean "NTC (MW)" column (the raw v7 text column
        # was removed by the analyst). EPM capa columns override below.
        ntc_op = _clean(d.get("NTC (MW)"))
        nf, nr = (ntc_op.split("/", 1)[0].strip(), ntc_op.split("/", 1)[1].strip()) \
            if "/" in ntc_op else (ntc_op, ntc_op)
        mw_fwd, mw_rev, flag = nf, nr, ""

        # EPM scope = both endpoints are internal EPM countries (TUR/GEO/ARM/AZE).
        # NB: differs from v7's own Internal/External flag, which counts BG/RO as
        # internal — here they are external until added to zcmap.
        scope = "internal" if (f_int and t_int) else "external"
        loss = _DEFAULT_LOSS if scope == "internal" else ""

        # EPM columns (if present and filled) are AUTHORITATIVE — they override
        # the auto-derived type / zones / capacities / loss and clear parse_flag.
        if has_epm and _clean(d.get("EPM zone from")):
            scope = _clean(d.get("EPM Type")).lower() or scope
            f_zone = _clean(d.get("EPM zone from"))
            t_zone = _clean(d.get("EPM zone to")) or t_zone_auto
            ef, er = _clean(d.get("EPM cap fwd (MW)")), _clean(d.get("EPM cap rev (MW)"))
            el = _clean(d.get("EPM loss"))
            if ef or er:
                mw_fwd, mw_rev, flag = ef, er, ""
            if el:
                loss = el
        else:
            f_zone, t_zone = f_zone_auto, t_zone_auto

        rows.append({
            "from_country": f_iso, "from_substation": from_sub, "from_zone": f_zone,
            "to_country": t_iso, "to_substation": to_sub, "to_zone": t_zone,
            "scope": scope,
            "voltage_kv": _clean(d.get("Voltage (kV)")),
            "n_circuits": _clean(d.get("Circuits")),
            "length_km": _clean(d.get("Length (km)")),
            "mw_fwd": mw_fwd, "mw_rev": mw_rev, "loss_pu": loss,
            "max_cap": _clean(d.get("Max Capacity (MW)")),
            "ntc_op": ntc_op,
            "ntc_comm": _clean(d.get("Commercial NTC (MW)")),
            "epm_cap_basis": _clean(d.get("EPM cap — basis")),
            "status": status,
            "earliest_entry": _clean(d.get("COD")),
            "project": _clean(d.get("Project")),
            "parse_flag": flag,
            "note": _clean(d.get("Notes")),
            "source": _clean(d.get("Source")),
        })
    return pd.DataFrame(rows)


# ── Annotation: write EPM columns back into a copy of the v7 workbook ──────────

def _find_header(ws):
    """Return (header_row_1based, {name: col_index}); first occurrence wins."""
    for rr in range(1, 11):
        names: dict[str, int] = {}
        found = False
        for cc in range(1, ws.max_column + 1):
            v = ws.cell(rr, cc).value
            if isinstance(v, str):
                key = v.split("\n")[0].strip()     # ignore "\n(subtitle)" second line
                names.setdefault(key, cc)          # keep FIRST "Border" (= scope flag)
                if key == "From country":
                    found = True
        if found:
            return rr, names
    return None, {}


def _col_ntc(col: dict) -> int:
    """Index of v7's source NTC text column (the one carrying 'Capacity', not the
    inserted 'NTC (MW)' menu column)."""
    for name, idx in col.items():
        if str(name).startswith("NTC") and "Capacity" in str(name):
            return idx
    return col.get("NTC / Capacity (MW)")


def _read_prior_epm(path: Path) -> dict:
    """Read previously-saved EPM cells from the annotated 'All' sheet, keyed by line."""
    from openpyxl import load_workbook
    wb = load_workbook(path, data_only=True)
    if _SHEET not in wb.sheetnames:
        return {}
    ws = wb[_SHEET]
    hrow, col = _find_header(ws)
    if hrow is None or "EPM zone from" not in col:
        return {}
    prior: dict[tuple, list] = {}
    r = hrow + 1
    while r <= ws.max_row:
        fc = ws.cell(r, col["From country"]).value
        if fc is None or str(fc).strip() == "":
            break
        key = (str(fc).strip(), _clean(ws.cell(r, col["To country"]).value),
               _clean(ws.cell(r, col["Route"]).value))
        # tolerate schema changes (an older workbook may have fewer EPM columns)
        prior[key] = [_clean(ws.cell(r, col[h]).value) if h in col else ""
                      for h in EPM_HEADERS]
        r += 1
    return prior


def annotate(src_v7: Path, out_path: Path) -> None:
    """Append/refresh the EPM columns (coloured) on the master 'All' sheet of v7."""
    from openpyxl import load_workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    zone_map = _load_zone_map()
    prior = _read_prior_epm(out_path) if out_path.exists() else {}
    # Always annotate from the pristine source, carrying over prior EPM edits.
    wb = load_workbook(src_v7)
    ws = wb[_SHEET]
    hdr_fill = PatternFill("solid", fgColor="BDD7EE")   # light blue
    chk_fill = PatternFill("solid", fgColor="FFC7CE")   # light red (needs review)
    wrap = Alignment(wrap_text=True, vertical="center", horizontal="center")
    n_check = 0

    def _hdr(rr, cc, name):
        c = ws.cell(rr, cc, f"{name}\n{_SUBTITLES[name]}" if name in _SUBTITLES else name)
        c.fill, c.font, c.alignment = hdr_fill, Font(bold=True), wrap
        return c

    hrow, col = _find_header(ws)

    # ── Insert the capacity-menu columns right after "Max Capacity (MW)" ───────
    maxcap_col = col.get("Max Capacity (MW)")
    if maxcap_col and not all(h in col for h in CAP_MENU):
        ws.insert_cols(maxcap_col + 1, len(CAP_MENU))
    hrow, col = _find_header(ws)                  # indices shifted by the insert
    maxcap_col = col.get("Max Capacity (MW)")
    for j, h in enumerate(CAP_MENU):
        _hdr(hrow, maxcap_col + 1 + j, h)
    if "Max Capacity (MW)" in col:               # add subtitle to existing header
        _hdr(hrow, maxcap_col, "Max Capacity (MW)")
    hrow, col = _find_header(ws)
    ntc_col = _col_ntc(col)
    cap_ntc_col = col["NTC (MW)"]
    cap_comm_col = col["Commercial NTC (MW)"]
    src_col = col.get("Source")
    notes_col = col.get("Notes")
    type_col = col.get("Type")

    # ── EPM decision block at the very end ────────────────────────────────────
    start = ws.max_column + 1
    for i, h in enumerate(EPM_HEADERS):
        _hdr(hrow, start + i, h)

    r = hrow + 1
    while r <= ws.max_row:
        fc = ws.cell(r, col["From country"]).value
        if fc is None or str(fc).strip() == "":
            break
        route = ws.cell(r, col["Route"]).value
        tc = ws.cell(r, col["To country"]).value
        ntc = ws.cell(r, ntc_col).value
        from_sub, to_sub = _split_route(route)
        f_iso, f_zone, f_int = _resolve_zone(str(fc).strip(), from_sub, zone_map)
        t_iso, t_zone, t_int = _resolve_zone(str(tc).strip(), to_sub, zone_map)
        scope = "internal" if (f_int and t_int) else "external"
        # NTC (operational) parsed from v7's own NTC text
        mw_fwd, mw_rev, flag = parse_ntc(ntc)

        # Capacity menu: thermal Max Capacity + firm Commercial NTC from TEİAŞ
        teias = _teias_match(route)
        base_src = _clean(ws.cell(r, src_col).value) if src_col else ""
        max_cap = cf = cr = ""
        if teias:
            max_cap, cf, cr, tnote = teias
            ws.cell(r, maxcap_col, max_cap)
            ws.cell(r, cap_comm_col, _capstr(str(cf), str(cr)))
            if notes_col:
                cur = _clean(ws.cell(r, notes_col).value)
                if tnote not in cur:
                    ws.cell(r, notes_col, (cur + " | " + tnote) if cur else tnote)
        if src_col:
            ws.cell(r, src_col, _tag_source(base_src, bool(teias)))
        ws.cell(r, cap_ntc_col, _capstr(mw_fwd, mw_rev))   # operational NTC ref

        # EPM cap default: NTC → fall back to Max Capacity → Commercial NTC
        ef, er, src_used = mw_fwd, mw_rev, "NTC"
        if not _clean(ef) and not _clean(er):
            if teias and _clean(str(max_cap)):
                ef = er = str(max_cap); src_used = "MAX"
            elif teias and (_clean(str(cf)) or _clean(str(cr))):
                ef, er, src_used = str(cf), str(cr), "COMM"
            else:
                src_used = ""
        cap_doubt = bool(flag) or src_used in ("MAX", "COMM")
        typ = _clean(ws.cell(r, type_col).value) if type_col else ""
        basis = _basis(_clean(ws.cell(r, col["Status"]).value), typ, ef, er, src_used, cap_doubt)
        loss = _DEFAULT_LOSS if scope == "internal" else ""

        auto = [scope, f_zone, t_zone, ef, er, loss, basis]
        doubt = [False, f_zone == "UNMAPPED", t_zone == "UNMAPPED",
                 cap_doubt, cap_doubt, False, False]

        key = (str(fc).strip(), _clean(tc), _clean(route))
        pri = prior.get(key, [""] * len(EPM_HEADERS))
        for i in range(len(EPM_HEADERS)):
            edited = pri[i] not in ("", None)
            val = pri[i] if edited else auto[i]   # pre-fill even doubtful cells
            c = ws.cell(r, start + i, val)
            if doubt[i] and not edited:
                c.fill = chk_fill                 # keep RED = please verify
                n_check += 1
        r += 1

    wb.save(out_path)
    print(f"Wrote {out_path}")
    print(f"  Capacity menu inserted after 'Max Capacity (MW)': {', '.join(CAP_MENU)}.")
    print(f"  EPM block appended ({len(EPM_HEADERS)} cols); EPM cap pre-filled (default NTC), RED = verify.")
    print(f"  {n_check} EPM cap cell(s) highlighted RED.")
    if prior:
        print(f"  Preserved your prior EPM edits from the existing workbook.")


def reconcile_report(new: pd.DataFrame, old_path: Path) -> None:
    print("\n" + "=" * 78)
    print("RECONCILIATION REPORT")
    print("=" * 78)
    print(f"v7 lines parsed : {len(new)}  ({(new.scope=='internal').sum()} internal, "
          f"{(new.scope=='external').sum()} external)")
    print(f"by status       : " + ", ".join(f"{k}={v}" for k, v in new.status.value_counts().items()))
    unmapped = new[(new.from_zone == "UNMAPPED") | (new.to_zone == "UNMAPPED")]
    if len(unmapped):
        print(f"\n[!] {len(unmapped)} line(s) with UNMAPPED zone — add them to substation_zone_map.csv:")
        for r in unmapped.itertuples():
            print(f"     {r.from_country}:{r.from_substation} → {r.to_country}:{r.to_substation}")
    flagged = new[new.parse_flag.str.startswith(("manual", "review"))]
    if len(flagged):
        print(f"\n[!] {len(flagged)} capacity value(s) need review:")
        for r in flagged.itertuples():
            print(f"     {r.from_zone}→{r.to_zone:<16} raw='{r.ntc_raw}'  -> fwd={r.mw_fwd} rev={r.mw_rev}  ({r.parse_flag})")
    if old_path.exists():
        print(f"\nPrevious reference_lines.csv backed up alongside (*.bak).")
    print("=" * 78 + "\n")


def main() -> int:
    ap = argparse.ArgumentParser(
        description="v7 ⇄ reference_lines.csv. Default: build reference_lines.csv "
                    "from the EPM-annotated workbook (EPM columns authoritative)."
    )
    ap.add_argument("--annotate", action="store_true",
                    help="append/refresh EPM columns on a copy of v7 "
                         "(blacksea_crossborder_lines_v7_epm_included.xlsx) for editing")
    ap.add_argument("--src", type=Path, default=_DEFAULT_V7,
                    help="source v7 workbook used by --annotate")
    ap.add_argument("--v7", type=Path, default=None,
                    help="workbook to build from (default: epm_included if present, else v7)")
    ap.add_argument("--dry-run", action="store_true",
                    help="build: write reference_lines.NEW.csv instead of overwriting")
    args = ap.parse_args()

    if args.annotate:
        if not args.src.exists():
            print(f"ERROR: source workbook not found: {args.src}", file=sys.stderr)
            return 1
        annotate(args.src, _EPM_INCLUDED)
        return 0

    # build mode — prefer the EPM-annotated workbook when it exists
    src = args.v7 or (_EPM_INCLUDED if _EPM_INCLUDED.exists() else _DEFAULT_V7)
    if not src.exists():
        print(f"ERROR: workbook not found: {src}", file=sys.stderr)
        return 1
    print(f"Building from: {src.name}")

    new = build(src)
    out = _OUT.with_suffix(".NEW.csv") if args.dry_run else _OUT
    if not args.dry_run and _OUT.exists():
        _OUT.replace(_OUT.with_suffix(".csv.bak"))
    header = (
        "# Cross-border transmission reference data — Black Sea region (SOURCE OF TRUTH)\n"
        "# Generated by build_reference_lines_from_v7.py from the EPM-annotated v7 workbook.\n"
        "# EPM columns (zones / cap fwd-rev / loss) in v7 are authoritative; auto-parse is fallback.\n"
        "# mw_fwd = from_zone→to_zone ; mw_rev = reverse. loss_pu feeds pLossFactorInternal (internal only).\n"
        "# status: existing | committed | candidate  (existing+committed→pTransferLimit, candidate→pNewTransmission)\n"
        "# parse_flag flags rows still inferred from free text — fill the EPM cols in v7 to clear them.\n"
    )
    with open(out, "w", encoding="utf-8", newline="") as fh:
        fh.write(header)
        new.to_csv(fh, index=False)
    print(f"Wrote {out}  ({len(new)} rows)")
    reconcile_report(new, _OUT.with_suffix(".csv.bak"))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
